#!/usr/bin/env python3
"""
update_partner.py – Mise à jour ciblée d'un partenaire UpJunoo
================================================================
Workflow :
1. Connexion manuelle + filtre sur /manage-owners.
2. Le script charge la liste complète des partenaires.
3. Boucle : tu saisis un nom → le script extrait tout (conducteurs + flotte + véhicules).
4. Export par partenaire : output/partners/<nom>/data.json, data.html, data.xlsx
5. Mise à jour des fichiers globaux : output/partners.json, partners.html, partners.xlsx
"""

import asyncio
import json
import time
import traceback
import html as html_module
from datetime import datetime
from pathlib import Path

import aiohttp
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
)
from webdriver_manager.chrome import ChromeDriverManager

# ─── Configuration ──────────────────────────────────────────────────────────────
BASE_URL = "https://upjunoo-server-new.junooapps.com"
OUTPUT_DIR = Path(__file__).parent / "output"
PARTNERS_DIR = OUTPUT_DIR / "partners"
GLOBAL_JSON = OUTPUT_DIR / "partners.json"
GLOBAL_HTML = OUTPUT_DIR / "partners.html"
GLOBAL_XLSX = OUTPUT_DIR / "partners.xlsx"

# ─── Excel Styling ──────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, cols):
    for col_idx, title in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ═══════════════════════════════════════════════════════════════════════════════
#  VEHICLE ENRICHMENT (via HTTP — aiohttp)
# ═══════════════════════════════════════════════════════════════════════════════

def find_key(obj, key):
    """Recherche récursive d'une clé dans un dict/list."""
    if isinstance(obj, dict):
        if key in obj and obj[key] not in [None, "", "null"]:
            return obj[key]
        for v in obj.values():
            res = find_key(v, key)
            if res is not None:
                return res
    elif isinstance(obj, list):
        for item in obj:
            res = find_key(item, key)
            if res is not None:
                return res
    return None


async def _enrich_drivers_vehicles(drivers, cookies_dict):
    """Récupère les infos véhicule de chaque conducteur via HTTP."""
    enriched = 0
    errors = 0
    semaphore = asyncio.Semaphore(10)

    async def fetch_one(session, drv):
        nonlocal enriched, errors
        url = drv.get("view_profile", "")
        if not url or url == "N/A":
            return
        async with semaphore:
            try:
                async with session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                    if resp.status in (401, 403) or '/login' in str(resp.url):
                        print("    🔒 Session expirée – enrichissement interrompu.")
                        return
                    if resp.status != 200:
                        errors += 1
                        return
                    html_text = await resp.text()
                    soup = BeautifulSoup(html_text, 'html.parser')
                    app_div = soup.find('div', id='app')
                    if not app_div or not app_div.get('data-page'):
                        return
                    data_page = json.loads(app_div['data-page'])
                    vi = {"type": "N/A", "marque": "N/A", "modele": "N/A", "matricule": "N/A"}
                    v = find_key(data_page, 'vehicle_type_name')
                    if v: vi["type"] = str(v).upper()
                    v = find_key(data_page, 'car_make_name')
                    if v: vi["marque"] = str(v).upper()
                    v = find_key(data_page, 'car_model_name')
                    if v: vi["modele"] = str(v).upper()
                    v = find_key(data_page, 'car_number')
                    if v: vi["matricule"] = str(v).upper()
                    drv["vehicle"] = vi
                    enriched += 1
                    print(f"      ✅ {drv.get('nom','?')} → {vi['type']} | {vi['matricule']}")
            except asyncio.TimeoutError:
                errors += 1
            except Exception as e:
                errors += 1
                print(f"      ⚠️ {drv.get('nom','?')}: {e}")

    to_process = [d for d in drivers if d.get("view_profile", "N/A") != "N/A"]
    if not to_process:
        print("    ℹ️ Aucun lien profil conducteur — enrichissement véhicule ignoré.")
        return

    print(f"    🔄 Enrichissement véhicules ({len(to_process)} conducteurs)…")
    async with aiohttp.ClientSession(cookies=cookies_dict) as session:
        for i in range(0, len(to_process), 20):
            batch = to_process[i:i + 20]
            await asyncio.gather(*(fetch_one(session, d) for d in batch))
            done = min(i + 20, len(to_process))
            print(f"    ⏳ {done}/{len(to_process)} traités…")

    print(f"    ✅ {enriched}/{len(to_process)} véhicules enrichis ({errors} erreurs).")


def enrich_drivers_vehicles(drivers, selenium_driver):
    """Point d'entrée synchrone pour l'enrichissement async."""
    cookies = {c['name']: c['value'] for c in selenium_driver.get_cookies()}
    asyncio.run(_enrich_drivers_vehicles(drivers, cookies))


# ═══════════════════════════════════════════════════════════════════════════════
#  STATS
# ═══════════════════════════════════════════════════════════════════════════════

def compute_stats(partner):
    """Calcule les statistiques pour un partenaire."""
    drivers = partner.get("drivers", [])
    fleet = partner.get("fleet", [])
    nb_approuves = sum(
        1 for d in drivers
        if d.get("statut_approuve", "").strip().upper() in ["APPROUVÉ", "APPROVED", "ACTIF"]
    )
    return {
        "nb_conducteurs": len(drivers),
        "nb_conducteurs_approuves": nb_approuves,
        "nb_conducteurs_desapprouves": len(drivers) - nb_approuves,
        "nb_flottes": len(fleet),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPORT INDIVIDUEL
# ═══════════════════════════════════════════════════════════════════════════════

def export_partner_json(partner, partner_dir):
    """Export JSON individuel."""
    path = partner_dir / "data.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(partner, f, ensure_ascii=False, indent=2)
    print(f"    📄 JSON : {path}")


def export_partner_html(partner, partner_dir):
    """Export HTML individuel (style moderne)."""
    path = partner_dir / "data.html"
    nom = html_module.escape(partner.get("nom", "N/A"))
    email = html_module.escape(partner.get("email", "N/A"))
    tel = html_module.escape(partner.get("telephone", "N/A"))
    owner_id = html_module.escape(partner.get("owner_id", "N/A"))
    doc_url = partner.get("document_url", "#")
    stats = partner.get("stats", {})
    drivers = partner.get("drivers", [])
    fleet = partner.get("fleet", [])
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Driver rows
    driver_rows = ""
    for d in drivers:
        v = d.get("vehicle", {})
        if v:
            v_info = (
                f'<div class="vehicle-info">'
                f'<strong>Type:</strong> {html_module.escape(v.get("type", "N/A"))}<br>'
                f'<strong>Marque:</strong> {html_module.escape(v.get("marque", "N/A"))}<br>'
                f'<strong>Modèle:</strong> {html_module.escape(v.get("modele", "N/A"))}<br>'
                f'<strong>Matricule:</strong> {html_module.escape(v.get("matricule", "N/A"))}'
                f'</div>'
            )
        else:
            v_info = html_module.escape(d.get("type_vehicule", "N/A"))
        ds = d.get("statut_approuve", "N/A")
        badge_cls = "approved" if ds.strip().upper() in ["APPROUVÉ", "APPROVED", "ACTIF"] else "disapproved"
        driver_rows += f"""
            <tr>
                <td>{html_module.escape(d.get('nom', ''))}</td>
                <td>{html_module.escape(d.get('emplacement', ''))}</td>
                <td>{html_module.escape(d.get('telephone', ''))}</td>
                <td>{html_module.escape(d.get('type_transport', ''))}</td>
                <td><span class="badge {badge_cls}">{html_module.escape(ds)}</span></td>
                <td>{v_info}</td>
            </tr>"""

    # Fleet rows
    fleet_rows = ""
    for v in fleet:
        fleet_rows += f"""
            <tr>
                <td>{html_module.escape(v.get('type_vehicule', ''))}</td>
                <td>{html_module.escape(v.get('marque', ''))}</td>
                <td>{html_module.escape(v.get('modele', ''))}</td>
                <td>{html_module.escape(v.get('plaque', ''))}</td>
                <td>{html_module.escape(v.get('statut', ''))}</td>
            </tr>"""

    drivers_table = (
        '<table><thead><tr><th>Nom</th><th>Emplacement</th><th>Téléphone</th>'
        '<th>Type Transport</th><th>Statut</th><th>Véhicule</th></tr></thead><tbody>'
        + driver_rows + '</tbody></table>'
    ) if drivers else '<div class="empty">Aucun conducteur trouvé.</div>'

    fleet_table = (
        '<table><thead><tr><th>Type Véhicule</th><th>Marque</th><th>Modèle</th>'
        '<th>Plaque</th><th>Statut</th></tr></thead><tbody>'
        + fleet_rows + '</tbody></table>'
    ) if fleet else '<div class="empty">Aucun véhicule trouvé.</div>'

    html_content = f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{nom} - Rapport Partenaire</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            margin: 0;
            padding: 20px;
            color: #333;
        }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        .section {{
            background: white;
            padding: 30px;
            margin: 20px 0;
            border-radius: 10px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
        }}
        h1 {{ color: #667eea; margin-bottom: 10px; font-size: 2.2em; text-align: center; }}
        .meta {{ text-align: center; color: #999; font-size: 0.85em; margin-bottom: 20px; }}
        h2 {{
            color: #555;
            margin-top: 30px;
            margin-bottom: 20px;
            border-bottom: 2px solid #667eea;
            padding-bottom: 10px;
        }}
        .partner-info {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .info-card {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            border-left: 4px solid #667eea;
        }}
        .info-label {{ font-weight: bold; color: #667eea; margin-bottom: 5px; }}
        .info-value {{ color: #333; word-break: break-all; }}
        .stats {{
            display: flex;
            justify-content: center;
            gap: 20px;
            margin: 20px 0;
            flex-wrap: wrap;
        }}
        .stat {{
            text-align: center;
            color: white;
            padding: 15px 25px;
            border-radius: 8px;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            min-width: 120px;
        }}
        .stat-number {{ font-size: 2em; font-weight: bold; display: block; }}
        .stat-label {{ font-size: 0.9em; opacity: 0.9; }}
        .stat.green {{ background: #27ae60; }}
        .stat.red {{ background: #e74c3c; }}
        .stat.blue {{ background: #2980b9; }}
        .stat.purple {{ background: #667eea; }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        }}
        th, td {{ padding: 12px 15px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            font-weight: bold;
            font-size: 0.9em;
        }}
        tr:nth-child(even) {{ background: #f8f9fa; }}
        tr:hover {{ background: #e3f2fd; transition: background 0.3s; }}
        .vehicle-info {{ font-size: 0.85em; line-height: 1.4; }}
        .badge {{
            display: inline-block;
            padding: 3px 10px;
            border-radius: 12px;
            font-size: 0.8em;
            font-weight: 600;
        }}
        .badge.approved {{ background: #d5f5e3; color: #1e8449; }}
        .badge.disapproved {{ background: #fadbd8; color: #c0392b; }}
        a {{ color: #667eea; text-decoration: none; font-weight: bold; }}
        a:hover {{ text-decoration: underline; }}
        .empty {{ text-align: center; color: #999; font-style: italic; padding: 40px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="section">
            <h1>{nom}</h1>
            <div class="meta">Mis à jour le {now}</div>
            <div class="stats">
                <div class="stat green">
                    <span class="stat-number">{stats.get('nb_conducteurs_approuves', 0)}</span>
                    <span class="stat-label">Approuvés</span>
                </div>
                <div class="stat red">
                    <span class="stat-number">{stats.get('nb_conducteurs_desapprouves', 0)}</span>
                    <span class="stat-label">Désapprouvés</span>
                </div>
                <div class="stat blue">
                    <span class="stat-number">{stats.get('nb_flottes', 0)}</span>
                    <span class="stat-label">Véhicules</span>
                </div>
                <div class="stat purple">
                    <span class="stat-number">{stats.get('nb_conducteurs', 0)}</span>
                    <span class="stat-label">Conducteurs</span>
                </div>
            </div>
            <div class="partner-info">
                <div class="info-card">
                    <div class="info-label">Email</div>
                    <div class="info-value">{email}</div>
                </div>
                <div class="info-card">
                    <div class="info-label">Téléphone</div>
                    <div class="info-value">{tel}</div>
                </div>
                <div class="info-card">
                    <div class="info-label">Owner ID</div>
                    <div class="info-value">{owner_id}</div>
                </div>
                <div class="info-card">
                    <div class="info-label">Document</div>
                    <div class="info-value"><a href="{html_module.escape(doc_url)}" target="_blank">Voir Document</a></div>
                </div>
            </div>
        </div>
        <div class="section">
            <h2>Conducteurs ({len(drivers)})</h2>
            {drivers_table}
        </div>
        <div class="section">
            <h2>Flotte ({len(fleet)})</h2>
            {fleet_table}
        </div>
    </div>
</body>
</html>"""

    with open(path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"    🌐 HTML : {path}")


def export_partner_xlsx(partner, partner_dir):
    """Export Excel individuel (3 onglets)."""
    path = partner_dir / "data.xlsx"
    drivers = partner.get("drivers", [])
    fleet = partner.get("fleet", [])
    stats = partner.get("stats", {})

    wb = Workbook()

    # Sheet 1: Résumé
    ws1 = wb.active
    ws1.title = "Résumé"
    style_header(ws1, ["Champ", "Valeur"])
    fields = [
        ("Nom", partner.get("nom", "")),
        ("Email", partner.get("email", "")),
        ("Téléphone", partner.get("telephone", "")),
        ("Owner ID", partner.get("owner_id", "")),
        ("Conducteurs Approuvés", stats.get("nb_conducteurs_approuves", 0)),
        ("Conducteurs Désapprouvés", stats.get("nb_conducteurs_desapprouves", 0)),
        ("Véhicules (Flotte)", stats.get("nb_flottes", 0)),
        ("Total Conducteurs", stats.get("nb_conducteurs", 0)),
        ("Mis à jour", partner.get("updated_at", "")),
    ]
    for field, val in fields:
        ws1.append([field, val])
    auto_width(ws1)

    # Sheet 2: Conducteurs
    ws2 = wb.create_sheet("Conducteurs")
    cols2 = ["Nom", "Emplacement", "Téléphone", "Type Transport", "Statut",
             "Type Véhicule", "Marque", "Modèle", "Matricule"]
    style_header(ws2, cols2)
    for d in drivers:
        v = d.get("vehicle", {})
        ws2.append([
            d.get("nom", ""),
            d.get("emplacement", ""),
            d.get("telephone", ""),
            d.get("type_transport", ""),
            d.get("statut_approuve", ""),
            v.get("type", d.get("type_vehicule", "")),
            v.get("marque", ""),
            v.get("modele", ""),
            v.get("matricule", ""),
        ])
    auto_width(ws2)
    if ws2.dimensions:
        ws2.auto_filter.ref = ws2.dimensions

    # Sheet 3: Flotte
    ws3 = wb.create_sheet("Flotte")
    cols3 = ["Type Véhicule", "Marque", "Modèle", "Plaque Immatriculation", "Statut"]
    style_header(ws3, cols3)
    for v in fleet:
        ws3.append([
            v.get("type_vehicule", ""),
            v.get("marque", ""),
            v.get("modele", ""),
            v.get("plaque", ""),
            v.get("statut", ""),
        ])
    auto_width(ws3)
    if ws3.dimensions:
        ws3.auto_filter.ref = ws3.dimensions

    wb.save(path)
    print(f"    📊 Excel : {path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  FICHIERS GLOBAUX
# ═══════════════════════════════════════════════════════════════════════════════

def load_global_data():
    """Charge les données globales existantes."""
    if GLOBAL_JSON.exists():
        with open(GLOBAL_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def update_global_data(partner):
    """Met à jour le partenaire dans les fichiers globaux."""
    data = load_global_data()

    # Remplacer ou ajouter
    found = False
    for i, p in enumerate(data):
        if p.get("nom") == partner["nom"]:
            data[i] = partner
            found = True
            break
    if not found:
        data.append(partner)

    # Trier par nom
    data.sort(key=lambda x: x.get("nom", ""))

    # JSON
    PARTNERS_DIR.mkdir(parents=True, exist_ok=True)
    with open(GLOBAL_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # HTML
    _export_global_html(data)

    # XLSX
    _export_global_xlsx(data)

    print(f"  📦 Fichiers globaux mis à jour ({len(data)} partenaires au total)")


def _export_global_html(data):
    """Génère le rapport HTML global."""
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    total_partners = len(data)
    total_drivers = sum(p.get("stats", {}).get("nb_conducteurs", 0) for p in data)
    total_fleet = sum(p.get("stats", {}).get("nb_flottes", 0) for p in data)

    css = """
    * { margin:0; padding:0; box-sizing:border-box; }
    body { font-family:'Segoe UI',Tahoma,sans-serif; background:#f4f6f9; color:#333; padding:20px; }
    .header { background:linear-gradient(135deg,#1abc9c,#16a085); color:#fff; padding:30px; border-radius:10px; margin-bottom:20px; }
    .header h1 { font-size:1.8em; }
    .header .meta { opacity:.85; margin-top:5px; }
    .global-stats { display:flex; gap:15px; margin-bottom:25px; flex-wrap:wrap; }
    .stat-card { background:#fff; border-radius:8px; padding:20px; flex:1; min-width:180px;
                 box-shadow:0 2px 8px rgba(0,0,0,.08); text-align:center; }
    .stat-card .number { font-size:2em; font-weight:700; color:#1abc9c; }
    .stat-card .label  { font-size:.9em; color:#777; margin-top:5px; }
    .partner-card { background:#fff; border-radius:8px; margin-bottom:20px;
                    box-shadow:0 2px 8px rgba(0,0,0,.08); overflow:hidden; }
    .partner-header { background:#2c3e50; color:#fff; padding:15px 20px;
                      display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; }
    .partner-header h2 { font-size:1.2em; }
    .partner-info { padding:15px 20px; display:flex; gap:20px; flex-wrap:wrap; font-size:.9em; color:#555; }
    .partner-info span { background:#ecf0f1; padding:4px 10px; border-radius:4px; }
    .badge { display:inline-block; padding:3px 10px; border-radius:12px; font-size:.8em; font-weight:600; }
    .badge-actif   { background:#d5f5e3; color:#1e8449; }
    .badge-inactif { background:#fadbd8; color:#c0392b; }
    .section-title { padding:10px 20px; font-weight:600; background:#f8f9fa; border-top:1px solid #eee; }
    table { width:100%; border-collapse:collapse; font-size:.85em; }
    th { background:#34495e; color:#fff; padding:8px 12px; text-align:left; }
    td { padding:8px 12px; border-bottom:1px solid #eee; }
    tr:hover td { background:#f9f9f9; }
    .empty { padding:15px 20px; color:#999; font-style:italic; }
    .stats-bar { padding:10px 20px; display:flex; gap:15px; font-size:.85em; flex-wrap:wrap; }
    .stats-bar span { padding:4px 10px; border-radius:4px; }
    .stats-bar .approved    { background:#d5f5e3; color:#1e8449; }
    .stats-bar .disapproved { background:#fadbd8; color:#c0392b; }
    .stats-bar .fleet       { background:#d6eaf8; color:#2471a3; }
    .updated { font-size:.75em; color:#95a5a6; margin-left:10px; }
    """

    parts = [f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Rapport Partners UpJunoo - {now}</title>
<style>{css}</style></head><body>
<div class="header"><h1>Rapport Partners UpJunoo</h1>
<div class="meta">Genere le {now}</div></div>
<div class="global-stats">
  <div class="stat-card"><div class="number">{total_partners}</div><div class="label">Partenaires</div></div>
  <div class="stat-card"><div class="number">{total_drivers}</div><div class="label">Conducteurs</div></div>
  <div class="stat-card"><div class="number">{total_fleet}</div><div class="label">Vehicules</div></div>
</div>"""]

    for idx, p in enumerate(data, 1):
        st = p.get("stats", {})
        updated = p.get("updated_at", "")
        if updated:
            try:
                dt = datetime.fromisoformat(updated)
                updated = dt.strftime("%d/%m/%Y %H:%M")
            except:
                pass

        parts.append(f"""<div class="partner-card">
<div class="partner-header"><h2>{idx}. {p.get('nom','N/A')}</h2>
<span class="updated">MAJ: {updated}</span></div>
<div class="partner-info">
  <span>Email: {p.get('email','N/A')}</span>
  <span>Tel: {p.get('telephone','N/A')}</span></div>
<div class="stats-bar">
  <span class="approved">{st.get('nb_conducteurs_approuves',0)} conducteurs approuves</span>
  <span class="disapproved">{st.get('nb_conducteurs_desapprouves',0)} conducteurs desapprouves</span>
  <span class="fleet">{st.get('nb_flottes',0)} vehicules</span></div>""")

        # Drivers table
        drivers = p.get("drivers", [])
        parts.append('<div class="section-title">Conducteurs</div>')
        if drivers:
            parts.append(
                "<table><thead><tr><th>Nom</th><th>Emplacement</th>"
                "<th>Telephone</th><th>Type Transport</th>"
                "<th>Statut</th><th>Vehicule</th></tr></thead><tbody>")
            for d in drivers:
                v = d.get("vehicle", {})
                if v:
                    v_str = f"{v.get('type','N/A')} | {v.get('marque','N/A')} {v.get('modele','N/A')} | {v.get('matricule','N/A')}"
                else:
                    v_str = d.get("type_vehicule", "N/A")
                ds = d.get("statut_approuve", "N/A")
                db = "badge-actif" if ds.strip().upper() in ["APPROUVÉ", "APPROVED", "ACTIF"] else "badge-inactif"
                parts.append(
                    f"<tr><td>{d.get('nom','')}</td><td>{d.get('emplacement','')}</td>"
                    f"<td>{d.get('telephone','')}</td><td>{d.get('type_transport','')}</td>"
                    f'<td><span class="badge {db}">{ds}</span></td>'
                    f"<td>{v_str}</td></tr>")
            parts.append("</tbody></table>")
        else:
            parts.append('<div class="empty">Aucun conducteur.</div>')

        # Fleet table
        fleet = p.get("fleet", [])
        parts.append('<div class="section-title">Flotte</div>')
        if fleet:
            parts.append(
                "<table><thead><tr><th>Type</th><th>Marque</th>"
                "<th>Modele</th><th>Plaque</th>"
                "<th>Statut</th></tr></thead><tbody>")
            for v in fleet:
                parts.append(
                    f"<tr><td>{v.get('type_vehicule','')}</td><td>{v.get('marque','')}</td>"
                    f"<td>{v.get('modele','')}</td><td>{v.get('plaque','')}</td>"
                    f"<td>{v.get('statut','')}</td></tr>")
            parts.append("</tbody></table>")
        else:
            parts.append('<div class="empty">Aucun vehicule.</div>')

        parts.append("</div>")  # close partner-card

    parts.append("</body></html>")

    with open(GLOBAL_HTML, "w", encoding="utf-8") as f:
        f.write("\n".join(parts))
    print(f"    🌐 HTML global : {GLOBAL_HTML}")


def _export_global_xlsx(data):
    """Génère le fichier Excel global (3 onglets)."""
    wb = Workbook()

    # Sheet 1: Résumé
    ws1 = wb.active
    ws1.title = "Résumé Partenaires"
    cols1 = ["Partenaire", "Email", "Téléphone",
             "Conducteurs Approuvés", "Conducteurs Désapprouvés", "Véhicules",
             "Total Conducteurs", "Mis à jour"]
    style_header(ws1, cols1)
    for p in data:
        s = p.get("stats", {})
        ws1.append([
            p.get("nom", ""),
            p.get("email", ""),
            p.get("telephone", ""),
            s.get("nb_conducteurs_approuves", 0),
            s.get("nb_conducteurs_desapprouves", 0),
            s.get("nb_flottes", 0),
            s.get("nb_conducteurs", 0),
            p.get("updated_at", ""),
        ])
    auto_width(ws1)
    if ws1.dimensions:
        ws1.auto_filter.ref = ws1.dimensions

    # Sheet 2: Conducteurs (tous partenaires)
    ws2 = wb.create_sheet("Conducteurs")
    cols2 = ["Partenaire", "Nom", "Emplacement", "Téléphone",
             "Type Transport", "Statut", "Type Véhicule", "Marque", "Modèle", "Matricule"]
    style_header(ws2, cols2)
    for p in data:
        for d in p.get("drivers", []):
            v = d.get("vehicle", {})
            ws2.append([
                p.get("nom", ""),
                d.get("nom", ""),
                d.get("emplacement", ""),
                d.get("telephone", ""),
                d.get("type_transport", ""),
                d.get("statut_approuve", ""),
                v.get("type", d.get("type_vehicule", "")),
                v.get("marque", ""),
                v.get("modele", ""),
                v.get("matricule", ""),
            ])
    auto_width(ws2)
    if ws2.dimensions:
        ws2.auto_filter.ref = ws2.dimensions

    # Sheet 3: Flotte (tous partenaires)
    ws3 = wb.create_sheet("Flotte")
    cols3 = ["Partenaire", "Type Véhicule", "Marque", "Modèle",
             "Plaque Immatriculation", "Statut"]
    style_header(ws3, cols3)
    for p in data:
        for v in p.get("fleet", []):
            ws3.append([
                p.get("nom", ""),
                v.get("type_vehicule", ""),
                v.get("marque", ""),
                v.get("modele", ""),
                v.get("plaque", ""),
                v.get("statut", ""),
            ])
    auto_width(ws3)
    if ws3.dimensions:
        ws3.auto_filter.ref = ws3.dimensions

    wb.save(GLOBAL_XLSX)
    print(f"    📊 Excel global : {GLOBAL_XLSX}")


# ═══════════════════════════════════════════════════════════════════════════════
#  SCRAPING : LISTE PARTENAIRES
# ═══════════════════════════════════════════════════════════════════════════════

def scrape_partner_list(driver):
    """Extrait la liste complète des partenaires du tableau manage-owners."""
    partners = []
    try:
        print("\n🔍 Extraction de la liste des partenaires…")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr")))
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")

        for i, row in enumerate(rows):
            try:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) < 6:
                    continue
                nom = cells[0].text.strip()
                email = cells[1].text.strip()
                telephone = cells[2].text.strip()
                document_url = "N/A"
                profile_url = "N/A"
                owner_id = "N/A"
                statut_approbation = "N/A"

                try:
                    link_doc = cells[3].find_element(By.TAG_NAME, "a")
                    document_url = link_doc.get_attribute("href")
                    if document_url and "/document/" in document_url:
                        owner_id = document_url.split("/document/")[-1]
                        profile_url = f"{BASE_URL}/manage-owners/view-profile/{owner_id}"
                except:
                    pass

                try:
                    toggle = row.find_element(
                        By.CSS_SELECTOR,
                        "input[type='checkbox'], .custom-switch input, .switch input, .toggle input")
                    statut_approbation = "Actif" if toggle.is_selected() else "Inactif"
                except:
                    try:
                        toggle_el = row.find_element(
                            By.CSS_SELECTOR,
                            ".custom-control-input, .form-check-input, input[role='switch']")
                        statut_approbation = "Actif" if toggle_el.is_selected() else "Inactif"
                    except:
                        pass

                partners.append({
                    "nom": nom,
                    "email": email,
                    "telephone": telephone,
                    "statut_approbation": statut_approbation,
                    "document_url": document_url,
                    "profile_url": profile_url,
                    "owner_id": owner_id,
                })
            except Exception as e:
                print(f"  ⚠️ Erreur ligne {i}: {e}")

        print(f"✅ {len(partners)} partenaires indexés.")
    except Exception as e:
        print(f"❌ Erreur extraction liste : {e}")
    return partners


def find_partner(partners_index, name):
    """Trouve un partenaire par nom (case-insensitive, partial match)."""
    name_lower = name.strip().lower()
    # Match exact
    for p in partners_index:
        if p["nom"].strip().lower() == name_lower:
            return p
    # Match partiel
    matches = [p for p in partners_index if name_lower in p["nom"].strip().lower()]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        print(f"  ⚠️ Plusieurs correspondances pour '{name}':")
        for i, m in enumerate(matches):
            print(f"    {i + 1}. {m['nom']} ({m['email']})")
        try:
            choix = int(input("  👉 Choisissez le numéro : "))
            if 1 <= choix <= len(matches):
                return matches[choix - 1]
        except:
            pass
    return None


# ═══════════════════════════════════════════════════════════════════════════════
#  SCRAPING : CONDUCTEURS + FLOTTE (profil partenaire)
# ═══════════════════════════════════════════════════════════════════════════════

def wait_for_table_stable(driver, timeout=30, stable_delay=2):
    """Attend que le tableau dans l'onglet actif soit chargé et stabilisé.
    Vérifie que le nombre de lignes ne change plus pendant `stable_delay` secondes.
    """
    selectors = [
        ".tab-pane.active table tbody tr",
        ".tab-pane.show table tbody tr",
        "table tbody tr",
    ]

    def count_rows():
        for sel in selectors:
            rows = driver.find_elements(By.CSS_SELECTOR, sel)
            if rows:
                return len(rows), sel
        return 0, selectors[-1]

    print(f"    ⏳ Attente chargement du tableau (max {timeout}s)…")
    start = time.time()
    last_count = 0
    last_change = time.time()
    winning_sel = selectors[-1]

    while time.time() - start < timeout:
        current_count, winning_sel = count_rows()
        if current_count != last_count:
            last_count = current_count
            last_change = time.time()
        elif current_count > 0 and (time.time() - last_change) >= stable_delay:
            print(f"    ✅ Tableau stabilisé : {current_count} lignes détectées.")
            return driver.find_elements(By.CSS_SELECTOR, winning_sel)
        time.sleep(0.5)

    # Timeout — retourner ce qu'on a
    rows = driver.find_elements(By.CSS_SELECTOR, winning_sel)
    if rows:
        print(f"    ⚠️ Timeout atteint mais {len(rows)} lignes récupérées.")
    else:
        print(f"    ⚠️ Timeout atteint — aucune ligne trouvée.")
    return rows


def scrape_drivers_tab(driver, partner):
    """Navigue vers le profil et scrape l'onglet conducteurs (full auto)."""
    drivers = []
    if not partner.get("profile_url") or partner["profile_url"] == "N/A":
        print("    ⚠️ Pas d'URL de profil — conducteurs ignorés.")
        return []

    try:
        print("    🌐 Navigation vers le profil…")
        driver.get(partner["profile_url"])
        WebDriverWait(driver, 15).until(EC.presence_of_element_located(
            (By.XPATH,
             "//a[contains(text(), 'Détails du conducteur')] | //a[contains(text(), 'Driver Details')]")))

        # Clic automatique sur l'onglet conducteurs
        tab = driver.find_element(
            By.XPATH,
            "//a[contains(text(), 'Détails du conducteur')] | //a[contains(text(), 'Driver Details')]")
        driver.execute_script("arguments[0].click();", tab)
        print("    �️ Clic onglet 'Détails du conducteur'…")
        time.sleep(1)

        # Attente intelligente : le tableau doit se stabiliser
        rows = wait_for_table_stable(driver, timeout=30, stable_delay=2)

        for dr in rows:
            try:
                d_cells = dr.find_elements(By.TAG_NAME, "td")
                if len(d_cells) >= 5:
                    view_profile = "N/A"
                    try:
                        links = dr.find_elements(By.TAG_NAME, "a")
                        for link in links:
                            href = link.get_attribute("href") or ""
                            if "/document/" in href:
                                driver_id = href.split("/document/")[-1]
                                view_profile = f"{BASE_URL}/fleet-drivers/view-profile/{driver_id}"
                                break
                    except:
                        pass

                    drivers.append({
                        "nom": d_cells[0].text.strip(),
                        "emplacement": d_cells[1].text.strip(),
                        "telephone": d_cells[2].text.strip(),
                        "type_transport": d_cells[3].text.strip(),
                        "statut_approuve": d_cells[4].text.strip(),
                        "type_vehicule": d_cells[6].text.strip() if len(d_cells) > 6 else "N/A",
                        "view_profile": view_profile,
                    })
            except StaleElementReferenceException:
                continue

        print(f"    ✅ {len(drivers)} conducteurs extraits.")

    except Exception as e:
        print(f"    ❌ Erreur conducteurs : {e}")

    return drivers


def scrape_fleet_tab(driver, partner):
    """Scrape l'onglet flotte (clic automatique + attente stabilisation)."""
    fleet = []
    if not partner.get("profile_url") or partner["profile_url"] == "N/A":
        return []

    try:
        # Clic automatique sur l'onglet flotte
        try:
            fleet_tab = driver.find_element(
                By.XPATH,
                "//a[contains(text(), 'Détails de la flotte')] | //a[contains(text(), 'Fleet Details')]")
            driver.execute_script("arguments[0].click();", fleet_tab)
            print("    🖱️ Clic onglet 'Détails de la flotte'…")
            time.sleep(1)
        except NoSuchElementException:
            print("    ⚠️ Onglet 'Détails de la flotte' introuvable.")
            return []

        # Attente intelligente : le tableau doit se stabiliser
        rows = wait_for_table_stable(driver, timeout=30, stable_delay=2)

        for fr in rows:
            try:
                f_cells = fr.find_elements(By.TAG_NAME, "td")
                if len(f_cells) >= 4:
                    fleet.append({
                        "type_vehicule": f_cells[0].text.strip(),
                        "marque": f_cells[1].text.strip() if len(f_cells) > 1 else "N/A",
                        "modele": f_cells[2].text.strip() if len(f_cells) > 2 else "N/A",
                        "plaque": f_cells[3].text.strip() if len(f_cells) > 3 else "N/A",
                        "statut": f_cells[4].text.strip() if len(f_cells) > 4 else "N/A",
                    })
            except StaleElementReferenceException:
                continue

        print(f"    ✅ {len(fleet)} véhicules extraits.")

    except Exception as e:
        print(f"    ⚠️ Erreur flotte : {e}")

    return fleet


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def run():
    opts = Options()
    opts.add_argument("--window-size=1600,1000")

    browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)

    try:
        browser.get(f"{BASE_URL}/login/admin")

        print("\n" + "=" * 60)
        print(" UPDATE PARTNER — Mise à jour ciblée")
        print("=" * 60)
        print(f"1. Connectez-vous sur : {BASE_URL}/login/admin")
        print(f"2. Allez dans : Gestion par le partenaire > Approuvé Partenaires")
        print(f"   Lien direct : {BASE_URL}/manage-owners")
        print("3. Réglez le filtre de pagination sur '200' ou '500'.")
        print("=" * 60)

        input("\n👉 UNE FOIS PRÊT (FILTRE APPLIQUÉ), APPUYEZ SUR [ENTRÉE]…")

        # ── Phase 1 : Indexer la liste ──
        partners_index = scrape_partner_list(browser)
        if not partners_index:
            print("❌ Aucun partenaire trouvé. Fin du script.")
            return

        # ── Phase 2 : Traitement automatique de tous les partenaires ──
        total = len(partners_index)
        success = 0
        errors_list = []
        start_all = time.time()

        print(f"\n🚀 Lancement du traitement automatique de {total} partenaires…\n")

        for idx, partner in enumerate(partners_index, 1):
            print("\n" + "─" * 60)
            print(f"  [{idx}/{total}] 🔄 {partner['nom']}")
            print(f"     Email: {partner['email']} | Tel: {partner['telephone']}")

            try:
                # Conducteurs
                print(f"\n  🧑‍✈️ Phase Conducteurs…")
                drivers = scrape_drivers_tab(browser, partner)
                partner["drivers"] = drivers

                # Flotte
                print(f"\n  🚗 Phase Flotte…")
                fleet = scrape_fleet_tab(browser, partner)
                partner["fleet"] = fleet

                # Enrichissement véhicules (automatique)
                if drivers:
                    has_profiles = any(d.get("view_profile", "N/A") != "N/A" for d in drivers)
                    if has_profiles:
                        enrich_drivers_vehicles(drivers, browser)
                    else:
                        print("    ℹ️ Aucun lien document conducteur trouvé — véhicules non enrichis.")

                # Stats
                partner["stats"] = compute_stats(partner)
                partner["updated_at"] = datetime.now().isoformat()

                # Export individuel
                partner_dir = PARTNERS_DIR / partner["nom"]
                partner_dir.mkdir(parents=True, exist_ok=True)
                print(f"\n  📁 Export → {partner_dir}")
                export_partner_json(partner, partner_dir)
                export_partner_html(partner, partner_dir)
                export_partner_xlsx(partner, partner_dir)

                # Global
                update_global_data(partner)

                success += 1
                print(f"\n  ✅ {partner['nom']} terminé !")
                print(f"     • {len(drivers)} conducteurs ({partner['stats']['nb_conducteurs_approuves']} approuvés)")
                print(f"     • {len(fleet)} véhicules flotte")

            except Exception as e:
                errors_list.append(partner['nom'])
                print(f"\n  ❌ Erreur sur {partner['nom']}: {e}")
                traceback.print_exc()
                print("  ⏩ Passage au partenaire suivant…")

        # ── Bilan final ──
        elapsed = time.time() - start_all
        print("\n" + "=" * 60)
        print(f" SESSION TERMINÉE en {elapsed/60:.1f} minutes")
        print(f" ✅ {success}/{total} partenaires traités avec succès")
        if errors_list:
            print(f" ❌ {len(errors_list)} erreurs : {', '.join(errors_list)}")
        print("=" * 60)

    except Exception as e:
        print(f"\n❌ ERREUR CRITIQUE: {e}")
        traceback.print_exc()
    finally:
        print("\n👋 Fermeture du navigateur dans 5 secondes…")
        time.sleep(5)
        browser.quit()


if __name__ == "__main__":
    run()
