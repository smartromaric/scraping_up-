#!/usr/bin/env python3
import argparse
import json
import re
import shutil
import sys
import time
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait

if sys.platform == "win32":
    import io

    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE_URL = "https://upjunoo-server-new.junooapps.com"
OWNER_LOGIN_URL = f"{BASE_URL}/login/owner-login"
FLEET_URL = f"{BASE_URL}/manage-fleet"
UNIVERSAL_PASSWORD = "123456789@"
PARTNER_RE = re.compile(r"(partenaires?|partenaire)[-_\s]*(\d+)", re.IGNORECASE)


def log(msg: str, level: str = "INFO") -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] [{level}] {msg}", flush=True)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text


def normalize_partner(name: str) -> str:
    m = PARTNER_RE.search(name or "")
    if not m:
        return ""
    return f"partenaire{int(m.group(2))}"


def extract_partner_num(name: str) -> int:
    m = PARTNER_RE.search(name or "")
    return int(m.group(2)) if m else 0


def normalize_immat(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(value)).upper()


def derive_email(partner_name: str) -> str:
    m = PARTNER_RE.search(partner_name or "")
    if not m:
        return ""
    return f"partenaire{int(m.group(2))}@upjunoo.com"


def load_partner_email_map(data_root: Path) -> dict[str, str]:
    email_map: dict[str, str] = {}
    if not data_root.exists():
        return email_map

    for folder in sorted([p for p in data_root.iterdir() if p.is_dir()]):
        num = extract_partner_num(folder.name)
        if num <= 0:
            continue

        email = ""
        jf = folder / "data.json"
        if jf.exists():
            try:
                with open(jf, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                email = str(payload.get("email", "") or "").strip()
            except Exception:
                email = ""

        if not email:
            lower_name = folder.name.lower()
            prefix = "partenaires" if "partenaires" in lower_name else "partenaire"
            email = f"{prefix}{num}@upjunoo.com"

        email_map[f"partenaire{num}"] = email
        email_map[f"partenaires{num}"] = email

    return email_map


def read_excel_targets(excel_path: Path) -> dict[str, set[str]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    col_partner = None
    col_matricule = None
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        values = [normalize_text(c.value) for c in ws[row_idx]]
        if "partenaire" in values and "matricule" in values:
            header_row = row_idx
            col_partner = values.index("partenaire")
            col_matricule = values.index("matricule")
            break

    if header_row is None or col_partner is None or col_matricule is None:
        raise RuntimeError("Colonnes 'PARTENAIRE' et/ou 'Matricule' introuvables dans le fichier Excel.")

    targets = defaultdict(set)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        partner_raw = row[col_partner] if col_partner < len(row) else None
        matricule_raw = row[col_matricule] if col_matricule < len(row) else None
        partner = normalize_partner(str(partner_raw or ""))
        matricule = normalize_immat(matricule_raw)
        if not partner or not matricule:
            continue
        targets[partner].add(matricule)

    return dict(targets)


def make_driver(headless: bool = True) -> webdriver.Chrome:
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    try:
        return webdriver.Chrome(options=opts)
    except Exception:
        svc = Service(shutil.which("chromedriver.exe") or "chromedriver.exe")
        return webdriver.Chrome(service=svc, options=opts)


def login_partner(driver, email: str, password: str) -> bool:
    log(f"🔐 Connexion partenaire: {email}")
    try:
        driver.delete_all_cookies()
        driver.get(OWNER_LOGIN_URL)
        wait = WebDriverWait(driver, 25)
        email_input = wait.until(EC.presence_of_element_located((By.ID, "email-input")))
        pwd_input = driver.find_element(By.ID, "password-input")
        email_input.clear()
        email_input.send_keys(email)
        pwd_input.clear()
        pwd_input.send_keys(password)
        submit = driver.find_element(By.XPATH, "//button[@type='submit'] | //button[contains(@class, 'btn-success')]")
        driver.execute_script("arguments[0].click();", submit)
        wait.until(lambda d: "login" not in d.current_url.lower())
        return True
    except Exception as exc:
        log(f"❌ Login échoué: {exc}", "ERROR")
        return False


def set_page_size(driver, size: int = 500) -> None:
    selectors = ["select.form-select-sm", "select[name*='_length']", ".dataTables_length select", "select"]
    for sel in selectors:
        try:
            element = WebDriverWait(driver, 12).until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
            select_obj = Select(element)
            wanted = str(size)
            values = [o.get_attribute("value") for o in select_obj.options]
            target = wanted if wanted in values else max((v for v in values if str(v).isdigit()), key=int, default=None)
            if target:
                select_obj.select_by_value(target)
                log(f"📄 Pagination réglée sur {target}")
                time.sleep(2)
                return
        except Exception:
            continue


def wait_for_table_data(driver, timeout: int = 20) -> bool:
    start = time.time()
    ignore_words = ["chargement", "loading", "aucune", "no data", "na", "vide"]
    while (time.time() - start) < timeout:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        if rows:
            txt = (rows[0].text or "").strip().lower()
            if txt and not any(word in txt for word in ignore_words):
                return True
        time.sleep(1)
    return False


def load_fleet_page(driver) -> None:
    driver.get(FLEET_URL)
    try:
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr, table")))
    except TimeoutException:
        pass
    time.sleep(1)
    set_page_size(driver, 500)
    time.sleep(8)
    # Important: après changement de pagination, attendre que la vraie liste se recharge.
    if wait_for_table_data(driver, timeout=35):
        log("✅ Liste véhicules chargée après pagination 500")
    else:
        log("⚠️ Chargement liste véhicules lent/incomplet après pagination 500", "WARNING")


def scrape_fleet_rows(driver) -> list[dict]:
    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
    data = []
    for idx, row in enumerate(rows):
        try:
            tds = row.find_elements(By.TAG_NAME, "td")
            if len(tds) < 5:
                continue
            plaque = tds[4].text.strip()
            if not plaque:
                continue
            data.append({"idx": idx, "row": row, "plaque": plaque, "norm_plaque": normalize_immat(plaque)})
        except Exception:
            continue
    return data


def delete_row_at_index(driver, row_index: int) -> bool:
    try:
        result = driver.execute_script(
            """
            var rows = document.querySelectorAll('table tbody tr');
            if (!rows || arguments[0] >= rows.length) return 'no_row';
            var row = rows[arguments[0]];

            row.style.outline = '3px solid #ff0000';
            row.style.background = '#fff3cd';
            row.scrollIntoView({block:'center'});

            var btn = row.querySelector(
                'button[data-bs-toggle="dropdown"], button.dropdown-toggle, .btn-action'
            );
            if (!btn) {
                // Fallback: dernier bouton visible de la ligne
                var btns = row.querySelectorAll('button, a');
                for (var b = btns.length - 1; b >= 0; b--) {
                    var cand = btns[b];
                    if (cand && cand.offsetParent !== null) {
                        btn = cand;
                        break;
                    }
                }
            }
            if (!btn) return 'no_btn';

            // Bootstrap peut planter avec un élément non conforme -> fallback direct
            try {
                if (window.bootstrap && window.bootstrap.Dropdown && btn.matches('[data-bs-toggle="dropdown"], .dropdown-toggle')) {
                    var dd = window.bootstrap.Dropdown.getOrCreateInstance(btn);
                    if (dd && dd.show) {
                        dd.show();
                    } else {
                        btn.click();
                    }
                } else {
                    btn.click();
                }
            } catch (e) {
                btn.click();
            }

            return new Promise(function(resolve) {
                setTimeout(function() {
                    var menu = row.querySelector('.dropdown-menu.show') || row.querySelector('.dropdown-menu');
                    if (!menu) {
                        // Fallback global si menu hors de la ligne
                        menu = document.querySelector('.dropdown-menu.show');
                    }
                    if (!menu) { resolve('no_menu'); return; }
                    var items = menu.querySelectorAll('a, button, li, span');
                    for (var i = 0; i < items.length; i++) {
                        var t = (items[i].textContent || '').trim().toLowerCase();
                        if (t.includes('supprimer') || t.includes('delete') || t.includes('retirer')) {
                            items[i].click();
                            resolve('clicked');
                            return;
                        }
                    }
                    resolve('no_supprimer');
                }, 500);
            });
            """,
            row_index,
        )
        if result != "clicked":
            log(f"⚠️ Suppression index {row_index}: {result}", "WARNING")
            return False

        confirm = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".swal2-popup .swal2-confirm"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", confirm)
        time.sleep(0.2)
        try:
            ActionChains(driver).move_to_element(confirm).click().perform()
        except Exception:
            driver.execute_script("arguments[0].click();", confirm)

        time.sleep(0.5)
        try:
            ok = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".swal2-popup .swal2-confirm"))
            )
            driver.execute_script("arguments[0].click();", ok)
        except TimeoutException:
            pass

        end = time.time() + 6
        while time.time() < end:
            if not driver.find_elements(By.CSS_SELECTOR, ".swal2-container .swal2-popup"):
                return True
            time.sleep(0.2)
        return True
    except Exception as exc:
        log(f"❌ Erreur suppression ligne {row_index}: {exc}", "ERROR")
        try:
            driver.execute_script("var b=document.querySelector('.swal2-cancel'); if(b) b.click();")
        except Exception:
            pass
        return False


def delete_targets_for_partner(driver, partner: str, immats: set[str], dry_run: bool = False) -> dict:
    stats = {"target_count": len(immats), "deleted": 0, "failed": 0, "not_found": 0}
    remaining = set(immats)
    max_passes = 30

    for pass_num in range(1, max_passes + 1):
        load_fleet_page(driver)
        rows = scrape_fleet_rows(driver)
        if not rows:
            log(f"⚠️ Aucun véhicule affiché pour {partner}", "WARNING")

        matches = [r for r in rows if r["norm_plaque"] in remaining]
        if not matches:
            break

        matches.sort(key=lambda x: x["idx"], reverse=True)
        log(f"🗑️ {partner} | Pass {pass_num}: {len(matches)} véhicule(s) à supprimer")

        for item in matches:
            immat = item["norm_plaque"]
            if dry_run:
                log(f"   [DRY-RUN] Suppression simulée: {item['plaque']}")
                stats["deleted"] += 1
                remaining.discard(immat)
                continue

            ok = delete_row_at_index(driver, item["idx"])
            if ok:
                stats["deleted"] += 1
                remaining.discard(immat)
                time.sleep(0.5)
            else:
                stats["failed"] += 1

        if not remaining:
            break

    stats["not_found"] = len(remaining)
    if remaining:
        preview = ", ".join(sorted(list(remaining))[:10])
        more = "" if len(remaining) <= 10 else f" ... (+{len(remaining)-10})"
        log(f"ℹ️ {partner}: non trouvés/supprimés: {preview}{more}", "WARNING")
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Supprime les véhicules listés dans un Excel, partenaire par partenaire.")
    parser.add_argument(
        "--excel",
        default=r"C:\Users\c.romaric\Desktop\scraping\vps_deploy\Matrice_appel_errone_injoignable.xlsx",
        help="Chemin du fichier Excel filtré.",
    )
    parser.add_argument(
        "--data-root",
        default=r"C:\Users\c.romaric\Desktop\scraping\vps_deploy\output\organized_by_partner",
        help="Dossier des partenaires (pour récupérer l'email réel depuis data.json).",
    )
    parser.add_argument("--only", help="Traiter un partenaire uniquement (ex: partenaire1).")
    parser.add_argument("--start", type=int, help="Commencer à partir de ce numéro partenaire (ex: 51).")
    parser.add_argument("--dry-run", action="store_true", help="Simulation sans suppression réelle.")
    parser.add_argument("--no-headless", action="store_true", help="Afficher Chrome.")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    data_root = Path(args.data_root)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel introuvable: {excel_path}")

    targets_by_partner = read_excel_targets(excel_path)
    if args.only:
        key = normalize_partner(args.only)
        targets_by_partner = {k: v for k, v in targets_by_partner.items() if k == key}
    if args.start:
        targets_by_partner = {
            k: v for k, v in targets_by_partner.items() if extract_partner_num(k) >= args.start
        }

    if not targets_by_partner:
        log("Aucun véhicule à traiter selon les filtres.", "WARNING")
        return

    partner_email_map = load_partner_email_map(data_root)

    total_targets = sum(len(v) for v in targets_by_partner.values())
    log(f"📦 Cibles: {len(targets_by_partner)} partenaire(s), {total_targets} matricule(s) uniques")
    if args.dry_run:
        log("🧪 Mode DRY-RUN activé")

    driver = make_driver(headless=not args.no_headless)
    summary = {}
    try:
        for partner in sorted(targets_by_partner.keys(), key=lambda x: int(re.findall(r"\d+", x)[0])):
            immats = targets_by_partner[partner]
            email = partner_email_map.get(partner) or derive_email(partner)
            log(f"\n{'='*65}")
            log(f"👤 Partenaire: {partner} | 📧 {email} | 🎯 {len(immats)}")
            log(f"{'='*65}")

            if not email:
                log(f"❌ Email impossible à déduire pour {partner}", "ERROR")
                summary[partner] = {"target_count": len(immats), "deleted": 0, "failed": 0, "not_found": len(immats)}
                continue

            if not login_partner(driver, email, UNIVERSAL_PASSWORD):
                summary[partner] = {"target_count": len(immats), "deleted": 0, "failed": 0, "not_found": len(immats)}
                continue

            stats = delete_targets_for_partner(driver, partner, immats, dry_run=args.dry_run)
            summary[partner] = stats
            log(
                f"✅ {partner} | cible={stats['target_count']} | supprimés={stats['deleted']} | "
                f"échecs={stats['failed']} | non_trouvés={stats['not_found']}"
            )
            driver.get(f"{BASE_URL}/logout")
            time.sleep(1)
    finally:
        driver.quit()

    total_deleted = sum(v["deleted"] for v in summary.values())
    total_failed = sum(v["failed"] for v in summary.values())
    total_not_found = sum(v["not_found"] for v in summary.values())
    log(f"\n🏁 FIN | supprimés={total_deleted} | échecs={total_failed} | non_trouvés={total_not_found}")


if __name__ == "__main__":
    main()
