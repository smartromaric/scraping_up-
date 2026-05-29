#!/usr/bin/env python3
import argparse
import json
import re
import shutil
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def log(msg: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def normalize_partner_key(value: object) -> str:
    txt = normalize_text(value)
    txt = re.sub(r"[\s\-_]", "", txt)
    txt = txt.replace("partenaires", "partenaire")
    return txt


def normalize_matricule(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(value)).upper()


def canonical_phone(value: object) -> str:
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if not digits:
        return ""
    # Harmonisation vers format +2250XXXXXXXXX (10 derniers digits conservés)
    local = digits[-10:] if len(digits) >= 10 else digits
    if len(local) < 10:
        return ""
    return f"+225{local}"


def read_excel_phone_map(excel_path: Path) -> tuple[dict[str, dict[str, str]], list[str]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    col_partner = None
    col_matricule = None
    col_numero = None

    for row_idx in range(1, min(ws.max_row, 40) + 1):
        vals = [normalize_text(c.value) for c in ws[row_idx]]
        if "partenaire" in vals and "matricule" in vals and "numero" in vals:
            header_row = row_idx
            col_partner = vals.index("partenaire")
            col_matricule = vals.index("matricule")
            col_numero = vals.index("numero")
            break

    if header_row is None:
        raise RuntimeError("Colonnes PARTENAIRE/Matricule/Numero introuvables dans le fichier Excel.")

    mapping: dict[str, dict[str, str]] = defaultdict(dict)
    conflicts: list[str] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        partner = normalize_partner_key(row[col_partner] if col_partner < len(row) else "")
        matricule = normalize_matricule(row[col_matricule] if col_matricule < len(row) else "")
        phone = canonical_phone(row[col_numero] if col_numero < len(row) else "")

        if not partner or not matricule or not phone:
            continue

        prev = mapping[partner].get(matricule)
        if prev and prev != phone:
            conflicts.append(f"{partner} | {matricule}: {prev} vs {phone} (dernier conservé)")
        mapping[partner][matricule] = phone

    return dict(mapping), conflicts


def update_json_file(
    json_path: Path,
    partner_key: str,
    phone_map_for_partner: dict[str, str],
    dry_run: bool,
    backup: bool,
) -> dict:
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    drivers = data.get("drivers", [])
    updates = []
    hits = 0

    for idx, driver in enumerate(drivers):
        vehicle = driver.get("vehicle", {}) or {}
        matricule = normalize_matricule(vehicle.get("matricule", ""))
        if not matricule:
            continue

        target_phone = phone_map_for_partner.get(matricule)
        if not target_phone:
            continue

        hits += 1
        old_phone = driver.get("telephone", "")
        old_canon = canonical_phone(old_phone)
        if old_canon != target_phone:
            updates.append(
                {
                    "index": idx,
                    "nom": driver.get("nom", ""),
                    "matricule": vehicle.get("matricule", ""),
                    "old_phone": old_phone,
                    "new_phone": target_phone,
                }
            )
            if not dry_run:
                driver["telephone"] = target_phone

    written = False
    if updates and not dry_run:
        if backup:
            backup_path = json_path.with_suffix(json_path.suffix + ".bak")
            if not backup_path.exists():
                shutil.copy2(json_path, backup_path)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        written = True

    unmatched_in_json = len(phone_map_for_partner) - hits if len(phone_map_for_partner) >= hits else 0
    return {
        "partner": partner_key,
        "json_path": str(json_path),
        "excel_targets": len(phone_map_for_partner),
        "hits_in_json": hits,
        "updates": updates,
        "updated_count": len(updates),
        "unmatched_excel_matricules": unmatched_in_json,
        "written": written,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Met à jour les téléphones dans data.json via match PARTENAIRE+MATRICULE depuis Excel."
    )
    parser.add_argument(
        "--excel",
        default=r"C:\Users\c.romaric\Desktop\scraping\vps_deploy\Matrice_appel_errone_injoignable.xlsx",
        help="Chemin Excel source.",
    )
    parser.add_argument(
        "--data-root",
        default=r"C:\Users\c.romaric\Desktop\scraping\vps_deploy\output\organized_by_partner",
        help="Dossier racine des partenaires contenant data.json.",
    )
    parser.add_argument("--only", help="Filtrer un partenaire (ex: partenaire1).")
    parser.add_argument("--dry-run", action="store_true", help="Simulation sans écriture.")
    parser.add_argument("--no-backup", action="store_true", help="Ne pas créer de .bak avant écriture.")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    data_root = Path(args.data_root)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel introuvable: {excel_path}")
    if not data_root.exists():
        raise FileNotFoundError(f"Data root introuvable: {data_root}")

    phone_map_by_partner, conflicts = read_excel_phone_map(excel_path)
    if args.only:
        only_key = normalize_partner_key(args.only)
        phone_map_by_partner = {k: v for k, v in phone_map_by_partner.items() if k == only_key}

    log(f"📦 Partenaires Excel: {len(phone_map_by_partner)}")
    log(f"🎯 Cibles totales (partenaire+matricule): {sum(len(v) for v in phone_map_by_partner.values())}")
    if args.dry_run:
        log("🧪 Mode DRY-RUN (aucune écriture)")
    if conflicts:
        log(f"⚠️ Conflits Excel détectés: {len(conflicts)} (dernier numéro conservé)")

    results = []
    processed = 0
    for partner_dir in sorted([p for p in data_root.iterdir() if p.is_dir()]):
        partner_key = normalize_partner_key(partner_dir.name)
        if partner_key not in phone_map_by_partner:
            continue
        json_path = partner_dir / "data.json"
        if not json_path.exists():
            log(f"⚠️ data.json absent pour {partner_dir.name}")
            continue
        processed += 1
        log(f"\n🔎 {partner_dir.name}")
        res = update_json_file(
            json_path=json_path,
            partner_key=partner_key,
            phone_map_for_partner=phone_map_by_partner[partner_key],
            dry_run=args.dry_run,
            backup=not args.no_backup,
        )
        results.append(res)
        log(
            f"   hits={res['hits_in_json']} | modifs={res['updated_count']} | "
            f"non_trouvés={res['unmatched_excel_matricules']} | écrit={res['written']}"
        )

    total_updates = sum(r["updated_count"] for r in results)
    total_unmatched = sum(r["unmatched_excel_matricules"] for r in results)
    log(f"\n✅ Partenaires traités: {processed}")
    log(f"🛠️ Téléphones modifiés: {total_updates}")
    log(f"❓ Matricules Excel sans match JSON: {total_unmatched}")

    report = {
        "generated_at": datetime.now().isoformat(),
        "excel": str(excel_path),
        "data_root": str(data_root),
        "dry_run": args.dry_run,
        "conflicts_count": len(conflicts),
        "conflicts": conflicts[:200],
        "results": results,
    }
    report_path = data_root.parent / f"phone_update_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    log(f"📄 Rapport: {report_path}")


if __name__ == "__main__":
    main()
