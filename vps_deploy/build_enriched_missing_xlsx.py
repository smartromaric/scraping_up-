#!/usr/bin/env python3
"""
build_enriched_missing_xlsx.py
Croise all_partners_enriched.json avec missing_plates.xlsx
→ Génère 1_120_partners_enriched.xlsx avec uniquement les conducteurs
  dont le matricule figure dans missing_plates.xlsx

Usage:
    python3 build_enriched_missing_xlsx.py
"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ openpyxl non installé. Lance: pip3 install openpyxl")
    sys.exit(1)

# ─── Chemins ──────────────────────────────────────────────────────────────────
_SCRIPT_DIR   = Path(__file__).parent
_PROJECT_ROOT = _SCRIPT_DIR.parent if _SCRIPT_DIR.name == "vps_deploy" else _SCRIPT_DIR

ENRICHED_JSON   = _PROJECT_ROOT / "output" / "organized_by_partner" / "all_partners_enriched.json"
MISSING_JSON    = _PROJECT_ROOT / "output" / "missing_plates.json"
OUTPUT_XLSX     = _PROJECT_ROOT / "output" / "1_120_partners_enriched.xlsx"


def normalize_plate(plate: str) -> str:
    if not plate:
        return ""
    raw = str(plate).strip().upper()
    for ch in ("-", " ", ".", "/", "_"):
        raw = raw.replace(ch, "")
    return raw


def load_missing_plates(json_path: Path) -> set:
    """Charge les matricules manquants depuis missing_plates.json → set normalisé."""
    if not json_path.exists():
        print(f"❌ Fichier introuvable : {json_path}")
        sys.exit(1)
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    plates = set()
    for entry in data:
        matricule = str(entry.get("matricule") or "").strip()
        if matricule:
            plates.add(normalize_plate(matricule))
    print(f"✅ {len(plates)} matricules manquants chargés depuis {json_path.name}")
    return plates


def load_enriched(json_path: Path) -> list:
    """Charge all_partners_enriched.json."""
    if not json_path.exists():
        print(f"❌ Fichier introuvable : {json_path}")
        sys.exit(1)
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    print(f"✅ {len(data)} partenaires chargés depuis {json_path.name}")
    return data


def build_excel(partners: list, missing_set: set, output_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conducteurs sans carte grise"

    # ── En-têtes ──
    headers = ["Partenaire", "Email", "Conducteur", "Téléphone",
               "Matricule", "Type véhicule", "Marque", "Modèle"]
    ws.append(headers)

    # Style en-tête
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Largeurs colonnes
    widths = [20, 30, 25, 18, 18, 15, 15, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # ── Données ──
    count = 0
    alt_fill = PatternFill("solid", fgColor="F5F5F5")

    for partner in partners:
        partner_name = partner.get("nom", "")
        email = partner.get("email", "")
        drivers = partner.get("drivers", [])

        for driver in drivers:
            vehicle = driver.get("vehicle", {})
            matricule_raw = vehicle.get("matricule", "") or ""
            if not matricule_raw or matricule_raw.upper() == "N/A":
                continue

            norm = normalize_plate(matricule_raw)
            if norm not in missing_set:
                continue  # pas dans la liste manquante → on skip

            count += 1
            row_data = [
                partner_name,
                email,
                driver.get("nom", ""),
                driver.get("telephone", ""),
                matricule_raw,
                vehicle.get("type", ""),
                vehicle.get("marque", ""),
                vehicle.get("modele", ""),
            ]
            ws.append(row_data)

            # Alternance couleur
            if count % 2 == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = alt_fill

    # Freeze header
    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return count


def main():
    print("=" * 60)
    print("🔗 CROISEMENT enriched × missing_plates")
    print("=" * 60)

    missing_set = load_missing_plates(MISSING_JSON)
    partners    = load_enriched(ENRICHED_JSON)

    print(f"\n🔍 Croisement en cours...")
    count = build_excel(partners, missing_set, OUTPUT_XLSX)

    print(f"\n✅ {count} conducteurs trouvés avec matricule dans missing_plates")
    print(f"📊 Fichier généré : {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
