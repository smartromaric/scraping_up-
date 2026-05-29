#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def log(msg: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def normalize_partner_key(value: object) -> str:
    txt = normalize_text(value)
    txt = re.sub(r"[\s\-_]", "", txt)
    txt = txt.replace("partenaires", "partenaire")
    return txt


def normalize_matricule(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(value)).upper()


def canonical_phone(value: object) -> str:
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if not digits:
        return ""
    local = digits[-10:] if len(digits) >= 10 else digits
    if len(local) < 10:
        return ""
    return f"+225{local}"


def build_phone_map_from_datajson(data_root: Path) -> tuple[dict[str, dict[str, str]], list[str]]:
    phone_map: dict[str, dict[str, str]] = defaultdict(dict)
    conflicts: list[str] = []

    for partner_dir in sorted([p for p in data_root.iterdir() if p.is_dir()]):
        partner_key = normalize_partner_key(partner_dir.name)
        json_path = partner_dir / "data.json"
        if not json_path.exists():
            continue

        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as exc:
            conflicts.append(f"{partner_dir.name}: data.json illisible ({exc})")
            continue

        for d in data.get("drivers", []):
            vehicle = d.get("vehicle", {}) or {}
            mat = normalize_matricule(vehicle.get("matricule", ""))
            phone = canonical_phone(d.get("telephone", ""))
            if not mat or not phone:
                continue

            prev = phone_map[partner_key].get(mat)
            if prev and prev != phone:
                conflicts.append(f"{partner_key} | {mat}: {prev} vs {phone} (premier conservé)")
                continue
            phone_map[partner_key][mat] = phone

    return dict(phone_map), conflicts


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Met à jour les numéros dans Excel depuis les data.json par PARTENAIRE + Matricule."
    )
    parser.add_argument(
        "--excel",
        default=r"C:\Users\c.romaric\Desktop\scraping\vps_deploy\Matrice_appel_errone_injoignable.xlsx",
        help="Fichier Excel source.",
    )
    parser.add_argument(
        "--output",
        default=r"C:\Users\c.romaric\Desktop\scraping\vps_deploy\Matrice_appel_errone_injoignable_updated_from_json.xlsx",
        help="Fichier Excel de sortie.",
    )
    parser.add_argument(
        "--data-root",
        default=r"C:\Users\c.romaric\Desktop\scraping\vps_deploy\output\organized_by_partner",
        help="Dossier contenant les data.json partenaires.",
    )
    parser.add_argument("--only", help="Filtrer un partenaire (ex: partenaire1).")
    parser.add_argument("--dry-run", action="store_true", help="Simulation sans sauvegarde.")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    output_path = Path(args.output)
    data_root = Path(args.data_root)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel introuvable: {excel_path}")
    if not data_root.exists():
        raise FileNotFoundError(f"data-root introuvable: {data_root}")

    phone_map_by_partner, conflicts = build_phone_map_from_datajson(data_root)
    if args.only:
        key = normalize_partner_key(args.only)
        phone_map_by_partner = {k: v for k, v in phone_map_by_partner.items() if k == key}

    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    col_partner = None
    col_matricule = None
    col_numero = None
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        values = [normalize_text(c.value) for c in ws[row_idx]]
        if "partenaire" in values and "matricule" in values and "numero" in values:
            header_row = row_idx
            col_partner = values.index("partenaire")
            col_matricule = values.index("matricule")
            col_numero = values.index("numero")
            break

    if header_row is None:
        raise RuntimeError("Colonnes PARTENAIRE/Matricule/Numero introuvables dans le fichier Excel.")

    # Colonnes d'audit ajoutées en fin de tableau
    existing_headers = [ws.cell(row=header_row, column=i).value for i in range(1, ws.max_column + 1)]
    header_index = {normalize_text(v): idx + 1 for idx, v in enumerate(existing_headers) if v is not None}

    def ensure_audit_col(header_name: str) -> int:
        normalized = normalize_text(header_name)
        if normalized in header_index:
            return header_index[normalized]
        col = ws.max_column + 1
        ws.cell(row=header_row, column=col).value = header_name
        header_index[normalized] = col
        return col

    col_numero_avant = ensure_audit_col("numero_avant")
    col_numero_apres = ensure_audit_col("numero_apres")
    col_statut = ensure_audit_col("statut_update")

    scanned = 0
    matched = 0
    updated = 0
    missing = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        partner_raw = ws.cell(row=row_idx, column=col_partner + 1).value
        mat_raw = ws.cell(row=row_idx, column=col_matricule + 1).value
        numero_cell = ws.cell(row=row_idx, column=col_numero + 1)

        partner = normalize_partner_key(partner_raw)
        mat = normalize_matricule(mat_raw)
        if not partner or not mat:
            continue

        if args.only and partner != normalize_partner_key(args.only):
            continue

        scanned += 1
        target_phone = phone_map_by_partner.get(partner, {}).get(mat)
        old_phone_raw = numero_cell.value
        old_phone = canonical_phone(old_phone_raw)

        ws.cell(row=row_idx, column=col_numero_avant).value = old_phone_raw if old_phone_raw is not None else ""

        if not target_phone:
            missing += 1
            ws.cell(row=row_idx, column=col_numero_apres).value = ""
            ws.cell(row=row_idx, column=col_statut).value = "NO_MATCH_JSON"
            continue

        matched += 1
        ws.cell(row=row_idx, column=col_numero_apres).value = target_phone
        if old_phone != target_phone:
            updated += 1
            ws.cell(row=row_idx, column=col_statut).value = "UPDATED"
            if not args.dry_run:
                numero_cell.value = target_phone
        else:
            ws.cell(row=row_idx, column=col_statut).value = "UNCHANGED"

    log(f"[STATS] Lignes scannees: {scanned}")
    log(f"[STATS] Match trouves: {matched}")
    log(f"[STATS] Numeros modifies: {updated}")
    log(f"[STATS] Sans match JSON: {missing}")
    if conflicts:
        log(f"[WARN] Conflits data.json detectes: {len(conflicts)}")

    if args.dry_run:
        log("[DRY-RUN] Aucun fichier sauvegarde")
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log(f"[OK] Fichier genere: {output_path}")


if __name__ == "__main__":
    main()
