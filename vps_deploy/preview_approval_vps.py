"""
Pré-audit approbation carte grise — Rapport sans action (headless)
===========================================================
Analyse combien de véhicules sont prêts pour approbation automatique

Usage:
  python3 preview_approval_vps.py              # Rapport complet tous partenaires
  python3 preview_approval_vps.py --slack    # Envoie aussi sur Slack

Sortie:
  📊 RAPPORT PRÉ-APPROBATION
  Total immatriculations (Excel)   : 150
  Images carte grise trouvées      : 127
  ─────────────────────────────────────────
  Partenaires analysés             : 45
  Véhicules EN ATTENTE total       : 89
  ✅ Prêts pour approbation       : 72 (Excel + Image OK)
  ❌ Bloqués                       : 17
     - Pas dans Excel              : 8
     - Pas d'image                 : 9
"""

import json
import os
import re
import sys
import urllib.request
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# ─── Configuration ──────────────────────────────────────────────────────────
OUTPUT_DIR = Path(__file__).parent / "output"
ORGANIZED_DIR = OUTPUT_DIR / "organized_by_partner"
IMAGES_OCR_DIR = Path(__file__).parent / "images_ocr"
EXCEL_PATH = OUTPUT_DIR / "immatriculations.xlsx"

_ALT_EXCEL_PATHS = [
    Path.home() / "Downloads" / "output" / "immatriculations.xlsx",
    Path(__file__).parent / "immatriculations.xlsx",
]

WEBHOOK_URL = os.getenv("WEBHOOK_URL", "")

# ─── Utilitaires ─────────────────────────────────────────────────────────────

def log(message: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {message}")

def send_slack(message: str, color: str = "#36a64f"):
    if not WEBHOOK_URL:
        return
    try:
        payload = json.dumps({
            "username": "UpJunoo Bot",
            "icon_emoji": ":car:",
            "attachments": [{"color": color, "text": message}]
        }).encode("utf-8")
        req = urllib.request.Request(
            WEBHOOK_URL, data=payload,
            headers={"Content-Type": "application/json"}
        )
        urllib.request.urlopen(req, timeout=10)
    except Exception as e:
        log(f"⚠️ Slack erreur: {e}")

def normalize_plaque(plaque: str) -> str:
    if not plaque:
        return ""
    return plaque.upper().replace(" ", "").replace("-", "").replace("_", "")

def find_excel_path() -> Path:
    if EXCEL_PATH.exists():
        return EXCEL_PATH
    for alt in _ALT_EXCEL_PATHS:
        if alt.exists():
            return alt
    for candidate in Path.home().rglob("immatriculations*.xlsx"):
        return candidate
    return None

# ═══════════════════════════════════════════════════════════════════════════════
#  CHARGEMENT DONNÉES
# ═══════════════════════════════════════════════════════════════════════════════

def load_immatriculations_index(excel_path: Path) -> dict:
    """Charge Excel → index {plaque_norm → données}"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("❌ openpyxl non installé: pip3 install openpyxl")
        sys.exit(1)
    
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    headers = {}
    immat_col = fichier_col = url_col = None
    
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value.lower().strip()] = idx
    
    for key, col in headers.items():
        if 'immatriculation' in key:
            immat_col = col
        elif key in ['nom du fichier', 'fichier', 'nom_fichier', 'filename']:
            fichier_col = col
        elif 'url' in key:
            url_col = col
    
    index = {}
    total_rows = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        plaque = row[immat_col-1] if immat_col else None
        if not plaque:
            continue
        
        plaque_norm = normalize_plaque(str(plaque))
        fichier = row[fichier_col-1] if fichier_col else None
        url = row[url_col-1] if url_col else None
        
        index[plaque_norm] = {
            "plaque": str(plaque).strip(),
            "fichier": str(fichier) if fichier else "",
            "url_image": str(url) if url else ""
        }
    
    wb.close()
    log(f"✅ {len(index)} immatriculations indexées (sur {total_rows} lignes)")
    return index

def load_images_index() -> dict:
    """Scan images_ocr/ → index {nom_fichier_sans_ext → Path}"""
    if not IMAGES_OCR_DIR.exists():
        log(f"⚠️ Dossier images_ocr introuvable: {IMAGES_OCR_DIR}")
        return {}
    
    images = {}
    for ext in ['*.jpeg', '*.jpg', '*.png']:
        for img_path in IMAGES_OCR_DIR.glob(ext):
            key = img_path.stem.lower()
            images[key] = img_path
    
    log(f"✅ {len(images)} images trouvées dans images_ocr/")
    return images

def load_partners_data() -> list:
    """Charge all_partners_enriched.json ou scan organized_by_partner/"""
    json_file = ORGANIZED_DIR / "all_partners_enriched.json"
    
    if json_file.exists():
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        log(f"✅ {len(data)} partenaires chargés depuis JSON")
        return data
    
    # Fallback: scan dossiers
    partners = []
    if ORGANIZED_DIR.exists():
        for partner_dir in ORGANIZED_DIR.iterdir():
            if partner_dir.is_dir() and partner_dir.name != "UNASSIGNED_DRIVERS":
                data_file = partner_dir / "data.json"
                if data_file.exists():
                    with open(data_file, "r", encoding="utf-8") as f:
                        partners.append(json.load(f))
    
    log(f"✅ {len(partners)} partenaires scannés depuis dossiers")
    return partners

# ═══════════════════════════════════════════════════════════════════════════════
#  ANALYSE PRÉ-APPROBATION
# ═══════════════════════════════════════════════════════════════════════════════

def analyze_approval_readiness(immat_index: dict, images_index: dict, partners: list) -> dict:
    """Analyse quels véhicules sont prêts pour approbation"""
    
    stats = {
        "total_partners": len(partners),
        "total_vehicles": 0,
        "in_excel": 0,
        "has_image": 0,
        "ready_for_approval": 0,
        "blocked_no_excel": 0,
        "blocked_no_image": 0,
        "by_partner": []
    }
    
    PARTNER_RE = re.compile(r'^\s*partenaires?[-_]?\s*(\d+)\s*$', re.I)
    
    for partner in partners:
        partner_name = partner.get('nom', 'Unknown')
        
        # Skip non-standard partners
        if not PARTNER_RE.match(partner_name) and partner_name != 'UNASSIGNED_DRIVERS':
            continue
        if partner_name == 'UNASSIGNED_DRIVERS':
            continue
        
        drivers = partner.get('drivers', [])
        
        partner_stats = {
            "name": partner_name,
            "total_drivers": len(drivers),
            "ready": 0,
            "blocked_no_excel": 0,
            "blocked_no_image": 0
        }
        
        for driver in drivers:
            stats["total_vehicles"] += 1
            
            vehicle = driver.get('vehicle', {})
            plaque = vehicle.get('matricule', '')
            if not plaque or plaque == 'N/A':
                continue
            
            plaque_norm = normalize_plaque(plaque)
            
            # Check Excel
            in_excel = plaque_norm in immat_index
            if in_excel:
                stats["in_excel"] += 1
            
            # Check image
            # Cherche par plaque ou par nom de fichier référencé
            has_image = False
            
            # 1. Par plaque dans le nom du fichier image
            for img_key in images_index:
                if plaque_norm in img_key or plaque.replace('-', '').lower() in img_key:
                    has_image = True
                    break
            
            # 2. Par fichier référencé dans Excel
            if in_excel and not has_image:
                ref_file = immat_index[plaque_norm].get('fichier', '')
                if ref_file:
                    ref_key = Path(ref_file).stem.lower()
                    if ref_key in images_index:
                        has_image = True
            
            if has_image:
                stats["has_image"] += 1
            
            # Déterminer statut
            if in_excel and has_image:
                stats["ready_for_approval"] += 1
                partner_stats["ready"] += 1
            elif not in_excel:
                stats["blocked_no_excel"] += 1
                partner_stats["blocked_no_excel"] += 1
            elif not has_image:
                stats["blocked_no_image"] += 1
                partner_stats["blocked_no_image"] += 1
        
        stats["by_partner"].append(partner_stats)
    
    return stats

# ═══════════════════════════════════════════════════════════════════════════════
#  AFFICHAGE RAPPORT
# ═══════════════════════════════════════════════════════════════════════════════

def print_report(stats: dict, immat_index: dict, images_index: dict):
    """Affiche le rapport final"""
    
    print()
    print("=" * 60)
    print("📊 RAPPORT PRÉ-APPROBATION CARTES GRISES")
    print("=" * 60)
    print()
    print("📁 SOURCES DE DONNÉES:")
    print(f"   • Immatriculations (Excel)  : {len(immat_index)} plaques")
    print(f"   • Images carte grise          : {len(images_index)} fichiers")
    print(f"   • Partenaires à traiter       : {stats['total_partners']}")
    print()
    print("🚗 VÉHICULES ANALYSÉS:")
    print(f"   • Total véhicules            : {stats['total_vehicles']}")
    print()
    print("✅ PRÊTS POUR APPROBATION:")
    print(f"   • Prêts (Excel + Image OK)   : {stats['ready_for_approval']}")
    print()
    print("❌ BLOQUÉS:")
    print(f"   • Pas dans l'Excel           : {stats['blocked_no_excel']}")
    print(f"   • Image manquante            : {stats['blocked_no_image']}")
    print()
    print("=" * 60)
    
    # Top 10 partenaires avec le plus de véhicules prêts
    sorted_partners = sorted(
        stats['by_partner'], 
        key=lambda x: x['ready'], 
        reverse=True
    )[:10]
    
    if sorted_partners:
        print("🏆 TOP PARTENAIRES (plus de véhicules prêts):")
        for p in sorted_partners:
            status = "✅" if p['ready'] > 0 else "❌"
            print(f"   {status} {p['name']:<20} | Prêts: {p['ready']:<3} | Bloqués: {p['blocked_no_excel'] + p['blocked_no_image']}")
        print()
    
    # Partenaires bloqués
    blocked_partners = [p for p in stats['by_partner'] if p['ready'] == 0 and p['total_drivers'] > 0][:5]
    if blocked_partners:
        print("⚠️  PARTENAIRES AVEC BLOQUAGES:")
        for p in blocked_partners:
            print(f"   • {p['name']}: {p['blocked_no_excel']} sans Excel, {p['blocked_no_image']} sans image")
        print()
    
    print("=" * 60)
    print(f"💡 Pour approuver: python3 approve_fleet_vps.py")
    print("=" * 60)

def generate_slack_report(stats: dict, immat_index: dict, images_index: dict) -> str:
    """Génère le message Slack"""
    return f"""📊 *Rapport Pré-Approbation Cartes Grises*

📁 *Sources:*
• Excel: {len(immat_index)} plaques
• Images: {len(images_index)} fichiers
• Partenaires: {stats['total_partners']}

🚗 *Analyse:*
• Total véhicules: {stats['total_vehicles']}

✅ *Prêts:* {stats['ready_for_approval']} véhicules

❌ *Bloqués:* {stats['blocked_no_excel'] + stats['blocked_no_image']}
   - Pas dans Excel: {stats['blocked_no_excel']}
   - Image manquante: {stats['blocked_no_image']}

💡 Lancez: `python3 approve_fleet_vps.py`"""

# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Pré-audit approbation carte grise")
    parser.add_argument("--slack", action="store_true", help="Envoie rapport sur Slack")
    args = parser.parse_args()
    
    log("🔍 DÉMARRAGE PRÉ-AUDIT APPROBATION")
    log("=" * 50)
    
    # 1. Charger Excel
    excel_path = find_excel_path()
    if not excel_path:
        log("❌ Fichier Excel immatriculations introuvable!")
        sys.exit(1)
    log(f"📄 Excel trouvé: {excel_path}")
    
    immat_index = load_immatriculations_index(excel_path)
    if not immat_index:
        log("❌ Aucune immatriculation chargée!")
        sys.exit(1)
    
    # 2. Charger images
    images_index = load_images_index()
    
    # 3. Charger partenaires
    partners = load_partners_data()
    if not partners:
        log("❌ Aucun partenaire trouvé!")
        sys.exit(1)
    
    # 4. Analyser
    log("🔍 Analyse des correspondances...")
    stats = analyze_approval_readiness(immat_index, images_index, partners)
    
    # 5. Afficher rapport
    print_report(stats, immat_index, images_index)
    
    # 6. Slack si demandé
    if args.slack and WEBHOOK_URL:
        slack_msg = generate_slack_report(stats, immat_index, images_index)
        send_slack(slack_msg, "#36a64f" if stats['ready_for_approval'] > 0 else "#ff0000")
        log("📤 Rapport envoyé sur Slack")

if __name__ == "__main__":
    main()
