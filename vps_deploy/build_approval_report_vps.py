#!/usr/bin/env python3
"""
build_approval_report_vps.py
Consolide tous les fleet_approval_report_*.json de chaque partenaire
en un seul rapport HTML KPI global.

Usage:
    python3 build_approval_report_vps.py
"""

import json
import os
import re
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

# ─── Chemins ──────────────────────────────────────────────────────────────────
_SCRIPT_DIR   = Path(__file__).parent
_PROJECT_ROOT = _SCRIPT_DIR.parent if _SCRIPT_DIR.name == "vps_deploy" else _SCRIPT_DIR

ORGANIZED_DIR       = _PROJECT_ROOT / "output" / "organized_by_partner"
OUTPUT_HTML         = _PROJECT_ROOT / "output" / "approval_kpi_report_global.html"
OUTPUT_MISSING_JSON = _PROJECT_ROOT / "output" / "missing_plates.json"
OUTPUT_MISSING_XLSX = _PROJECT_ROOT / "output" / "missing_plates.xlsx"

PARTNER_NAME_RE = re.compile(r'^\s*(partenaires?)[-_\s]*(\d+)\s*$', re.I)

WEBHOOK_URL = os.getenv("WEBHOOK_URL", "").strip()


def send_slack(message: str, color: str = "#36a64f"):
    if not WEBHOOK_URL:
        return
    try:
        payload = json.dumps({
            "username": "UpJunoo Bot",
            "icon_emoji": ":bar_chart:",
            "attachments": [{"color": color, "text": message}]
        }).encode("utf-8")
        req = urllib.request.Request(
            WEBHOOK_URL, data=payload,
            headers={"Content-Type": "application/json"}
        )
        urllib.request.urlopen(req, timeout=10)
        print("📨 Slack notifié")
    except Exception as e:
        print(f"⚠️  Slack erreur: {e}")

def extract_partner_number(name: str) -> int:
    m = PARTNER_NAME_RE.match(name or "")
    return int(m.group(2)) if m else 0


def load_all_reports(organized_dir: Path) -> list:
    """Parcourt tous les dossiers partenaire et charge le dernier JSON de chaque."""
    all_results = []
    partner_summaries = []
    skipped = []

    if not organized_dir.exists():
        print(f"❌ Dossier introuvable : {organized_dir}")
        sys.exit(1)

    partner_dirs = sorted(
        [d for d in organized_dir.iterdir() if d.is_dir() and d.name != "UNASSIGNED_DRIVERS"],
        key=lambda d: extract_partner_number(d.name)
    )

    print(f"📂 {len(partner_dirs)} dossiers partenaires trouvés")

    for partner_dir in partner_dirs:
        json_files = sorted(partner_dir.glob("fleet_approval_report_*.json"), reverse=True)
        if not json_files:
            skipped.append(partner_dir.name)
            continue

        latest = json_files[0]  # le plus récent
        try:
            with open(latest, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            print(f"⚠️  Erreur lecture {latest}: {e}")
            skipped.append(partner_dir.name)
            continue

        vehicles = data.get("vehicles", [])
        approved    = sum(1 for v in vehicles if v.get("statut_approbation") == "Oui")
        not_found   = sum(1 for v in vehicles if "non trouvé" in (v.get("raison") or "").lower())
        img_missing = sum(1 for v in vehicles if "image" in (v.get("raison") or "").lower() and v.get("statut_approbation") != "Oui")
        failed      = sum(1 for v in vehicles if v.get("statut_approbation") != "Oui" and "non trouvé" not in (v.get("raison") or "").lower() and "image" not in (v.get("raison") or "").lower() and v.get("raison"))
        skipped_v   = sum(1 for v in vehicles if "déjà" in (v.get("raison") or "").lower() or "skip" in (v.get("raison") or "").lower())

        partner_summaries.append({
            "name": partner_dir.name,
            "json_file": latest.name,
            "generated_at": data.get("generated_at", ""),
            "total": len(vehicles),
            "approved": approved,
            "not_found": not_found,
            "img_missing": img_missing,
            "failed": failed,
            "skipped": skipped_v,
        })
        all_results.extend(vehicles)

    print(f"✅ {len(partner_summaries)} partenaires avec rapport | ⏭️  {len(skipped)} sans rapport")
    if skipped:
        print(f"   Sans rapport : {', '.join(skipped[:10])}{'...' if len(skipped) > 10 else ''}")

    return all_results, partner_summaries


def generate_html(all_results: list, summaries: list, output_path: Path):
    ts = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Stats globales
    total_p     = len(summaries)
    total_v     = sum(s["total"] for s in summaries)
    approved    = sum(s["approved"] for s in summaries)
    not_found   = sum(s["not_found"] for s in summaries)
    img_missing = sum(s["img_missing"] for s in summaries)
    failed      = sum(s["failed"] for s in summaries)
    skipped     = sum(s["skipped"] for s in summaries)
    pct = round(approved / total_v * 100, 1) if total_v > 0 else 0

    # ── Tableau résumé par partenaire ──
    summary_rows = ""
    for s in summaries:
        pct_p = round(s["approved"] / s["total"] * 100, 1) if s["total"] > 0 else 0
        bar = f'<div style="background:#eee;border-radius:6px;height:10px;width:100px;display:inline-block;vertical-align:middle"><div style="background:#27ae60;height:10px;border-radius:6px;width:{pct_p}%"></div></div>'
        summary_rows += f"""
        <tr>
            <td><strong>{s['name']}</strong></td>
            <td style="text-align:center">{s['total']}</td>
            <td style="text-align:center;color:#27ae60"><strong>{s['approved']}</strong></td>
            <td style="text-align:center;color:#e74c3c">{s['not_found']}</td>
            <td style="text-align:center;color:#e67e22">{s['img_missing']}</td>
            <td style="text-align:center;color:#c0392b">{s['failed']}</td>
            <td style="text-align:center">{bar} <small>{pct_p}%</small></td>
            <td style="font-size:11px;color:#aaa">{s['generated_at'][:16] if s['generated_at'] else '-'}</td>
        </tr>"""

    # ── Tableau détail véhicule ──
    detail_rows = ""
    for r in all_results:
        statut = r.get("statut_approbation", "Non")
        raison = r.get("raison", "")
        if statut == "Oui":
            badge = '<span class="badge badge-success">✅ Approuvé</span>'
        elif "non trouvé" in raison.lower():
            badge = '<span class="badge badge-secondary">❌ Non trouvé Excel</span>'
        elif "image" in raison.lower():
            badge = '<span class="badge badge-warning">⚠️ Image manquante</span>'
        elif "déjà" in raison.lower() or "skip" in raison.lower():
            badge = '<span class="badge badge-info">⏭️ Déjà uploadé</span>'
        else:
            badge = f'<span class="badge badge-danger">🔴 Erreur</span>'

        detail_rows += f"""
        <tr>
            <td>{r.get('partenaire', '')}</td>
            <td><code>{r.get('plate_raw', '')}</code></td>
            <td>{r.get('immatriculation_officielle', '-') or '-'}</td>
            <td><small>{r.get('nom_fichier_image', '-') or '-'}</small></td>
            <td>{badge}</td>
            <td><small style="color:#888">{raison}</small></td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Rapport Global Approbation — UpJunoo</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#333}}
.header{{background:linear-gradient(135deg,#1a1a2e 0%,#16213e 50%,#0f3460 100%);color:white;padding:30px 40px}}
.header h1{{font-size:26px;margin-bottom:4px}}
.header p{{opacity:.7;font-size:13px}}
.container{{max-width:1500px;margin:0 auto;padding:24px 16px}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;margin-bottom:24px}}
.kpi-card{{background:white;border-radius:12px;padding:18px;box-shadow:0 2px 8px rgba(0,0,0,.07);text-align:center}}
.kpi-card .val{{font-size:34px;font-weight:700;margin-bottom:2px}}
.kpi-card .lbl{{font-size:12px;color:#888}}
.green .val{{color:#27ae60}}.red .val{{color:#e74c3c}}.orange .val{{color:#e67e22}}
.blue .val{{color:#2980b9}}.gray .val{{color:#95a5a6}}.purple .val{{color:#8e44ad}}
.progress-wrap{{background:white;border-radius:12px;padding:18px;box-shadow:0 2px 8px rgba(0,0,0,.07);margin-bottom:24px}}
.progress-wrap h3{{font-size:14px;margin-bottom:10px;color:#555}}
.bar{{height:22px;background:#eee;border-radius:11px;overflow:hidden}}
.bar-fill{{height:100%;background:linear-gradient(90deg,#27ae60,#2ecc71);border-radius:11px;
  display:flex;align-items:center;justify-content:center;color:white;font-size:12px;font-weight:600}}
.section{{background:white;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,.07);overflow:hidden;margin-bottom:24px}}
.section-header{{padding:16px 20px;border-bottom:1px solid #eee;display:flex;justify-content:space-between;align-items:center}}
.section-header h2{{font-size:16px}}
table{{width:100%;border-collapse:collapse}}
thead th{{background:#f8f9fa;padding:10px 14px;text-align:left;font-size:11px;
  text-transform:uppercase;color:#888;letter-spacing:.4px;border-bottom:2px solid #eee}}
tbody td{{padding:10px 14px;border-bottom:1px solid #f0f0f0;font-size:13px}}
tbody tr:hover{{background:#fafbfc}}
.badge{{padding:3px 9px;border-radius:20px;font-size:11px;font-weight:600;white-space:nowrap}}
.badge-success{{background:#d4edda;color:#155724}}
.badge-danger{{background:#f8d7da;color:#721c24}}
.badge-warning{{background:#fff3cd;color:#856404}}
.badge-secondary{{background:#e2e3e5;color:#383d41}}
.badge-info{{background:#d1ecf1;color:#0c5460}}
.search-input{{padding:7px 12px;border:1px solid #ddd;border-radius:8px;font-size:13px;width:220px}}
.footer{{text-align:center;padding:16px;color:#aaa;font-size:11px}}
code{{background:#f4f4f4;padding:1px 5px;border-radius:4px;font-size:12px}}
</style>
</head>
<body>
<div class="header">
  <h1>🚗 Rapport Global Approbation Carte Grise</h1>
  <p>UpJunoo — Généré le {ts} — {total_p} partenaires consolidés</p>
</div>
<div class="container">

  <!-- KPI Cards -->
  <div class="kpi-grid">
    <div class="kpi-card blue"><div class="val">{total_p}</div><div class="lbl">Partenaires</div></div>
    <div class="kpi-card gray"><div class="val">{total_v}</div><div class="lbl">Véhicules EN ATTENTE</div></div>
    <div class="kpi-card green"><div class="val">{approved}</div><div class="lbl">✅ Approuvés</div></div>
    <div class="kpi-card red"><div class="val">{not_found}</div><div class="lbl">❌ Non trouvé Excel</div></div>
    <div class="kpi-card orange"><div class="val">{img_missing}</div><div class="lbl">⚠️ Image manquante</div></div>
    <div class="kpi-card red"><div class="val">{failed}</div><div class="lbl">🔴 Erreur</div></div>
    <div class="kpi-card purple"><div class="val">{skipped}</div><div class="lbl">⏭️ Déjà uploadé</div></div>
  </div>

  <!-- Barre progression -->
  <div class="progress-wrap">
    <h3>Taux d'approbation global : <strong>{pct}%</strong> ({approved} / {total_v} véhicules)</h3>
    <div class="bar">
      <div class="bar-fill" style="width:{pct}%">{pct}%</div>
    </div>
  </div>

  <!-- Résumé par partenaire -->
  <div class="section">
    <div class="section-header">
      <h2>📋 Résumé par partenaire ({total_p})</h2>
      <input class="search-input" type="text" id="searchPartner" placeholder="Filtrer partenaire..."
             onkeyup="filterTable('summaryTable','searchPartner')">
    </div>
    <table id="summaryTable">
      <thead>
        <tr>
          <th>Partenaire</th><th>Total</th><th>✅ Approuvés</th>
          <th>❌ Non trouvé</th><th>⚠️ Image</th><th>🔴 Erreur</th>
          <th>Progression</th><th>Dernière exécution</th>
        </tr>
      </thead>
      <tbody>{summary_rows}</tbody>
    </table>
  </div>

  <!-- Détail véhicule -->
  <div class="section">
    <div class="section-header">
      <h2>🔍 Détail par véhicule ({len(all_results)})</h2>
      <input class="search-input" type="text" id="searchDetail" placeholder="Rechercher plaque..."
             onkeyup="filterTable('detailTable','searchDetail')">
    </div>
    <table id="detailTable">
      <thead>
        <tr>
          <th>Partenaire</th><th>Matricule flotte</th><th>Matricule officiel</th>
          <th>Image</th><th>Statut</th><th>Raison</th>
        </tr>
      </thead>
      <tbody>{detail_rows}</tbody>
    </table>
  </div>

</div>
<div class="footer">UpJunoo Bot — build_approval_report_vps.py — {ts}</div>
<script>
function filterTable(tableId, inputId) {{
  var val = document.getElementById(inputId).value.toLowerCase();
  document.querySelectorAll('#' + tableId + ' tbody tr').forEach(function(row) {{
    row.style.display = row.textContent.toLowerCase().includes(val) ? '' : 'none';
  }});
}}
</script>
</body>
</html>"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ Rapport HTML généré : {output_path}")


def export_missing_plates(all_results: list):
    """Exporte les plaques non trouvées dans Excel en JSON + xlsx."""
    missing = [
        {"partenaire": r["partenaire"], "matricule": r["plate_raw"]}
        for r in all_results
        if "non trouvé" in (r.get("raison") or "").lower()
    ]

    # ── JSON ──
    OUTPUT_MISSING_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_MISSING_JSON, "w", encoding="utf-8") as f:
        json.dump(missing, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON plaques manquantes : {OUTPUT_MISSING_JSON}  ({len(missing)} entrées)")

    # ── Excel ──
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plaques manquantes"

        # En-têtes
        ws.append(["Partenaire", "Matricule"])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

        # Style en-tête
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1a1a2e")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Données
        for row in missing:
            ws.append([row["partenaire"], row["matricule"]])

        wb.save(OUTPUT_MISSING_XLSX)
        print(f"✅ Excel plaques manquantes : {OUTPUT_MISSING_XLSX}")
    except ImportError:
        print("⚠️  openpyxl non installé — Excel ignoré (pip3 install openpyxl)")

    return len(missing)


def main():
    print("=" * 60)
    print("📊 CONSOLIDATION RAPPORT APPROBATION GLOBAL")
    print("=" * 60)
    all_results, summaries = load_all_reports(ORGANIZED_DIR)

    if not summaries:
        print("❌ Aucun rapport JSON trouvé. Lancez d'abord approve_fleet_vps.py")
        sys.exit(1)

    # Stats rapides
    total_v  = sum(s["total"] for s in summaries)
    approved = sum(s["approved"] for s in summaries)
    print(f"\n📊 RÉSUMÉ GLOBAL")
    print(f"   Partenaires avec rapport : {len(summaries)}")
    print(f"   Total véhicules          : {total_v}")
    print(f"   ✅ Approuvés             : {approved}")
    print(f"   ❌ Non trouvé Excel      : {sum(s['not_found'] for s in summaries)}")
    print(f"   ⚠️  Image manquante      : {sum(s['img_missing'] for s in summaries)}")
    print(f"   🔴 Erreurs               : {sum(s['failed'] for s in summaries)}")
    print()

    generate_html(all_results, summaries, OUTPUT_HTML)
    print(f"\n🎉 Rapport disponible : {OUTPUT_HTML}")

    # ── Export plaques manquantes ──
    nb_missing = export_missing_plates(all_results)

    # Notification Slack
    approved = sum(s["approved"] for s in summaries)
    total_v  = sum(s["total"] for s in summaries)
    pct = round(approved / total_v * 100, 1) if total_v > 0 else 0
    color = "#36a64f" if approved > 0 else "#ffaa00"
    send_slack(
        f"📊 *Rapport Global Approbation Carte Grise*\n"
        f"• Partenaires consolidés : {len(summaries)}\n"
        f"• Total véhicules EN ATTENTE : {total_v}\n"
        f"• ✅ Approuvés : {approved} ({pct}%)\n"
        f"• ❌ Non trouvé Excel : {sum(s['not_found'] for s in summaries)}\n"
        f"• ⚠️ Image manquante : {sum(s['img_missing'] for s in summaries)}\n"
        f"• 🔴 Erreurs : {sum(s['failed'] for s in summaries)}\n"
        f"• 📋 Plaques sans carte grise : {nb_missing}\n"
        f"📄 Rapport : {OUTPUT_HTML}",
        color
    )


if __name__ == "__main__":
    main()
