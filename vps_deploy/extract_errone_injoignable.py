from pathlib import Path
import argparse
import unicodedata

from openpyxl import Workbook, load_workbook


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def should_keep_status(appel_value: object) -> bool:
    normalized = normalize_text(appel_value)
    return "injoignable" in normalized or "erron" in normalized


def extract_rows(input_file: Path, output_file: Path) -> tuple[int, int]:
    wb = load_workbook(input_file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = 3
    headers = [cell.value for cell in ws[header_row]]
    if "Appel" not in headers:
        raise ValueError("Colonne 'Appel' introuvable dans la feuille.")
    appel_col_idx = headers.index("Appel")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title

    # Conserver les lignes de contexte (titre + en-tete)
    for row in ws.iter_rows(min_row=1, max_row=header_row, values_only=True):
        out_ws.append(list(row))

    total_data_rows = 0
    kept_rows = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(cell is None for cell in row):
            continue
        total_data_rows += 1
        appel_value = row[appel_col_idx]
        if should_keep_status(appel_value):
            out_ws.append(list(row))
            kept_rows += 1

    output_file.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_file)
    return total_data_rows, kept_rows


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extrait les lignes avec Appel = Injoignable ou Numero errone."
    )
    parser.add_argument(
        "--input",
        default=r"C:\Users\c.romaric\Desktop\scraping\vps_deploy\Matrice de suivi opérationnel — Chauffeur _ Livreur.xlsx",
        help="Chemin du fichier Excel source.",
    )
    parser.add_argument(
        "--output",
        default=r"C:\Users\c.romaric\Desktop\scraping\vps_deploy\Matrice_appel_errone_injoignable.xlsx",
        help="Chemin du fichier Excel de sortie.",
    )
    args = parser.parse_args()

    input_file = Path(args.input)
    output_file = Path(args.output)

    if not input_file.exists():
        raise FileNotFoundError(f"Fichier introuvable: {input_file}")

    total_data_rows, kept_rows = extract_rows(input_file, output_file)
    print(f"Fichier cree: {output_file}")
    print(f"Lignes analysees: {total_data_rows}")
    print(f"Lignes conservees (Injoignable/Errone): {kept_rows}")


if __name__ == "__main__":
    main()
