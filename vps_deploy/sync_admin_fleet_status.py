import argparse
import csv
import json
import os
import re
import sys
import time
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait

# Encodage UTF-8 Windows
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ─── Config ───────────────────────────────────────────────────────────────────
BASE_URL             = "https://upjunoo-server-new.junooapps.com"
ADMIN_LOGIN_URL      = f"{BASE_URL}/login/admin"
MANAGE_OWNERS_URL    = f"{BASE_URL}/manage-owners"

_SCRIPT_DIR          = Path(__file__).parent
OUTPUT_DIR           = _SCRIPT_DIR / "output"
ORGANIZED_DIR        = OUTPUT_DIR / "organized_by_partner"
PARTNER_DETAILS_CSV  = OUTPUT_DIR / "partner_drivers_details.csv"
PARTNER_DETAILS_XLSX = OUTPUT_DIR / "partner_drivers_details.xlsx"
CSV_FIELDS = [
    "nom_partenaire",
    "email",
    "chauffeur_nom",
    "chauffeur_telephone",
    "emplacement",
    "type_transport",
    "type_vehicule",
    "assignment_status",
    "vehicle_type",
    "vehicle_marque",
    "vehicle_modele",
    "vehicle_matricule",
]

# REMPLACEZ PAR VOS IDENTIFIANTS ADMIN
ADMIN_EMAIL    = "admin@upjunoo.com"
ADMIN_PASSWORD = "123456789"

def log(msg: str, level: str = "INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    # Couleurs basiques pour la visibilité
    prefix = f"[{ts}] [{level}]"
    if level == "OK": prefix = f"\033[92m{prefix}\033[0m"
    elif level == "WARNING": prefix = f"\033[93m{prefix}\033[0m"
    elif level == "ERROR": prefix = f"\033[91m{prefix}\033[0m"
    print(f"{prefix} {msg}", flush=True)

def normalize_immat(text: str) -> str:
    if not text: return ""
    return re.sub(r'[^a-zA-Z0-9]', '', text).upper()

def make_driver(headless: bool = True) -> webdriver.Chrome:
    opts = Options()
    if headless: opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1920,1080")
    try:
        return webdriver.Chrome(options=opts)
    except:
        svc = Service(shutil.which("chromedriver.exe") or "chromedriver.exe")
        return webdriver.Chrome(service=svc, options=opts)

def wait_for_table_data(driver, timeout=30, min_rows=1):
    """Attend que le tableau contienne au moins min_rows lignes réelles."""
    start = time.time()
    ignore_mots = ["chargement", "loading", "aucune", "no data", "vide", "en attente"]
    while (time.time() - start) < timeout:
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            if len(rows) >= min_rows:
                txt = rows[0].text.lower()
                if not any(m in txt for m in ignore_mots):
                    return True
            time.sleep(1)
        except:
            time.sleep(1)
    return False

def wait_table_row_count_stable(driver, checks=3, pause_s=2):
    """
    Vérifie que le nombre de lignes du tableau reste stable sur plusieurs checks.
    Utile après changement de pagination (ex: 500 entrées).
    """
    last_count = -1
    stable_hits = 0
    for _ in range(checks * 3):
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            count = len(rows)
            if count == last_count and count > 0:
                stable_hits += 1
                if stable_hits >= checks:
                    return count
            else:
                stable_hits = 0
            last_count = count
        except:
            stable_hits = 0
        time.sleep(pause_s)
    return last_count

def find_row_by_cell_text_with_pagination(driver, target_text: str, max_pages: int = 25):
    """
    Cherche une ligne contenant target_text dans une cellule (td) en parcourant les pages DataTables.
    Retourne l'élément <tr> trouvé, ou None.
    """
    q = (target_text or "").strip().lower()
    if not q:
        return None

    next_btn_xpath = "//li[contains(@class, 'next') and not(contains(@class, 'disabled'))]/a"

    for _page in range(max_pages):
        if not wait_for_table_data(driver, timeout=15, min_rows=1):
            break

        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        for r in rows:
            if not r.is_displayed():
                continue
            try:
                cells = r.find_elements(By.TAG_NAME, "td")
                for c in cells:
                    if q in (c.text or "").strip().lower():
                        return r
            except:
                continue

        # Si non trouvé, tenter "next" DataTables
        try:
            next_btn = driver.find_element(By.XPATH, next_btn_xpath)
            if next_btn and next_btn.is_displayed():
                driver.execute_script("arguments[0].click();", next_btn)
                time.sleep(3)
                continue
        except:
            pass

        break

    return None

def set_page_size(driver, size=500):
    try:
        time.sleep(3)
        selectors = ["select.form-select-sm", "select[name*='_length']", ".dataTables_length select", "select.form-select"]
        element = None
        for sel in selectors:
            try:
                element = driver.find_element(By.CSS_SELECTOR, sel)
                if element: break
            except: continue
        
        if element:
            s = Select(element); target = str(size)
            current_val = ""
            try:
                current_val = s.first_selected_option.get_attribute("value")
            except: pass
            
            if current_val != target:
                highlight(driver, element, "orange")
                # Vérifier si l'option existe avant de la sélectionner
                options = [o.get_attribute("value") for o in s.options]
                if target in options:
                    s.select_by_value(target)
                else:
                    # Sélectionner la plus grande disponible
                    s.select_by_index(len(s.options) - 1)
                
                log(f"   [PAGE] Passage à {target} (ou max)... Attente de chargement...")
                time.sleep(5)
                wait_for_table_data(driver, 60, min_rows=2)
                return True
        return False
    except Exception as e:
        log(f"   [PAGE] Erreur : {e}", "WARNING")
        return False

def highlight(driver, element, color="blue", duration=1):
    try:
        driver.execute_script(f"arguments[0].style.border='4px solid {color}';", element)
        if duration > 0: time.sleep(duration)
    except: pass

def build_partner_csv_rows(partner_name: str, partner_email: str, p_data: dict):
    rows = []
    items = p_data.get("drivers") or p_data.get("data") or []
    for d in items:
        v = d.get("vehicle") or {}
        rows.append({
            "nom_partenaire": partner_name or "",
            "email": partner_email or "",
            "chauffeur_nom": d.get("nom") or d.get("chauffeur_nom") or "",
            "chauffeur_telephone": d.get("telephone") or d.get("chauffeur_telephone") or "",
            "emplacement": d.get("emplacement") or "",
            "type_transport": d.get("type_transport") or "",
            "type_vehicule": v.get("statut_flotte") or d.get("type_vehicule") or "",
            "assignment_status": d.get("assignment_status") or "",
            "vehicle_type": v.get("type") or d.get("vehicle_type") or "",
            "vehicle_marque": v.get("marque") or d.get("vehicle_marque") or "",
            "vehicle_modele": v.get("modele") or d.get("vehicle_modele") or "",
            "vehicle_matricule": v.get("matricule") or d.get("vehicle_matricule") or d.get("immatriculation") or "",
        })
    return rows

def upsert_partner_rows_in_csv(csv_path: Path, partner_email: str, new_rows: list):
    existing_rows = []
    if csv_path.exists():
        with open(csv_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                existing_rows.append({field: row.get(field, "") for field in CSV_FIELDS})

    filtered_rows = [
        r for r in existing_rows
        if (r.get("email") or "").strip().lower() != (partner_email or "").strip().lower()
    ]
    filtered_rows.extend({field: r.get(field, "") for field in CSV_FIELDS} for r in new_rows)

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(filtered_rows)

def upsert_partner_rows_in_xlsx(xlsx_path: Path, partner_email: str, sheet_name: str, new_rows: list):
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        existing_rows = []
        row_iter = ws.iter_rows(min_row=2, values_only=True)
        for values in row_iter:
            row = {}
            for idx, field in enumerate(CSV_FIELDS):
                row[field] = values[idx] if idx < len(values) and values[idx] is not None else ""
            existing_rows.append(row)
        wb.remove(ws)
    else:
        existing_rows = []

    filtered_rows = [
        r for r in existing_rows
        if (r.get("email") or "").strip().lower() != (partner_email or "").strip().lower()
    ]
    filtered_rows.extend({field: r.get(field, "") for field in CSV_FIELDS} for r in new_rows)

    ws = wb.create_sheet(title=sheet_name)
    ws.append(CSV_FIELDS)
    for row in filtered_rows:
        ws.append([row.get(field, "") for field in CSV_FIELDS])

    wb.save(xlsx_path)

# ─── Main Logic ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--range", help="Partner range (e.g. 1-10)")
    parser.add_argument("--only", help="Specific partner name")
    parser.add_argument("--no-headless", action="store_true")
    args = parser.parse_args()

    driver = make_driver(headless=not args.no_headless)
    wait = WebDriverWait(driver, 20)
    total_corrections = 0

    try:
        # 1. Connexion Admin
        log(f"[ADMIN] Connexion à {ADMIN_EMAIL}...")
        driver.get(ADMIN_LOGIN_URL)
        try:
            e = wait.until(EC.presence_of_element_located((By.ID, "email-input")))
            e.send_keys(ADMIN_EMAIL)
            log("   [LOGIN] Email saisi.")
            p = driver.find_element(By.ID, "password-input")
            p.send_keys(ADMIN_PASSWORD)
            log("   [LOGIN] Password saisi.")
            btn = driver.find_element(By.XPATH, "//button[@type='submit']")
            highlight(driver, btn, "green")
            btn.click()
            log("   [LOGIN] Clic sur Submit.")
            wait.until(lambda d: "login" not in d.current_url.lower())
            log("[OK] Connexion Admin réussie.")
        except Exception as login_err:
            log(f"[ERROR] Échec de la connexion : {login_err}", "ERROR")
            raise login_err

        # 2. Liste des partenaires
        partners = [
            d for d in ORGANIZED_DIR.iterdir()
            if d.is_dir() and d.name.upper() != "UNASSIGNED_DRIVERS"
        ]
        partners.sort(key=lambda d: int(re.search(r'\d+', d.name).group()) if re.search(r'\d+', d.name) else 0)

        if args.only:
            partners = [p for p in partners if p.name.lower() == args.only.lower()]
        elif args.range:
            s, e = map(int, args.range.split('-'))
            partners = [p for p in partners if s <= (int(re.search(r'\d+', p.name).group()) if re.search(r'\d+', p.name) else 0) <= e]

        log(f"[INFO] {len(partners)} partenaires à vérifier.")

        for p_dir in partners:
            try:
                data_path = p_dir / "data.json"
                if not data_path.exists(): continue
                with open(data_path, "r", encoding="utf-8") as f: p_data = json.load(f)
                
                p_email = p_data.get("email") or p_data.get("partner_email")
                if not p_email: continue

                log(f"\n[SYNC] Partenaire : {p_dir.name} ({p_email})")

                # 3. Navigation manage-owners
                driver.get(MANAGE_OWNERS_URL)
                time.sleep(6)
                wait_for_table_data(driver, 30)
                set_page_size(driver, 500)
                # Stabiliser après changement du nombre d'entrées affichées
                wait_for_table_data(driver, 45, min_rows=1)
                log("   [PAGE] Attente de stabilisation après passage à 500...")
                time.sleep(8)
                stable_count = wait_table_row_count_stable(driver, checks=3, pause_s=2)
                log(f"   [PAGE] Tableau stabilisé (~{stable_count} lignes visibles).")
                
                target_row = None
                try:
                    log(f"   🔎 Recherche de {p_email}...")
                    # 1) Optionnel: si la barre de recherche existe, essayer de filtrer
                    try:
                        search_box = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located(
                                (By.CSS_SELECTOR, "input[type='search'], .dataTables_filter input")
                            )
                        )
                        search_box.clear()
                        search_box.send_keys(p_email)
                        time.sleep(3)
                    except:
                        log("   ⚠️ Barre de recherche introuvable, scan tableau...", "WARNING")

                    # 2) Scan robuste (avec pagination) pour ne pas dépendre d'un état DataTables
                    # (comme dans partner_auto_assign_v3.py: on continue si non trouvé)
                    target_row = find_row_by_cell_text_with_pagination(driver, p_email, max_pages=25)

                    # 3) Si on ne trouve pas par email, tenter via le nom du partenaire (souvent ce qui est affiché)
                    if not target_row:
                        alt_name = p_dir.name.replace("_", " ").replace("-", " ")
                        log(f"   🔁 Email introuvable dans le tableau, essai sur le nom: {alt_name}...")
                        target_row = find_row_by_cell_text_with_pagination(driver, alt_name, max_pages=25)

                    if not target_row:
                        log(f"   [WARN] Impossible de localiser {p_email}", "WARNING")
                        continue
                    
                    highlight(driver, target_row, "blue")
                    profile_btn = None
                    try:
                        # Chercher d'abord un lien contenant 'profile' ou 'view' ou 'edit'
                        profile_btn = target_row.find_element(By.XPATH, ".//a[contains(@href, 'profile') or contains(@href, 'view') or contains(@href, 'edit')]")
                    except:
                        # Sinon chercher tout bouton ou lien avec une icône
                        btns = target_row.find_elements(By.CSS_SELECTOR, "a.btn, button.btn, a i, button i")
                        if btns:
                            # Souvent le bouton de profil est le premier ou le dernier des boutons d'action
                            profile_btn = btns[-1]
                    
                    if not profile_btn:
                        log(f"   ⚠️ Bouton profil introuvable", "WARNING")
                        continue

                    highlight(driver, profile_btn, "green")
                    log(f"   [CLICK] Ouverture du profil...")
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", profile_btn)
                    time.sleep(1)
                    old_url = driver.current_url
                    driver.execute_script("arguments[0].click();", profile_btn)
                    
                    # Attendre que la navigation profil démarre vraiment avant de chercher les onglets
                    try:
                        WebDriverWait(driver, 20).until(
                            lambda d: (
                                d.current_url != old_url
                                or "profile" in d.current_url.lower()
                                or len(d.find_elements(By.XPATH, "//a[contains(., 'flotte') or contains(., 'Flotte') or contains(., 'Détails de la flotte')]")) > 0
                            )
                        )
                        log(f"   [NAV] Transition après clic profil détectée (URL: {driver.current_url})")
                    except:
                        log(f"   [WARN] Transition lente après clic profil (URL actuelle: {driver.current_url})", "WARNING")
                    
                    # Buffer supplémentaire pour laisser charger les infos du profil
                    time.sleep(30)
                except Exception as e:
                    log(f"   ❌ Erreur profil : {e} (URL actuelle: {driver.current_url})", "ERROR"); continue

                # 4. Onglet Flotte
                try:
                    log("   [INFO] Recherche de l'onglet 'Détails de la flotte'...")
                    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                    fleet_tab = None
                    tab_selectors = [
                        "//a[contains(text(), 'flotte') or contains(text(), 'Flotte')]",
                        "//span[contains(text(), 'flotte') or contains(text(), 'Flotte')]",
                        "//a[contains(., 'Détails de la flotte')]",
                        "//button[contains(., 'Flotte')]"
                    ]
                    for sel in tab_selectors:
                        try:
                            els = driver.find_elements(By.XPATH, sel)
                            for el in els:
                                if el.is_displayed():
                                    fleet_tab = el; break
                            if fleet_tab: break
                        except: continue
                    
                    if not fleet_tab:
                        log("   ⚠️ Onglet flotte introuvable.", "WARNING")
                        continue
                    
                    highlight(driver, fleet_tab, "green")
                    log(f"   [CLICK] Passage à la flotte ({fleet_tab.text or 'onglet'})...")
                    driver.execute_script("arguments[0].click();", fleet_tab)
                    time.sleep(5)
                    set_page_size(driver, 500)
                except Exception as e:
                    log(f"   ❌ Erreur onglet flotte : {e}", "ERROR"); continue

                # 5. Audit de la flotte
                try:
                    wait_for_table_data(driver, 30)
                    rows_fleet = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
                    log(f"   🔎 Scan de {len(rows_fleet)} véhicules...")
                    
                    corrections_p = 0
                    missing_on_site_immats = set()
                    items = p_data.get("drivers") or p_data.get("data") or []
                    
                    # Créer un dictionnaire local pour un accès rapide par immat normalisée
                    local_map = {}
                    for d in items:
                        v = d.get("vehicle", {})
                        immat = normalize_immat(v.get("matricule") or d.get("vehicle_matricule") or d.get("immatriculation"))
                        if immat: local_map[immat] = d

                    for rf in rows_fleet:
                        if not rf.is_displayed(): continue
                        cells = rf.find_elements(By.TAG_NAME, "td")
                        if len(cells) < 5: continue
                        
                        # Détection dynamique des colonnes
                        txt_cells = [c.text.strip() for c in cells]
                        
                        # Recherche de l'immat (colonne 3 ou 4 généralement)
                        site_immat = ""
                        for idx in [4, 3, 2]:
                            if idx < len(txt_cells) and len(txt_cells[idx]) >= 3:
                                test_im = normalize_immat(txt_cells[idx])
                                if test_im and any(c.isdigit() for c in test_im):
                                    site_immat = test_im; break
                        
                        if not site_immat: continue
                        
                        # Statut sur le site
                        is_approved_site = "APPROUV" in rf.text.upper()
                        
                        # Chauffeur sur le site
                        site_driver = ""
                        status_kw = ["APPROUVE", "ATTENTE", "REJETE", "PENDING", "APPROVED", "REJECTED", "MODIFIER"]
                        # On cherche le chauffeur dans les colonnes après le statut
                        for idx in range(len(txt_cells)-1, 4, -1):
                            t = txt_cells[idx]
                            if t and t != "-" and len(t) > 2 and not any(kw in t.upper() for kw in status_kw):
                                site_driver = t; break

                        # Synchronisation avec le JSON local
                        d = local_map.get(site_immat)
                        if d:
                            v = d.get("vehicle", {})
                            changed = False
                            
                            # 1. Synchronisation du Statut d'approbation
                            current_local_approved = (v.get("statut_flotte") == "APPROUVÉ" or d.get("admin_approval_status") == "approved")
                            
                            if is_approved_site and not current_local_approved:
                                log(f"   📝 [STATUT] {site_immat} : EN ATTENTE -> APPROUVÉ", "OK")
                                v["statut_flotte"] = "APPROUVÉ"
                                d["admin_approval_status"] = "approved"
                                d["admin_approved_at"] = datetime.now().isoformat()
                                changed = True
                            elif not is_approved_site and current_local_approved:
                                log(f"   📝 [STATUT] {site_immat} : APPROUVÉ -> EN ATTENTE (Site dit Non Approuvé)", "WARNING")
                                v["statut_flotte"] = "EN ATTENTE"
                                d["admin_approval_status"] = "pending"
                                changed = True

                            # 2. Synchronisation de l'Affectation (Driver)
                            has_driver_site = (site_driver != "")
                            local_done = (d.get("assignment_status") == "DONE")

                            if is_approved_site:
                                if has_driver_site and not local_done:
                                    log(f"   📝 [ASSIGN] {site_immat} : -> DONE (Chauffeur détecté: {site_driver})", "OK")
                                    d["assignment_status"] = "DONE"
                                    changed = True
                                elif not has_driver_site and local_done:
                                    log(f"   🚨 [CONFLIT] {site_immat} : DONE -> VIDE (Chauffeur absent sur Admin)", "WARNING")
                                    missing_on_site_immats.add(site_immat)
                                    if "assignment_status" in d: del d["assignment_status"]
                                    changed = True
                            
                            if changed:
                                highlight(driver, rf, "orange")
                                corrections_p += 1
                                total_corrections += 1
                    
                    partner_rows = build_partner_csv_rows(p_dir.name, p_email, p_data)
                    partner_rows_with_assignment = [
                        row for row in partner_rows
                        if (row.get("assignment_status") or "").strip()
                    ]
                    partner_rows_missing_on_site = [
                        row for row in partner_rows
                        if normalize_immat(row.get("vehicle_matricule") or "") in missing_on_site_immats
                    ]

                    upsert_partner_rows_in_xlsx(
                        PARTNER_DETAILS_XLSX,
                        p_email,
                        "assigned_drivers",
                        partner_rows_with_assignment
                    )
                    upsert_partner_rows_in_xlsx(
                        PARTNER_DETAILS_XLSX,
                        p_email,
                        "missing_on_site",
                        partner_rows_missing_on_site
                    )
                    if corrections_p > 0:
                        log(
                            f"   ✅ {corrections_p} synchronisations effectuées "
                            f"(assignés: {len(partner_rows_with_assignment)}, absents site: {len(partner_rows_missing_on_site)}, fichier: {PARTNER_DETAILS_XLSX.name}).",
                            "OK"
                        )
                    else:
                        log(
                            f"   ✨ Tout est synchronisé "
                            f"(assignés: {len(partner_rows_with_assignment)}, absents site: {len(partner_rows_missing_on_site)}, fichier: {PARTNER_DETAILS_XLSX.name}).",
                            "OK"
                        )
                except Exception as e:
                    log(f"   ❌ Erreur audit : {e}", "ERROR")

            except Exception as e:
                log(f"❌ Erreur sur {p_dir.name}: {e}")

        log("\n" + "="*50)
        log(f"🏁 BILAN DE LA SYNCHRONISATION")
        log(f"📊 Partenaires vérifiés : {len(partners)}")
        log(f"🛠️  Corrections effectuées : {total_corrections}")
        log("="*50)

    except Exception as e:
        log(f"💥 Erreur critique : {e}", "ERROR")
    finally:
        driver.quit()

if __name__ == "__main__": main()
