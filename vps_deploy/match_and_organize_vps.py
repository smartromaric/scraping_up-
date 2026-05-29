"""
Script VPS - Matching et organisation des données Partenaires + Conducteurs (Mode headless)
============================================================================================
1. Charge partenaires.json + conducteurs_vehicles.json
2. Fait un matching intelligent
3. Crée des dossiers par partenaire
4. Exporte JSON, CSV, HTML par partenaire
5. Notification Slack à la fin
"""

import json
import os
import re
import traceback
import urllib.request
from pathlib import Path
from datetime import datetime

# ─── Configuration ──────────────────────────────────────────────────────────────
OUTPUT_DIR = Path(__file__).parent / "output"
PARTNERS_FILE = OUTPUT_DIR / "partenaires.json"
DRIVERS_FILE = OUTPUT_DIR / "conducteurs_vehicles.json"
ORGANIZED_DIR = OUTPUT_DIR / "organized_by_partner"
LOG_FILE = OUTPUT_DIR / "match_organize.log"
WEBHOOK_URL = os.getenv("WEBHOOK_URL", "")

# Filtre : Partenaire[s]?-?N avec N >= 1 (sans limite haute)
PARTNER_NAME_RE = re.compile(r'^\s*partenaires?-?\s*(\d+)\s*$', re.I)
PARTNER_MIN = 1


def log(message):
    """Log avec timestamp"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    full_msg = f"[{timestamp}] {message}"
    print(full_msg)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(full_msg + "\n")


def send_slack(message, color="#36a64f"):
    """Envoie une notification Slack via webhook."""
    if not WEBHOOK_URL:
        return
    try:
        payload = json.dumps({
            "username": os.getenv("SLACK_BOT_NAME", "UpJunoo Bot"),
            "icon_emoji": os.getenv("SLACK_ICON_EMOJI", ":car:"),
            "attachments": [{"color": color, "text": message}]
        }).encode("utf-8")
        req = urllib.request.Request(WEBHOOK_URL, data=payload, headers={"Content-Type": "application/json"})
        urllib.request.urlopen(req, timeout=10)
    except Exception as e:
        log(f"⚠️ Slack erreur: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
#  LOADING DATA
# ═══════════════════════════════════════════════════════════════════════════════

def load_json(file_path):
    """Charge un fichier JSON."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log(f"❌ Erreur chargement {file_path}: {e}")
        return []

def sanitize_folder_name(name):
    """Convertit un nom en dossier valide."""
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    name = re.sub(r'\s+', '_', name)
    name = name[:100]
    return name


# ═══════════════════════════════════════════════════════════════════════════════
#  MATCHING LOGIC
# ═══════════════════════════════════════════════════════════════════════════════

def normalize_text(text):
    """Normalise un texte pour le matching."""
    return text.lower().strip()

def calculate_similarity(str1, str2):
    """Calcule un score de similarité simple entre deux chaînes."""
    str1_norm = normalize_text(str1)
    str2_norm = normalize_text(str2)
    
    # Exact match
    if str1_norm == str2_norm:
        return 100
    
    # Contient
    if str1_norm in str2_norm or str2_norm in str1_norm:
        return 80
    
    # Mots communs
    words1 = set(str1_norm.split())
    words2 = set(str2_norm.split())
    common_words = len(words1 & words2)
    if common_words > 0:
        similarity = (common_words * 2) / (len(words1) + len(words2)) * 100
        return similarity
    
    return 0

def match_drivers_to_partners(partners, drivers):
    """
    Fait un matching entre partenaires et chauffeurs.
    Retourne une liste enrichie de partenaires avec leurs chauffeurs.
    """
    log("🔍 Matching des chauffeurs avec les partenaires...")
    
    # Créer un dictionnaire pour tracer les chauffeurs assignés
    assigned_drivers = set()
    
    for partner in partners:
        partner['drivers'] = partner.get('drivers', [])
        partner_name = partner.get('nom', '')
        
        # Si pas de chauffeurs vides, on cherche
        if not partner['drivers']:
            matching_drivers = []
            
            for driver in drivers:
                driver_name = driver.get('nom', '')
                
                # Chercher une correspondance
                similarity = calculate_similarity(partner_name, driver_name)
                
                if similarity > 50:  # Seuil de similarité
                    matching_drivers.append({
                        'driver': driver,
                        'similarity': similarity
                    })
            
            # Trier par similarité et garder les meilleures correspondances
            matching_drivers.sort(key=lambda x: x['similarity'], reverse=True)
            
            # Ajouter les chauffeurs matchés
            for item in matching_drivers:
                if item['driver']['nom'] not in assigned_drivers:
                    # Ensure driver has vehicle object
                    driver = item['driver'].copy()
                    if 'vehicle' not in driver:
                        driver['vehicle'] = {
                            "type": "N/A",
                            "marque": "N/A",
                            "modele": "N/A",
                            "matricule": "N/A"
                        }
                    partner['drivers'].append(driver)
                    assigned_drivers.add(driver['nom'])
        else:
            # Marquer les chauffeurs existants comme assignés
            for driver in partner['drivers']:
                assigned_drivers.add(driver.get('nom', ''))
    
    # Ajouter les chauffeurs non-assignés à un partenaire "Unassigned"
    unassigned_drivers = [d for d in drivers if d.get('nom') not in assigned_drivers]
    if unassigned_drivers:
        # Ensure all unassigned drivers have vehicle object
        for driver in unassigned_drivers:
            if 'vehicle' not in driver:
                driver['vehicle'] = {
                    "type": "N/A",
                    "marque": "N/A",
                    "modele": "N/A",
                    "matricule": "N/A"
                }
        
        partners.append({
            'nom': 'UNASSIGNED_DRIVERS',
            'email': 'N/A',
            'telephone': 'N/A',
            'document_url': 'N/A',
            'owner_id': 'N/A',
            'drivers': unassigned_drivers
        })
    
    # Ensure all drivers have vehicle object
    for partner in partners:
        for driver in partner.get('drivers', []):
            if 'vehicle' not in driver:
                driver['vehicle'] = {
                    "type": "N/A",
                    "marque": "N/A",
                    "modele": "N/A",
                    "matricule": "N/A"
                }
    
    return partners


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPORT FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def export_partner_data(partner, partner_dir):
    """Exporte les données d'un partenaire dans JSON, CSV et HTML."""
    
    # JSON Export (already includes all data)
    json_path = partner_dir / "data.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(partner, f, ensure_ascii=False, indent=2)
    
    # Excel Export
    from openpyxl import Workbook
    xlsx_path = partner_dir / "data.xlsx"
    wb = Workbook()
    
    # Feuille Partenaire
    ws_partner = wb.active
    ws_partner.title = "Partenaire"
    ws_partner.append(["Nom", "Email", "Téléphone", "URL Document", "Owner ID"])
    ws_partner.append([
        partner.get('nom', 'N/A'),
        partner.get('email', 'N/A'),
        partner.get('telephone', 'N/A'),
        partner.get('document_url', 'N/A'),
        partner.get('owner_id', 'N/A')
    ])
    
    # Feuille Conducteurs
    if partner.get('drivers'):
        ws_drivers = wb.create_sheet("Conducteurs")
        ws_drivers.append(["Nom", "Emplacement", "Téléphone", "Type Transport", "Type Véhicule", "Marque", "Modèle", "Matricule", "URL Document"])
        
        for driver in partner.get('drivers', []):
            vehicle = driver.get('vehicle', {})
            ws_drivers.append([
                driver.get('nom', 'N/A'),
                driver.get('emplacement', 'N/A'),
                driver.get('telephone', 'N/A'),
                driver.get('type_transport', 'N/A'),
                vehicle.get('type', 'N/A'),
                vehicle.get('marque', 'N/A'),
                vehicle.get('modele', 'N/A'),
                vehicle.get('matricule', 'N/A'),
                driver.get('document_url', 'N/A')
            ])
    
    wb.save(xlsx_path)
    
    # HTML Export - Updated to show all driver information
    html_path = partner_dir / "data.html"
    
    drivers_html = ""
    if partner.get('drivers'):
        drivers_html = "".join([
            f'''
            <tr>
                <td>{d.get("nom", "N/A")}</td>
                <td>{d.get("emplacement", "N/A")}</td>
                <td>{d.get("telephone", "N/A")}</td>
                <td>{d.get("type_transport", "N/A")}</td>
                <td>{d.get("type_vehicule", "N/A")}</td>
                <td>
                    <strong>Type:</strong> {d.get("vehicle", {}).get("type", "N/A")}<br>
                    <strong>Marque:</strong> {d.get("vehicle", {}).get("marque", "N/A")}<br>
                    <strong>Modèle:</strong> {d.get("vehicle", {}).get("modele", "N/A")}<br>
                    <strong>Matricule:</strong> {d.get("vehicle", {}).get("matricule", "N/A")}
                </td>
                <td><a href="{d.get("document_url", "#")}" target="_blank">📄 Document</a></td>
            </tr>
            '''
            for d in partner.get('drivers', [])
        ])
    
    html_content = f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <title>{partner.get('nom', 'Partner')} - Data</title>
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                margin: 0;
                padding: 20px;
                color: #333;
            }}
            .container {{
                max-width: 1200px;
                margin: 0 auto;
            }}
            .section {{
                background: white;
                padding: 30px;
                margin: 20px 0;
                border-radius: 10px;
                box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            }}
            h1 {{
                color: #667eea;
                margin-bottom: 10px;
                font-size: 2.2em;
                text-align: center;
            }}
            h2 {{
                color: #555;
                margin-top: 30px;
                margin-bottom: 20px;
                border-bottom: 2px solid #667eea;
                padding-bottom: 10px;
            }}
            .partner-info {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                gap: 20px;
                margin: 20px 0;
            }}
            .info-card {{
                background: #f8f9fa;
                padding: 20px;
                border-radius: 8px;
                border-left: 4px solid #667eea;
            }}
            .info-label {{
                font-weight: bold;
                color: #667eea;
                margin-bottom: 5px;
            }}
            .info-value {{
                color: #333;
                word-break: break-all;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
                background: white;
                border-radius: 8px;
                overflow: hidden;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            }}
            th, td {{
                padding: 15px;
                text-align: left;
                border-bottom: 1px solid #ddd;
            }}
            th {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                font-weight: bold;
                font-size: 0.9em;
            }}
            tr:nth-child(even) {{
                background: #f8f9fa;
            }}
            tr:hover {{
                background: #e3f2fd;
                transition: background 0.3s;
            }}
            .vehicle-info {{
                font-size: 0.85em;
                line-height: 1.4;
            }}
            a {{
                color: #667eea;
                text-decoration: none;
                font-weight: bold;
            }}
            a:hover {{
                text-decoration: underline;
            }}
            .no-data {{
                text-align: center;
                color: #999;
                font-style: italic;
                padding: 40px;
            }}
            .stats {{
                display: flex;
                justify-content: center;
                gap: 30px;
                margin: 20px 0;
            }}
            .stat {{
                text-align: center;
                background: #667eea;
                color: white;
                padding: 15px 25px;
                border-radius: 8px;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            }}
            .stat-number {{
                font-size: 2em;
                font-weight: bold;
                display: block;
            }}
            .stat-label {{
                font-size: 0.9em;
                opacity: 0.9;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="section">
                <h1>📋 {partner.get('nom', 'Partner')}</h1>
                
                <div class="stats">
                    <div class="stat">
                        <span class="stat-number">{len(partner.get('drivers', []))}</span>
                        <span class="stat-label">Conducteurs</span>
                    </div>
                </div>
                
                <div class="partner-info">
                    <div class="info-card">
                        <div class="info-label">Email</div>
                        <div class="info-value">{partner.get('email', 'N/A')}</div>
                    </div>
                    <div class="info-card">
                        <div class="info-label">Téléphone</div>
                        <div class="info-value">{partner.get('telephone', 'N/A')}</div>
                    </div>
                    <div class="info-card">
                        <div class="info-label">Owner ID</div>
                        <div class="info-value">{partner.get('owner_id', 'N/A')}</div>
                    </div>
                    <div class="info-card">
                        <div class="info-label">Document</div>
                        <div class="info-value"><a href="{partner.get('document_url', '#')}" target="_blank">📄 Voir Document</a></div>
                    </div>
                </div>
            </div>
            
            <div class="section">
                <h2>🚗 Conducteurs Associés</h2>
                {f'''
                <table>
                    <thead>
                        <tr>
                            <th>Nom</th>
                            <th>Emplacement</th>
                            <th>Téléphone</th>
                            <th>Type Transport</th>
                            <th>Statut Véhicule</th>
                            <th>Informations Véhicule</th>
                            <th>Document</th>
                        </tr>
                    </thead>
                    <tbody>
                        {drivers_html}
                    </tbody>
                </table>
                ''' if drivers_html else '<div class="no-data">Aucun conducteur associé à ce partenaire</div>'}
            </div>
        </div>
    </body>
    </html>
    """
    
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN PROCESS
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    log("=" * 60)
    log("📊 ORGANISATION DES DONNÉES (VPS Mode)")
    log("=" * 60)
    
    try:
        # 1. Vérifier les fichiers requis
        if not PARTNERS_FILE.exists():
            msg = f"❌ Fichier manquant: {PARTNERS_FILE}"
            log(msg)
            send_slack(msg, "#ff0000")
            return
        
        if not DRIVERS_FILE.exists():
            msg = f"❌ Fichier manquant: {DRIVERS_FILE}"
            log(msg)
            send_slack(msg, "#ff0000")
            return
        
        # 2. Load data
        log("Chargement des données...")
        partners = load_json(PARTNERS_FILE)
        drivers = load_json(DRIVERS_FILE)
        log(f"  {len(partners)} partenaires chargés")
        log(f"  {len(drivers)} chauffeurs chargés")
        
        # 3. Filter partners by pattern
        log("Filtrage des partenaires (Partenaire-N)...")
        filtered_partners = []
        for partner in partners:
            partner_name = partner.get('nom', '')
            match = PARTNER_NAME_RE.match(partner_name)
            if match:
                partner_num = int(match.group(1))
                if partner_num >= PARTNER_MIN:
                    filtered_partners.append(partner)
                else:
                    log(f"  Ignoré: {partner_name} (numéro < {PARTNER_MIN})")
            else:
                log(f"  Ignoré: {partner_name} (pattern non reconnu)")
        
        log(f"  {len(filtered_partners)} partenaires conservés après filtrage")
        partners = filtered_partners
        
        # Create vehicle mapping from drivers data
        vehicle_map = {}
        for driver in drivers:
            phone = driver.get('telephone', '').strip()
            if phone:
                vehicle_map[phone] = driver.get('vehicle', {})
        log(f"  ✅ Mapping véhicule créé pour {len(vehicle_map)} chauffeurs (par téléphone)")
        
        # 3. Match data
        enriched_partners = match_drivers_to_partners(partners, drivers)
        
        # 4. Update vehicle data for all drivers
        log("🔄 Mise à jour des données véhicule...")
        updated_count = 0
        for partner in enriched_partners:
            for driver in partner.get('drivers', []):
                driver_phone = driver.get('telephone', '').strip()
                if driver_phone in vehicle_map:
                    driver['vehicle'] = vehicle_map[driver_phone]
                    updated_count += 1
        log(f"  ✅ {updated_count} véhicules mis à jour")
        
        # 5. Create organized directory structure
        log("📁 Création de la structure de dossiers...")
        ORGANIZED_DIR.mkdir(parents=True, exist_ok=True)
        
        # 6. Export for each partner
        log("📤 Exportation des données par partenaire...")
        for idx, partner in enumerate(enriched_partners, 1):
            partner_name = partner.get('nom', f'Partner_{idx}')
            folder_name = sanitize_folder_name(partner_name)
            partner_dir = ORGANIZED_DIR / folder_name
            partner_dir.mkdir(parents=True, exist_ok=True)
            
            num_drivers = len(partner.get('drivers', []))
            
            if idx % 10 == 0 or idx == len(enriched_partners):
                log(f"  [{idx}/{len(enriched_partners)}] {partner_name} ({num_drivers} conducteurs)")
            
            export_partner_data(partner, partner_dir)
        
        # 7. Create summary report
        log("📊 Création du rapport de synthèse...")
        summary_path = ORGANIZED_DIR / "SUMMARY.txt"
        with open(summary_path, "w", encoding="utf-8") as f:
            f.write("RÉSUMÉ DE L'ORGANISATION DES DONNÉES\n")
            f.write("=" * 60 + "\n")
            f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(f"Total Partenaires: {len(enriched_partners)}\n")
            f.write(f"Total Chauffeurs: {len(drivers)}\n")
            f.write(f"Chauffeurs Assignés: {sum(len(p.get('drivers', [])) for p in enriched_partners if p.get('nom') != 'UNASSIGNED_DRIVERS')}\n")
            f.write(f"Chauffeurs Non-Assignés: {len([p for p in enriched_partners if p.get('nom') == 'UNASSIGNED_DRIVERS'])}\n\n")
            
            f.write("DÉTAILS PAR PARTENAIRE:\n")
            f.write("-" * 60 + "\n")
            for partner in enriched_partners:
                num_drivers = len(partner.get('drivers', []))
                f.write(f"\n• {partner.get('nom', 'N/A')}\n")
                f.write(f"  Email: {partner.get('email', 'N/A')}\n")
                f.write(f"  Téléphone: {partner.get('telephone', 'N/A')}\n")
                f.write(f"  Conducteurs: {num_drivers}\n")
        
        # 8. Create global JSON with enriched data
        enriched_json = ORGANIZED_DIR / "all_partners_enriched.json"
        with open(enriched_json, "w", encoding="utf-8") as f:
            json.dump(enriched_partners, f, ensure_ascii=False, indent=2)
        
        # 9. Excel global (tout le monde)
        log("📊 Création du fichier Excel global...")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb_global = Workbook()
        
        # Feuille Partenaires
        ws_p = wb_global.active
        ws_p.title = "Partenaires"
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        p_headers = ["Nom", "Email", "Téléphone", "URL Document", "Owner ID", "Nb Conducteurs"]
        ws_p.append(p_headers)
        for cell in ws_p[1]:
            cell.fill = header_fill
            cell.font = header_font
        for p in enriched_partners:
            ws_p.append([
                p.get('nom', 'N/A'),
                p.get('email', 'N/A'),
                p.get('telephone', 'N/A'),
                p.get('document_url', 'N/A'),
                p.get('owner_id', 'N/A'),
                len(p.get('drivers', []))
            ])
        
        # Feuille Conducteurs
        ws_d = wb_global.create_sheet("Conducteurs")
        d_headers = ["Partenaire", "Nom", "Emplacement", "Téléphone", "Type Transport", "Type Véhicule", "Marque", "Modèle", "Matricule", "URL Document"]
        ws_d.append(d_headers)
        for cell in ws_d[1]:
            cell.fill = header_fill
            cell.font = header_font
        for p in enriched_partners:
            for d in p.get('drivers', []):
                v = d.get('vehicle', {})
                ws_d.append([
                    p.get('nom', 'N/A'),
                    d.get('nom', 'N/A'),
                    d.get('emplacement', 'N/A'),
                    d.get('telephone', 'N/A'),
                    d.get('type_transport', 'N/A'),
                    v.get('type', 'N/A'),
                    v.get('marque', 'N/A'),
                    v.get('modele', 'N/A'),
                    v.get('matricule', 'N/A'),
                    d.get('document_url', 'N/A')
                ])
        
        global_xlsx = ORGANIZED_DIR / "all_partners_enriched.xlsx"
        wb_global.save(global_xlsx)
        log(f"  ✅ Excel global: {global_xlsx}")
        
        # 10. HTML global (tout le monde)
        log("📊 Création du fichier HTML global...")
        all_rows_html = ""
        for p in enriched_partners:
            for d in p.get('drivers', []):
                v = d.get('vehicle', {})
                all_rows_html += f"""
                <tr>
                    <td>{p.get('nom', 'N/A')}</td>
                    <td>{d.get('nom', 'N/A')}</td>
                    <td>{d.get('emplacement', 'N/A')}</td>
                    <td>{d.get('telephone', 'N/A')}</td>
                    <td>{d.get('type_transport', 'N/A')}</td>
                    <td>{v.get('type', 'N/A')}</td>
                    <td>{v.get('marque', 'N/A')}</td>
                    <td>{v.get('modele', 'N/A')}</td>
                    <td>{v.get('matricule', 'N/A')}</td>
                    <td><a href="{d.get('document_url', '#')}" target="_blank">📄</a></td>
                </tr>
                """
        
        total_all_drivers = sum(len(p.get('drivers', [])) for p in enriched_partners)
        global_html_content = f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Tous les Partenaires & Conducteurs</title>
            <style>
                body {{ font-family: 'Segoe UI', sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; margin: 0; padding: 20px; }}
                .container {{ max-width: 1400px; margin: 0 auto; }}
                .section {{ background: white; padding: 30px; margin: 20px 0; border-radius: 10px; box-shadow: 0 10px 30px rgba(0,0,0,0.2); }}
                h1 {{ color: #667eea; text-align: center; }}
                .stats {{ display: flex; justify-content: center; gap: 30px; margin: 20px 0; }}
                .stat {{ text-align: center; background: #667eea; color: white; padding: 15px 25px; border-radius: 8px; }}
                .stat-number {{ font-size: 2em; font-weight: bold; display: block; }}
                .stat-label {{ font-size: 0.9em; opacity: 0.9; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #ddd; font-size: 0.85em; }}
                th {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; font-weight: bold; position: sticky; top: 0; }}
                tr:nth-child(even) {{ background: #f8f9fa; }}
                tr:hover {{ background: #e3f2fd; }}
                a {{ color: #667eea; text-decoration: none; font-weight: bold; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="section">
                    <h1>📊 Tous les Partenaires & Conducteurs</h1>
                    <div class="stats">
                        <div class="stat"><span class="stat-number">{len(enriched_partners)}</span><span class="stat-label">Partenaires</span></div>
                        <div class="stat"><span class="stat-number">{total_all_drivers}</span><span class="stat-label">Conducteurs</span></div>
                    </div>
                    <table>
                        <thead>
                            <tr>
                                <th>Partenaire</th>
                                <th>Nom</th>
                                <th>Emplacement</th>
                                <th>Téléphone</th>
                                <th>Type Transport</th>
                                <th>Type Véhicule</th>
                                <th>Marque</th>
                                <th>Modèle</th>
                                <th>Matricule</th>
                                <th>Doc</th>
                            </tr>
                        </thead>
                        <tbody>
                            {all_rows_html}
                        </tbody>
                    </table>
                </div>
            </div>
        </body>
        </html>
        """
        global_html_path = ORGANIZED_DIR / "all_partners_enriched.html"
        with open(global_html_path, "w", encoding="utf-8") as f:
            f.write(global_html_content)
        log(f"  ✅ HTML global: {global_html_path}")
        
        # 11. Stats
        total_assigned = sum(len(p.get('drivers', [])) for p in enriched_partners if p.get('nom') != 'UNASSIGNED_DRIVERS')
        unassigned_list = [p for p in enriched_partners if p.get('nom') == 'UNASSIGNED_DRIVERS']
        total_unassigned = sum(len(p.get('drivers', [])) for p in unassigned_list)
        
        log("=" * 60)
        log("✨ ORGANISATION TERMINÉE!")
        log(f"📊 {len(enriched_partners)} partenaires")
        log(f"✅ {total_assigned} chauffeurs assignés")
        log(f"❌ {total_unassigned} chauffeurs non-assignés")
        log(f"📁 {ORGANIZED_DIR}")
        log("=" * 60)
        
        msg = (
            f"✅ Organisation des données terminée!\n"
            f"📊 {len(enriched_partners)} partenaires\n"
            f"✅ {total_assigned} chauffeurs assignés\n"
            f"❌ {total_unassigned} chauffeurs non-assignés\n"
            f"📁 {ORGANIZED_DIR}"
        )
        send_slack(msg, "#36a64f")
    
    except Exception as e:
        err_msg = f"❌ Erreur organisation: {str(e)}"
        log(f"❌ ERREUR CRITIQUE: {e}")
        traceback.print_exc()
        send_slack(err_msg, "#ff0000")


if __name__ == "__main__":
    main()
