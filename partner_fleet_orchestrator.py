#!/usr/bin/env python3
"""
partner_fleet_orchestrator.py
=============================

Automatisation web multi-partenaires (1 navigateur, boucle temps réel) :

  Phase A — Compte partenaire : owner-login → fleet-drivers → 10 premiers chauffeurs
  Phase B — Admin : manage-owners → profil → Détails de la flotte → vérif assignation + pastilles

Registre : output/partner_automation/state.json (fusion à chaque tour — ne supprime pas transfer_2000_done)

Preuves : si tableau non vide (owner fleet-drivers ou admin flotte ≥1 ligne), capture PNG
pleine page dans output/partner_automation/proofs/ + chemins dans proof_screenshots.

Usage :
  python partner_fleet_orchestrator.py --start 1 --end 20
  python partner_fleet_orchestrator.py --only 3 --headed
  python partner_fleet_orchestrator.py --loop --interval 120
  python partner_fleet_orchestrator.py --excel vps_deploy/DOSSIER_PARTENAIRES.xlsx --limit 5
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import re
import shutil
import sys
import time
import traceback
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import (
    InvalidSessionIdException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait

import quick_approve_all_vehicle_vps as qa
from partner_state_mobile import merge_partner_record

# ─── Config ───────────────────────────────────────────────────────────────────

BASE_URL = os.getenv("UPJUNOO_BASE_URL", "https://upjunoo-server-new.junooapps.com")
OWNER_LOGIN_URL = f"{BASE_URL}/login/owner-login"
FLEET_DRIVERS_URL = f"{BASE_URL}/fleet-drivers"
ADMIN_LOGIN_URL = f"{BASE_URL}/login/admin"
MANAGE_OWNERS_URL = f"{BASE_URL}/manage-owners"
OWNER_DASHBOARD_URL = f"{BASE_URL}/owner-dashboard"
VIEW_PROFILE_URL = f"{BASE_URL}/manage-owners/view-profile"
LOGOUT_URL = f"{BASE_URL}/logout"

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_CONFIG_JSON = SCRIPT_DIR / "partner_automation_config.json"
OUTPUT_DIR = SCRIPT_DIR / "output" / "partner_automation"
STATE_FILE = OUTPUT_DIR / "state.json"
LOG_FILE = OUTPUT_DIR / "orchestrator.log"
DEBUG_DIR = OUTPUT_DIR / "debug"
PROOFS_DIR = OUTPUT_DIR / "proofs"

DEFAULT_EMAIL_TEMPLATE = os.getenv("PARTNER_EMAIL_TEMPLATE", "campagne{index}@upjunoo.com")
DEFAULT_PARTNER_PASSWORD = os.getenv("PARTNER_PASSWORD", "123456789@")
ADMIN_EMAIL = os.getenv("UPJUNOO_EMAIL", "admin@upjunoo.com")
ADMIN_PASSWORD = os.getenv("UPJUNOO_PASSWORD", "Upjunoo@Admin")

DRIVERS_TOP_N = 10

# Indicateurs visuels (bordures + pastilles)
VIS_GREEN = "#2e7d32"
VIS_ORANGE = "#ef6c00"
VIS_RED = "#c62828"
VIS_BLUE = "#1565c0"
VIS_GRAY = "#757575"
VIS_PURPLE = "#6a1b9a"
VIS_TEAL = "#00695c"
VIS_NAVY = "#0d47a1"


def log(msg: str, level: str = "INFO") -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}][{level}] {msg}"
    try:
        print(line, flush=True)
    except UnicodeEncodeError:
        enc = getattr(sys.stdout, "encoding", None) or "utf-8"
        print(line.encode(enc, errors="replace").decode(enc, errors="replace"), flush=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with LOG_FILE.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def normalize_text(value: str) -> str:
    value = (value or "").strip().lower()
    value = unicodedata.normalize("NFD", value)
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", value)


def normalize_phone(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    return digits[-10:] if len(digits) >= 10 else digits


def names_match(a: str, b: str) -> bool:
    na, nb = normalize_text(a), normalize_text(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    if na in nb or nb in na:
        return True
    # Prénom/nom inversés ou partiel
    pa = set(na.split())
    pb = set(nb.split())
    return len(pa & pb) >= min(2, min(len(pa), len(pb)))


def is_browser_dead(exc: BaseException) -> bool:
    if isinstance(exc, InvalidSessionIdException):
        return True
    if isinstance(exc, WebDriverException):
        m = str(exc).lower()
        return any(
            k in m
            for k in ("invalid session", "session deleted", "disconnected", "not reachable")
        )
    return False


def make_driver(*, headed: bool = False) -> webdriver.Chrome:
    """Chrome via ChromeDriverManager (évite SeleniumManager bloqué sous Windows)."""
    return qa.setup_driver(headed=headed)


def wait_for_table(driver: webdriver.Chrome, timeout: int = 30, min_rows: int = 1) -> bool:
    ignore = ("chargement", "loading", "aucune", "no data", "vide")
    end = time.time() + timeout
    while time.time() < end:
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            if len(rows) >= min_rows:
                txt = (rows[0].text or "").lower()
                if not any(w in txt for w in ignore):
                    return True
        except Exception:
            pass
        time.sleep(0.8)
    return False


def inject_banner(driver: webdriver.Chrome, text: str, color: str = VIS_BLUE) -> None:
    """Bannière fixe en haut de page (visible en mode --headed)."""
    try:
        driver.execute_script(
            """
            let el = document.getElementById('pa-orch-banner');
            if (!el) {
                el = document.createElement('div');
                el.id = 'pa-orch-banner';
                document.body.prepend(el);
            }
            el.style.cssText = [
                'position:fixed','top:0','left:0','right:0','z-index:99999',
                'padding:10px 14px','color:#fff','font:600 13px sans-serif',
                'background:' + arguments[1], 'box-shadow:0 2px 8px rgba(0,0,0,.25)'
            ].join(';');
            el.textContent = arguments[0];
            document.body.style.marginTop = '48px';
            """,
            text,
            color,
        )
    except Exception:
        pass


def inject_progress_panel(
    driver: webdriver.Chrome,
    *,
    slot: int,
    total: int,
    step: str,
    detail: str,
    color: str = VIS_BLUE,
) -> None:
    try:
        driver.execute_script(
            """
            let root = document.getElementById('pa-progress-panel');
            if (!root) {
                root = document.createElement('div');
                root.id = 'pa-progress-panel';
                root.style.cssText = [
                    'position:fixed','top:52px','right:12px','z-index:99998',
                    'min-width:240px','padding:12px 14px','border-radius:8px',
                    'color:#fff','font:600 12px/1.45 sans-serif',
                    'box-shadow:0 4px 14px rgba(0,0,0,.35)'
                ].join(';');
                document.body.appendChild(root);
            }
            root.style.background = arguments[4];
            const pct = arguments[1] > 0
                ? Math.round(100 * arguments[0] / arguments[1]) : 0;
            root.innerHTML =
                '<div style="font-size:11px;opacity:.92">AUTOMATISATION PARTENAIRES</div>'
                + '<div style="font-size:17px;margin:6px 0">Campagne '
                + arguments[0] + ' &rarr; ' + arguments[1] + '</div>'
                + '<div style="background:rgba(255,255,255,.28);height:7px;border-radius:4px;margin:8px 0">'
                + '<div style="background:#fff;height:100%;width:' + pct
                + '%;border-radius:4px"></div></div>'
                + '<div style="font-size:13px">' + arguments[2] + '</div>'
                + '<div style="font-size:11px;margin-top:5px;opacity:.88">'
                + arguments[3] + '</div>';
            """,
            slot,
            total,
            step,
            detail,
            color,
        )
    except Exception:
        pass


def flash_viewport_border(driver: webdriver.Chrome, color: str, *, active: bool = True) -> None:
    try:
        driver.execute_script(
            "document.body.style.outline = arguments[0] ? ('5px solid ' + arguments[1]) : '';",
            active,
            color,
        )
    except Exception:
        pass


def show_phase_ui(
    driver: webdriver.Chrome,
    *,
    partner: dict[str, Any],
    slot_pos: int,
    total_slots: int,
    step: str,
    detail: str,
    color: str,
    show_ui: bool,
) -> None:
    if not show_ui:
        return
    idx = partner.get("index", "?")
    name = partner.get("name") or f"Campagne UPJUNOO {idx}"
    inject_progress_panel(
        driver,
        slot=slot_pos,
        total=total_slots,
        step=step,
        detail=f"{name} | {detail}",
        color=color,
    )
    inject_banner(
        driver,
        f"▶ Campagne {slot_pos}/{total_slots} (n°{idx}) — {step}",
        color,
    )
    flash_viewport_border(driver, color, active=True)


def log_campaign_plan(partners: list[dict[str, Any]], *, start: int, end: int) -> None:
    log(f"   Ordre : Campagne {start} → {end} ({len(partners)} slot(s))")
    for p in partners:
        src = "Excel" if p.get("login_from_excel") else "campagne@upjunoo"
        em = p.get("email", "")
        mask = em[:4] + "…@" + em.split("@", 1)[1] if "@" in em else em
        log(f"      [{p['index']:>2}] {p.get('name', '')} | {mask} ({src})")


def inject_legend(driver: webdriver.Chrome, phase: str) -> None:
    if phase == "owner":
        txt = (
            "PARTENAIRE fleet-drivers | vert=APPROUVÉ | orange=EN ATTENTE | "
            "rouge=DÉSAPPROUVÉ | bleu=ligne lue | gris=?"
        )
    else:
        txt = (
            "ADMIN flotte | vert=chauffeur assigné+OK | orange=assigné EN ATTENTE | "
            "rouge=absent / pas sur flotte | bleu=autre"
        )
    inject_banner(driver, txt, VIS_PURPLE)


def highlight_row(driver: webdriver.Chrome, row: Any, color: str, label: str = "") -> None:
    try:
        driver.execute_script(
            """
            const row = arguments[0];
            const color = arguments[1];
            const label = arguments[2];
            row.style.outline = '3px solid ' + color;
            row.style.backgroundColor = color + '33';
            row.style.transition = 'background .2s';
            if (label) {
                const key = 'pa_' + label;
                if (!row.dataset.paKey || row.dataset.paKey !== key) {
                    row.dataset.paKey = key;
                    const old = row.querySelector('.pa-orch-badge');
                    if (old) old.remove();
                    const td = row.querySelector('td:last-child') || row.querySelector('td');
                    if (td) {
                        const b = document.createElement('span');
                        b.className = 'pa-orch-badge';
                        b.textContent = label;
                        b.style.cssText = 'margin-left:6px;padding:3px 8px;border-radius:4px;'
                            + 'font-size:11px;font-weight:bold;color:#fff;background:' + color;
                        td.appendChild(b);
                    }
                }
            }
            """,
            row,
            color,
            label,
        )
    except Exception:
        pass


def _is_placeholder_row(row: Any) -> bool:
    try:
        txt = (row.text or "").strip().lower()
        if not txt:
            return True
        if any(w in txt for w in ("chargement", "loading", "aucune", "no data", "no records")):
            return True
        cells = row.find_elements(By.TAG_NAME, "td")
        if len(cells) < 3:
            return True
        first = (cells[0].text or "").strip().lower()
        if first in ("nom", "name", "#"):
            return True
    except Exception:
        return True
    return False


def table_column_map(driver: webdriver.Chrome) -> dict[str, int]:
    """Index des colonnes depuis thead (fleet-drivers / pilotes)."""
    out: dict[str, int] = {}
    for i, th in enumerate(driver.find_elements(By.CSS_SELECTOR, "table thead th")):
        key = normalize_text(th.text or "")
        if not key:
            continue
        if "nom" in key and "vehicule" not in key and "conducteur" not in key:
            out.setdefault("name", i)
        if "portable" in key or "telephone" in key:
            out.setdefault("phone", i)
        if "statut" in key:
            out.setdefault("status", i)
        if "vehicule" in key and "type" in key:
            out.setdefault("vehicle_type", i)
        if "conducteur" in key:
            out.setdefault("driver_on_fleet", i)
    return out


def _cell_text(cells: list, idx: int | None) -> str:
    if idx is None or idx >= len(cells):
        return ""
    return (cells[idx].text or "").strip()


def _owner_status_style(row_text: str) -> tuple[str, str]:
    u = (row_text or "").upper()
    if "DESAPPROUV" in u or "DÉSAPPROUV" in row_text.upper():
        return VIS_RED, "DESAPPROUVE"
    if "APPROUV" in u:
        return VIS_GREEN, "APPROUVE"
    if "ATTENTE" in u:
        return VIS_ORANGE, "EN_ATTENTE"
    return VIS_GRAY, "?"


def set_page_size(driver: webdriver.Chrome, size: int = 500) -> bool:
    """Pagination DataTables — sans planter si le select est vide / caché."""
    try:
        time.sleep(1.0)
        selects = driver.find_elements(
            By.CSS_SELECTOR,
            "select.form-select-sm, select[name*='length'], .dataTables_length select, select.form-select",
        )
        for el in selects:
            try:
                if not el.is_displayed():
                    continue
                s = Select(el)
                opts = [o for o in s.options if (o.get_attribute("value") or "").strip() != ""]
                if not opts:
                    continue
                values = [o.get_attribute("value") for o in opts]
                texts = [(o.text or "").strip() for o in opts]
                target = str(size)
                if target in values:
                    s.select_by_value(target)
                else:
                    for pref in (str(size), "100", "50", "25", "10", "All", "Tout"):
                        if pref in texts:
                            s.select_by_visible_text(pref)
                            break
                    else:
                        s.select_by_index(len(opts) - 1)
                log(f"   [PAGE] Pagination → {s.first_selected_option.text.strip()}", "OK")
                time.sleep(2.5)
                wait_for_table(driver, timeout=30, min_rows=1)
                return True
            except Exception:
                continue
        return False
    except Exception as e:
        log(f"   Pagination: {e}", "WARNING")
        return False


def try_search(driver: webdriver.Chrome, query: str) -> bool:
    if not (query or "").strip():
        return False
    try:
        box = WebDriverWait(driver, 8).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "input[type='search'], .dataTables_filter input"),
            ),
        )
        box.clear()
        time.sleep(0.2)
        box.send_keys(query)
        time.sleep(2.5)
        return True
    except Exception:
        return False


def wait_table_row_count_stable(driver: webdriver.Chrome, checks: int = 3, pause_s: float = 2) -> int:
    """Comme sync_admin_fleet_status — après pagination 500."""
    last_count = -1
    stable_hits = 0
    for _ in range(checks * 3):
        try:
            count = len(driver.find_elements(By.CSS_SELECTOR, "table tbody tr"))
            if count == last_count and count > 0:
                stable_hits += 1
                if stable_hits >= checks:
                    return count
            else:
                stable_hits = 0
            last_count = count
        except Exception:
            stable_hits = 0
        time.sleep(pause_s)
    return last_count


def find_row_by_cell_text_with_pagination(
    driver: webdriver.Chrome,
    target_text: str,
    max_pages: int = 25,
) -> Any | None:
    """Scan cellule par cellule + pagination DataTables (sync_admin_fleet_status)."""
    q = (target_text or "").strip().lower()
    if not q:
        return None
    next_btn_xpath = "//li[contains(@class, 'next') and not(contains(@class, 'disabled'))]/a"
    for _ in range(max_pages):
        if not wait_for_table(driver, timeout=15, min_rows=1):
            break
        for row in driver.find_elements(By.CSS_SELECTOR, "table tbody tr"):
            if not row.is_displayed():
                continue
            try:
                for cell in row.find_elements(By.TAG_NAME, "td"):
                    if q in (cell.text or "").strip().lower():
                        return row
            except Exception:
                continue
        try:
            nxt = driver.find_element(By.XPATH, next_btn_xpath)
            if nxt.is_displayed():
                driver.execute_script("arguments[0].click();", nxt)
                time.sleep(2.5)
                continue
        except Exception:
            pass
        break
    return None


def partner_search_queries(partner: dict[str, Any]) -> list[str]:
    """Email + nom dashboard avant libellés Campagne UPJUNOO."""
    idx = partner.get("index")
    email = (partner.get("email") or "").strip()
    name = (partner.get("name") or "").strip()
    display = (partner.get("display_name") or "").strip()
    queries: list[str] = []
    for q in (email, display, name):
        if q and q not in queries:
            queries.append(q)
    is_upjunoo_campagne = (
        email.lower().endswith("@upjunoo.com") and "campagne" in email.lower()
    )
    if is_upjunoo_campagne and idx is not None:
        for q in (
            f"Campagne UPJUNOO {idx}",
            f"Campagne UPJUNOO {idx}".upper(),
            f"campagne{idx}",
            f"partenaire{idx}",
            f"Partenaire {idx}",
        ):
            if q not in queries:
                queries.append(q)
    if email and "@" in email:
        local = email.split("@", 1)[0]
        if len(local) >= 4 and local not in queries:
            queries.append(local)
    return queries


def is_on_partner_profile(driver: webdriver.Chrome) -> bool:
    return "/manage-owners/view-profile/" in (driver.current_url or "").lower()


def extract_profile_uuid_from_row(row: Any) -> str:
    try:
        for link in row.find_elements(By.CSS_SELECTOR, "a[href*='view-profile']"):
            href = link.get_attribute("href") or ""
            m = re.search(r"/view-profile/([a-f0-9-]{36})", href, re.I)
            if m:
                return m.group(1).lower()
        for link in row.find_elements(By.CSS_SELECTOR, "a[href*='/document/']"):
            href = link.get_attribute("href") or ""
            m = re.search(r"/document/([a-f0-9-]{36})", href, re.I)
            if m:
                return m.group(1).lower()
    except Exception:
        pass
    return ""


def enrich_partners_from_state(
    partners: list[dict[str, Any]],
    state: dict[str, Any],
) -> None:
    pmap = state.get("partners") or {}
    for p in partners:
        prev = pmap.get(str(p.get("index")))
        if not prev:
            continue
        if prev.get("profile_uuid") and not p.get("profile_uuid"):
            p["profile_uuid"] = prev["profile_uuid"]
        if prev.get("display_name") and not p.get("display_name"):
            p["display_name"] = prev["display_name"]
        prev_name = (prev.get("name") or "").strip()
        cur_name = (p.get("name") or "").strip()
        if prev_name and (
            not cur_name
            or cur_name.startswith("Partenaire ")
            or cur_name.startswith("Campagne UPJUNOO")
        ):
            p["name"] = prev_name


def scrape_owner_display_name(driver: webdriver.Chrome) -> str:
    """Lit le nom affiché sur owner-dashboard (souvent = libellé dans manage-owners admin)."""
    try:
        url = driver.current_url.lower()
        if "owner-dashboard" not in url and "dashboard" not in url:
            driver.get(OWNER_DASHBOARD_URL)
            time.sleep(2)
        for sel in (
            "h1",
            "h2",
            ".page-title",
            ".profile-name",
            "[class*='partner']",
        ):
            for el in driver.find_elements(By.CSS_SELECTOR, sel):
                txt = (el.text or "").strip()
                if txt and len(txt) > 3 and "dashboard" not in txt.lower():
                    return txt
    except Exception:
        pass
    return ""


def navigate_manage_owners_table(driver: webdriver.Chrome) -> None:
    """Ouvre manage-owners avec pagination 500 et tableau stabilisé."""
    log("   [ADMIN] Ouverture manage-owners…")
    driver.get(MANAGE_OWNERS_URL)
    time.sleep(4)
    wait_for_table(driver, timeout=40)
    set_page_size(driver, 500)
    wait_for_table(driver, timeout=45, min_rows=1)
    time.sleep(4)
    n = wait_table_row_count_stable(driver, checks=3, pause_s=2)
    log(f"   [ADMIN] Tableau partenaires stabilisé (~{n} lignes visibles)")


def find_partner_row_admin(
    driver: webdriver.Chrome,
    partner: dict[str, Any],
) -> Any | None:
    """Recherche partenaire comme sync_admin_fleet_status / check_partner_drivers_admin."""
    navigate_manage_owners_table(driver)
    for query in partner_search_queries(partner):
        log(f"   [ADMIN] Recherche partenaire : {query!r}")
        try_search(driver, "")
        time.sleep(0.5)
        try_search(driver, query)
        row = find_row_by_cell_text_with_pagination(driver, query, max_pages=25)
        if row:
            log(f"   [ADMIN] Partenaire trouvé via {query!r}", "OK")
            return row
    return None


def open_partner_profile_by_uuid(driver: webdriver.Chrome, profile_uuid: str) -> bool:
    uuid = (profile_uuid or "").strip()
    if not uuid:
        return False
    url = f"{VIEW_PROFILE_URL}/{uuid}"
    log(f"   [ADMIN] Profil direct → {url}")
    try:
        driver.get(url)
        time.sleep(4)
        WebDriverWait(driver, 30).until(lambda d: is_on_partner_profile(d))
        return is_on_partner_profile(driver)
    except Exception as e:
        log(f"   [ADMIN] Profil UUID échoué: {e}", "WARNING")
        return False


def clear_session(driver: webdriver.Chrome) -> None:
    try:
        driver.get(LOGOUT_URL)
        time.sleep(0.8)
    except Exception:
        pass
    try:
        driver.delete_all_cookies()
    except Exception:
        pass


def _dismiss_alert(driver: webdriver.Chrome) -> str | None:
    try:
        from selenium.webdriver.common.alert import Alert

        alert = Alert(driver)
        txt = alert.text
        alert.accept()
        return txt
    except Exception:
        return None


def _save_login_debug(driver: webdriver.Chrome, tag: str) -> None:
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%H%M%S")
    path = DEBUG_DIR / f"{tag}_{ts}.png"
    try:
        driver.save_screenshot(str(path))
        log(f"   Capture → {path}", "WARNING")
    except Exception:
        pass


def _sanitize_proof_token(value: str, max_len: int = 40) -> str:
    s = re.sub(r"[^\w.\-]+", "_", (value or "").strip())
    return s[:max_len] or "unknown"


def save_full_page_screenshot(driver: webdriver.Chrome, path: Path) -> bool:
    """Capture pleine page (CDP Chrome) — headed ou headless."""
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        driver.execute_cdp_cmd("Page.enable", {})
        metrics = driver.execute_cdp_cmd("Page.getLayoutMetrics", {})
        content = metrics.get("contentSize") or metrics.get("cssContentSize") or {}
        width = int(content.get("width") or 1920)
        height = int(content.get("height") or 1080)
        width = max(width, 800)
        height = max(height, 600)
        result = driver.execute_cdp_cmd(
            "Page.captureScreenshot",
            {
                "captureBeyondViewport": True,
                "fromSurface": True,
                "clip": {
                    "x": 0,
                    "y": 0,
                    "width": width,
                    "height": height,
                    "scale": 1,
                },
            },
        )
        path.write_bytes(base64.b64decode(result["data"]))
        return True
    except Exception as e:
        log(f"   [PREUVE] CDP pleine page échouée ({e}) — repli viewport", "WARNING")
        try:
            driver.save_screenshot(str(path))
            return path.is_file()
        except Exception as e2:
            log(f"   [PREUVE] Capture impossible: {e2}", "WARNING")
            return False


def capture_list_proof_screenshot(
    driver: webdriver.Chrome,
    partner: dict[str, Any],
    *,
    phase: str,
    row_count: int,
) -> str | None:
    """
    Preuve PNG si le tableau n'est pas vide.
    phase: owner | admin
    Retourne chemin relatif à OUTPUT_DIR (ex. proofs/P01_...png).
    """
    if row_count < 1:
        return None

    idx = int(partner.get("index", 0))
    email = _sanitize_proof_token(str(partner.get("email", "")), 36)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    label = "owner_fleet-drivers" if phase == "owner" else "admin_flotte"
    filename = f"P{idx:02d}_{email}_{label}_{row_count}lignes_{ts}.png"
    full_path = PROOFS_DIR / filename

    try:
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.35)
    except Exception:
        pass

    if not save_full_page_screenshot(driver, full_path):
        return None

    rel = str(full_path.relative_to(OUTPUT_DIR)).replace("\\", "/")
    log(f"   [PREUVE] {phase.upper()} — {row_count} ligne(s) → {rel}", "OK")
    return rel


def _is_owner_logged_in(driver: webdriver.Chrome) -> bool:
    """Hors écran owner-login (URL ou formulaire)."""
    url = driver.current_url.lower()
    if "owner-login" in url:
        return False
    if any(p in url for p in ("/fleet-drivers", "/manage-fleet", "/dashboard", "/home")):
        return True
    try:
        email_inputs = driver.find_elements(By.ID, "email-input")
        if email_inputs and email_inputs[0].is_displayed():
            return False
        return "owner-login" not in url
    except Exception:
        return "owner-login" not in url


def _wait_owner_login_success(driver: webdriver.Chrome, timeout: float = 45) -> bool:
    end = time.time() + timeout
    last = ""
    while time.time() < end:
        _dismiss_alert(driver)
        if _is_owner_logged_in(driver):
            return True
        cur = driver.current_url
        if cur != last:
            last = cur
            log(f"   [OWNER] URL → {cur}")
        time.sleep(0.5)
    return _is_owner_logged_in(driver)


def owner_login(driver: webdriver.Chrome, email: str, password: str) -> bool:
    """
    Login partenaire (owner-login).
    Champs : #email-input / #password-input (comme les autres scripts vps_deploy).
    """
    log(f"   [OWNER] Connexion {email}")
    clear_session(driver)

    for attempt in range(1, 4):
        try:
            _dismiss_alert(driver)
            driver.get(OWNER_LOGIN_URL)
            wait = WebDriverWait(driver, 25)

            if _is_owner_logged_in(driver):
                log(f"   [OWNER] Déjà connecté → {driver.current_url}", "OK")
                return True

            email_el = wait.until(EC.element_to_be_clickable((By.ID, "email-input")))
            pwd_el = driver.find_element(By.ID, "password-input")
            email_el.clear()
            email_el.send_keys(email)
            pwd_el.clear()
            pwd_el.send_keys(password)

            for sel in (
                "//button[@type='submit']",
                "//button[contains(@class,'btn-success')]",
                "//button[contains(., 'Connexion') or contains(., 'connexion')]",
            ):
                try:
                    btn = driver.find_element(By.XPATH, sel)
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].click();", btn)
                        break
                except Exception:
                    continue
            else:
                pwd_el.submit()

            alert = _dismiss_alert(driver)
            if alert:
                log(f"   [OWNER] Alerte login: {alert}", "WARNING")

            if _wait_owner_login_success(driver, timeout=40):
                log(f"   [OWNER] OK → {driver.current_url}", "OK")
                return True

            log(
                f"   [OWNER] Tentative {attempt}/3 — toujours sur login "
                f"(URL: {driver.current_url})",
                "WARNING",
            )
            _save_login_debug(driver, f"owner_login_fail_{attempt}")

        except TimeoutException as e:
            if _is_owner_logged_in(driver):
                log(f"   [OWNER] OK (après timeout) → {driver.current_url}", "OK")
                return True
            log(f"   [OWNER] Timeout tentative {attempt}/3: {e}", "WARNING")
            _save_login_debug(driver, f"owner_login_timeout_{attempt}")
        except Exception as e:
            log(f"   [OWNER] Erreur tentative {attempt}/3: {e}", "WARNING")
            _save_login_debug(driver, f"owner_login_err_{attempt}")

        time.sleep(2 * attempt)

    log(
        "   [OWNER] Échec — vérifie email/mot de passe "
        f"({email} / mot de passe masqué). Test manuel sur owner-login.",
        "ERROR",
    )
    return False


def admin_login(driver: webdriver.Chrome) -> bool:
    """Réutilise la connexion admin éprouvée de quick_approve_all_vehicle_vps."""
    log(f"   [ADMIN] Connexion {ADMIN_EMAIL}")
    clear_session(driver)
    ok = qa.admin_login(driver, 0)
    if ok:
        log(f"   [ADMIN] OK → {driver.current_url}", "OK")
    else:
        log("   [ADMIN] Échec — vérifie UPJUNOO_EMAIL / UPJUNOO_PASSWORD", "ERROR")
        _save_login_debug(driver, "admin_login_fail")
    return ok


def scrape_fleet_drivers_top(
    driver: webdriver.Chrome,
    *,
    top_n: int = DRIVERS_TOP_N,
    visual: bool = True,
) -> list[dict[str, str]]:
    """Phase A : chauffeurs sur /fleet-drivers (table APPROVED DRIVERS)."""
    log(f"   [OWNER] Ouverture {FLEET_DRIVERS_URL}")
    driver.get(FLEET_DRIVERS_URL)
    time.sleep(3)
    inject_legend(driver, "owner")

    if not wait_for_table(driver, timeout=45, min_rows=1):
        log("   [OWNER] Tableau fleet-drivers vide / lent — retry sans pagination", "WARNING")
        time.sleep(3)

    # Pagination optionnelle (ne doit pas bloquer le scrape)
    if not set_page_size(driver, 10):
        log("   [OWNER] Pagination ignorée (on garde les lignes visibles)", "WARNING")

    wait_for_table(driver, timeout=20, min_rows=1)
    col = table_column_map(driver)
    if col:
        log(f"   [OWNER] Colonnes détectées : {col}")

    name_i = col.get("name", 0)
    phone_i = col.get("phone", 3)
    veh_i = col.get("vehicle_type")

    drivers: list[dict[str, str]] = []
    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
    log(f"   [OWNER] {len(rows)} ligne(s) brute(s) dans le tableau")

    for row in rows:
        if len(drivers) >= top_n:
            break
        if _is_placeholder_row(row):
            continue
        try:
            cells = row.find_elements(By.TAG_NAME, "td")
            name = _cell_text(cells, name_i)
            if not name or len(name) < 2:
                continue
            phone = _cell_text(cells, phone_i)
            if not normalize_phone(phone):
                for idx in range(len(cells)):
                    t = _cell_text(cells, idx)
                    if normalize_phone(t):
                        phone = t
                        break
            vehicle_type = _cell_text(cells, veh_i) if veh_i is not None else ""
            row_txt = row.text or ""
            status_i = col.get("status")
            status_cell = _cell_text(cells, status_i) if status_i is not None else ""
            if status_cell:
                owner_status = status_cell
            else:
                owner_status = (
                    "APPROUVÉ" if "APPROUV" in row_txt.upper()
                    else "EN ATTENTE" if "ATTENTE" in row_txt.upper()
                    else "DÉSAPPROUVÉ" if "DESAPPROUV" in row_txt.upper() or "DÉSAPPROUV" in row_txt
                    else "?"
                )
            color, label = _owner_status_style(row_txt)
            if visual:
                highlight_row(driver, row, color, label[:12])
            drivers.append(
                {
                    "name": name,
                    "phone": phone,
                    "vehicle_type": vehicle_type,
                    "owner_status": owner_status,
                    "scraped_at": datetime.now().isoformat(timespec="seconds"),
                },
            )
            log(f"      → {name} | {phone or '-'} | {owner_status} | véh.={vehicle_type or '-'}")
        except Exception as e:
            log(f"   [OWNER] Ligne ignorée: {e}", "WARNING")
            continue

    inject_banner(
        driver,
        f"PARTENAIRE — {len(drivers)} chauffeur(s) lu(s) sur fleet-drivers (max {top_n})",
        VIS_GREEN if drivers else VIS_RED,
    )
    log(f"   [OWNER] {len(drivers)} chauffeur(s) récupéré(s) (top {top_n})")
    return drivers


def open_partner_profile(driver: webdriver.Chrome, row: Any) -> bool:
    """Ouverture profil — navigation directe par UUID si possible."""
    uuid = extract_profile_uuid_from_row(row)
    if uuid and open_partner_profile_by_uuid(driver, uuid):
        log(f"   [ADMIN] Profil ouvert (UUID) → {driver.current_url}", "OK")
        return True
    try:
        highlight_row(driver, row, "#1565c0", "partenaire")
        profile_btn = None
        try:
            profile_btn = row.find_element(
                By.XPATH, ".//a[contains(@href, 'view-profile')]",
            )
        except Exception:
            pass
        if not profile_btn:
            try:
                profile_btn = row.find_element(
                    By.XPATH,
                    ".//a[contains(@href, 'profile') or contains(@href, 'view')]",
                )
            except Exception:
                pass
        if not profile_btn:
            return False
        href = (profile_btn.get_attribute("href") or "").strip()
        m = re.search(r"/view-profile/([a-f0-9-]{36})", href, re.I)
        if m:
            return open_partner_profile_by_uuid(driver, m.group(1).lower())
        log("   [ADMIN] Clic « voir profil »…")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", profile_btn)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", profile_btn)
        WebDriverWait(driver, 35).until(lambda d: is_on_partner_profile(d))
        time.sleep(2)
        ok = is_on_partner_profile(driver)
        if ok:
            log(f"   [ADMIN] Profil ouvert → {driver.current_url}", "OK")
        return ok
    except Exception as e:
        log(f"   [ADMIN] Profil: {e}", "WARNING")
        return False


def open_partner_profile_admin(driver: webdriver.Chrome, partner: dict[str, Any]) -> bool:
    """UUID direct si connu, sinon recherche + navigation profil."""
    uuid = (partner.get("profile_uuid") or "").strip()
    if uuid and open_partner_profile_by_uuid(driver, uuid):
        return True
    row = find_partner_row_admin(driver, partner)
    if not row:
        return False
    found_uuid = extract_profile_uuid_from_row(row)
    if found_uuid:
        partner["profile_uuid"] = found_uuid
        log(f"   [ADMIN] UUID mémorisé …{found_uuid[-8:]}", "OK")
        return open_partner_profile_by_uuid(driver, found_uuid)
    return open_partner_profile(driver, row)


def open_fleet_tab(driver: webdriver.Chrome) -> bool:
    if not is_on_partner_profile(driver):
        log("   [ADMIN] Onglet flotte refusé — pas sur view-profile", "WARNING")
        return False
    inject_legend(driver, "admin")
    for sel in (
        "//a[contains(., 'Détails de la flotte')]",
        "//a[contains(., 'flotte') or contains(., 'Flotte')]",
        "//span[contains(., 'flotte') or contains(., 'Flotte')]",
    ):
        try:
            for el in driver.find_elements(By.XPATH, sel):
                if el.is_displayed():
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    driver.execute_script("arguments[0].click();", el)
                    time.sleep(4)
                    wait_for_table(driver, timeout=30)
                    # Pagination seulement si le select est visible (évite not interactable)
                    set_page_size(driver, 500)
                    wait_for_table(driver, timeout=25)
                    return True
        except Exception:
            continue
    return False


def _looks_like_plate(text: str) -> bool:
    """Évite de confondre modèle (T55, Dzire) avec la colonne plaque."""
    t = (text or "").strip()
    if not t or t in ("-", "—", "N/A"):
        return False
    tu = t.upper()
    if "APPROUV" in tu or "ATTENTE" in tu:
        return False
    # Modèle court sans tiret (ex. T55) — pas une plaque
    if len(t) <= 5 and "-" not in t:
        return False
    if "-" in t and any(ch.isdigit() for ch in t):
        return True
    if len(t) >= 6 and any(ch.isdigit() for ch in t) and any(ch.isalpha() for ch in t):
        return True
    return bool(re.search(r"[A-Z]{2,}[-\s]?\d", t, re.I))


def parse_fleet_detail_rows(driver: webdriver.Chrome) -> list[dict[str, str]]:
    """Parse l'onglet Détails de la flotte (col. 4 = plaque, col. 8 = conducteur)."""
    rows_data: list[dict[str, str]] = []
    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
    for row in rows:
        try:
            if _is_placeholder_row(row):
                continue
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) < 6:
                continue
            texts = [(c.text or "").strip() for c in cells]
            if not any(t for t in texts if t and t not in ("-", "—", "N/A")):
                continue

            plate = ""
            driver_name = ""
            status = ""

            # Colonnes fixes tableau admin : type | marque | modèle | plaque | … | statut | … | conducteur
            if len(texts) >= 4 and _looks_like_plate(texts[3]):
                plate = texts[3]
            if len(texts) >= 8 and texts[7] not in ("-", "—"):
                driver_name = texts[7]
            if len(texts) >= 6 and not status:
                st = texts[5]
                if "APPROUV" in st.upper() or "ATTENTE" in st.upper():
                    status = st

            for t in texts:
                tu = t.upper()
                if not status and ("APPROUV" in tu or "ATTENTE" in tu):
                    status = t
                elif not plate and _looks_like_plate(t):
                    plate = t

            if not driver_name:
                for t in reversed(texts):
                    if t in ("-", "—", ""):
                        continue
                    if "APPROUV" in t.upper() or "ATTENTE" in t.upper():
                        continue
                    if len(t) > 3 and (not plate or t != plate):
                        driver_name = t
                        break

            if not (plate or driver_name or status):
                continue

            rows_data.append(
                {
                    "plate": plate,
                    "driver_name": driver_name if driver_name not in ("-", "—") else "",
                    "fleet_status": status,
                    "row_text": " | ".join(texts[:9]),
                },
            )
        except Exception:
            continue
    return rows_data


def admin_check_drivers(
    driver: webdriver.Chrome,
    partner: dict[str, Any],
    drivers: list[dict[str, str]],
    *,
    proof_sink: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Phase B : vérifie chaque chauffeur sur la flotte admin + indicateurs."""
    results: list[dict[str, Any]] = []
    email = partner.get("email", "")
    name = partner.get("name", "") or f"Partenaire {partner.get('index', '?')}"
    log(f"   [ADMIN] Cible : {name} | {email}")

    if not open_partner_profile_admin(driver, partner):
        log(
            f"   [ADMIN] Partenaire introuvable (essayé : "
            f"{partner_search_queries(partner)})",
            "WARNING",
        )
        _save_login_debug(driver, "partner_not_found")
        for d in drivers:
            results.append(
                {
                    **d,
                    "admin": {
                        "partner_found": False,
                        "vehicle_assigned": False,
                        "fleet_status": "",
                        "plate": "",
                        "visual": "red",
                        "reason": "partner_not_found",
                    },
                },
            )
        return results

    if not open_fleet_tab(driver):
        for d in drivers:
            results.append(
                {
                    **d,
                    "admin": {
                        "partner_found": True,
                        "vehicle_assigned": False,
                        "visual": "red",
                        "reason": "fleet_tab_failed",
                    },
                },
            )
        return results

    fleet_rows = parse_fleet_detail_rows(driver)
    log(f"   [ADMIN] {len(fleet_rows)} ligne(s) flotte lues")

    if proof_sink is not None and len(fleet_rows) >= 1:
        inject_banner(
            driver,
            f"ADMIN — {len(fleet_rows)} ligne(s) flotte (preuve)",
            VIS_TEAL,
        )
        admin_proof = capture_list_proof_screenshot(
            driver,
            partner,
            phase="admin",
            row_count=len(fleet_rows),
        )
        if admin_proof:
            proof_sink["admin"] = admin_proof

    # Marquer visuellement toutes les lignes flotte (véhicule / conducteur)
    for fr in fleet_rows:
        dname = fr.get("driver_name", "")
        st = (fr.get("fleet_status") or "").upper()
        color = VIS_GREEN if dname and "APPROUV" in st else (
            VIS_ORANGE if dname else VIS_GRAY
        )
        label = "ASSIGNE" if dname else "SANS_CHAUFFEUR"
        try:
            for row_el in driver.find_elements(By.CSS_SELECTOR, "table tbody tr"):
                if dname and names_match(dname, row_el.text or ""):
                    highlight_row(driver, row_el, color, label)
                    break
                if not dname and fr.get("plate") and fr["plate"] in (row_el.text or ""):
                    highlight_row(driver, row_el, VIS_ORANGE, "VIDE")
                    break
        except Exception:
            pass

    # Si owner n'a rien lu, on déduit les chauffeurs depuis la flotte admin
    if not drivers:
        log("   [ADMIN] Liste owner vide — chauffeurs déduits depuis la flotte", "WARNING")
        seen: set[str] = set()
        for fr in fleet_rows:
            dn = (fr.get("driver_name") or "").strip()
            if not dn or dn in ("-", "—"):
                continue
            key = normalize_text(dn)
            if key in seen:
                continue
            seen.add(key)
            drivers.append(
                {
                    "name": dn,
                    "phone": "",
                    "vehicle_type": "",
                    "owner_status": fr.get("fleet_status", "?"),
                    "scraped_at": datetime.now().isoformat(timespec="seconds"),
                    "source": "admin_fleet_fallback",
                },
            )

    for d in drivers:
        dname = d.get("name", "")
        match_row = None
        for fr in fleet_rows:
            fn = fr.get("driver_name", "")
            if fn and names_match(dname, fn):
                match_row = fr
                break

        assigned = bool(match_row and match_row.get("driver_name"))
        visual = "gray"
        reason = "ok"

        if not match_row:
            visual = "red"
            reason = "driver_not_on_fleet_table"
        elif assigned:
            st = (match_row.get("fleet_status") or "").upper()
            if "APPROUV" in st:
                visual = "green"
                reason = "assigned_approved"
            elif "ATTENTE" in st:
                visual = "orange"
                reason = "assigned_pending"
            else:
                visual = "blue"
                reason = "assigned_unknown_status"
        else:
            visual = "orange"
            reason = "not_assigned"

        # Pastille sur la ligne flotte si trouvée
        if match_row:
            try:
                for row_el in driver.find_elements(By.CSS_SELECTOR, "table tbody tr"):
                    if names_match(dname, row_el.text or ""):
                        colors = {
                            "green": "#2e7d32",
                            "orange": "#ef6c00",
                            "red": "#c62828",
                            "blue": "#1565c0",
                            "gray": "#757575",
                        }
                        highlight_row(
                            driver,
                            row_el,
                            colors.get(visual, "#757575"),
                            label=reason[:20],
                        )
                        break
            except Exception:
                pass

        results.append(
            {
                **d,
                "admin": {
                    "partner_found": True,
                    "vehicle_assigned": assigned,
                    "fleet_status": (match_row or {}).get("fleet_status", ""),
                    "plate": (match_row or {}).get("plate", ""),
                    "matched_driver_on_row": (match_row or {}).get("driver_name", ""),
                    "visual": visual,
                    "reason": reason,
                    "checked_at": datetime.now().isoformat(timespec="seconds"),
                },
            },
        )
        icon = {"green": "OK", "orange": "!", "red": "X", "blue": "?"}.get(visual, "?")
        log(
            f"      [{icon}] {dname} | assigné={assigned} | "
            f"{(match_row or {}).get('fleet_status', '-')}",
        )

    return results


def load_state() -> dict[str, Any]:
    if STATE_FILE.exists():
        try:
            with STATE_FILE.open(encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"version": 1, "partners": {}, "meta": {}}


def save_state(state: dict[str, Any]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    state["meta"]["updated_at"] = datetime.now().isoformat(timespec="seconds")
    with STATE_FILE.open("w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)
    log(f"   Etat sauvegardé → {STATE_FILE}")


def build_campaign_slot_partners(
    start: int,
    end: int,
    *,
    excel_path: Path | None = None,
    email_template: str = DEFAULT_EMAIL_TEMPLATE,
    password_default: str = DEFAULT_PARTNER_PASSWORD,
    limit: int = 0,
) -> list[dict[str, Any]]:
    """Slots Campagne start…end dans l'ordre ; Excel surcharge login si PARTENAIRE=n."""
    excel_by_slot: dict[int, dict[str, Any]] = {}
    if excel_path and excel_path.is_file():
        for row in load_partners_from_excel(excel_path):
            idx = int(row["index"])
            if start <= idx <= end:
                excel_by_slot[idx] = row

    partners: list[dict[str, Any]] = []
    for idx in range(start, end + 1):
        if idx in excel_by_slot:
            p = dict(excel_by_slot[idx])
            p["name"] = f"Campagne UPJUNOO {idx}"
            p["login_from_excel"] = True
        else:
            p = {
                "index": idx,
                "email": email_template.format(index=idx),
                "password": password_default,
                "name": f"Campagne UPJUNOO {idx}",
                "login_from_excel": False,
            }
        partners.append(p)
        if limit and len(partners) >= limit:
            break
    return partners


def load_partners_from_excel(path: Path, *, limit: int = 0) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        num, email, password = (list(row) + [None, None, None])[:3]
        if num is None:
            continue
        try:
            idx = int(num)
        except (TypeError, ValueError):
            continue
        em = str(email or "").strip()
        if not em:
            em = DEFAULT_EMAIL_TEMPLATE.format(index=idx)
        pwd = str(password or "").strip() or DEFAULT_PARTNER_PASSWORD
        out.append(
            {
                "index": idx,
                "email": em,
                "password": pwd,
                "name": f"Partenaire {idx}",
            },
        )
        if limit and len(out) >= limit:
            break
    return sorted(out, key=lambda p: p["index"])


def load_partner_config_overrides(config_path: Path | None = None) -> dict[int, dict[str, Any]]:
    """Lit partner_automation_config.json (profile_uuid, name, …)."""
    path = config_path or Path(os.getenv("PARTNER_CONFIG", str(DEFAULT_CONFIG_JSON)))
    if not path.is_file():
        return {}
    try:
        with path.open(encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return {}
    out: dict[int, dict[str, Any]] = {}
    for item in data.get("partners") or data.get("partners_example") or []:
        try:
            idx = int(item.get("index"))
        except (TypeError, ValueError):
            continue
        out[idx] = item
    return out


def apply_partner_overrides(
    partners: list[dict[str, Any]],
    config_path: Path | None = None,
) -> None:
    overrides = load_partner_config_overrides(config_path)
    for p in partners:
        o = overrides.get(p.get("index"))
        if not o:
            continue
        if o.get("name"):
            p["name"] = o["name"]
        if o.get("profile_uuid"):
            p["profile_uuid"] = o["profile_uuid"]
        if o.get("email"):
            p["email"] = o["email"]


def load_partners(
    args: argparse.Namespace,
    config_path: Path | None = None,
    *,
    state: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    excel_path: Path | None = None
    if not getattr(args, "no_excel", False):
        raw = (getattr(args, "excel", None) or "").strip()
        if raw:
            excel_path = Path(raw)
    partners = build_campaign_slot_partners(
        args.start,
        args.end,
        excel_path=excel_path,
        email_template=args.email_template,
        password_default=args.password,
        limit=args.limit or 0,
    )
    apply_partner_overrides(partners, config_path)
    if state:
        enrich_partners_from_state(partners, state)
    return partners


def process_partner(
    driver: webdriver.Chrome,
    partner: dict[str, Any],
    *,
    top_n: int,
    force: bool,
    show_ui: bool = False,
) -> dict[str, Any]:
    key = str(partner["index"])
    log(f"\n{'='*50}")
    log(f"PARTENAIRE {key} | {partner.get('email')} | {partner.get('name', '')}")
    log(f"{'='*50}")

    record: dict[str, Any] = {
        "index": partner["index"],
        "email": partner["email"],
        "name": partner.get("name", ""),
        "password_hint": "***",
        "last_owner_scan": None,
        "last_admin_check": None,
        "drivers": [],
        "errors": [],
    }

    # ── Phase A : owner / fleet-drivers ──
    if not owner_login(driver, partner["email"], partner["password"]):
        record["errors"].append("owner_login_failed")
        return record

    display = scrape_owner_display_name(driver)
    if display:
        partner["display_name"] = display
        record["display_name"] = display
        log(f"   [OWNER] Nom affiché dashboard : {display}")

    owner_scrape_ok = False
    owner_drivers: list[dict[str, Any]] = []
    try:
        owner_drivers = scrape_fleet_drivers_top(
            driver,
            top_n=top_n,
            visual=bool(show_ui),
        )
        owner_scrape_ok = True
        record["drivers"] = owner_drivers
        record["last_owner_scan"] = datetime.now().isoformat(timespec="seconds")
        log(f"   [OWNER] Scan OK — {len(owner_drivers)} chauffeur(s) sur fleet-drivers")
        if len(owner_drivers) > 0:
            owner_proof = capture_list_proof_screenshot(
                driver,
                partner,
                phase="owner",
                row_count=len(owner_drivers),
            )
            if owner_proof:
                record.setdefault("proof_screenshots", {})["owner"] = owner_proof
    except Exception as e:
        record["errors"].append(f"owner_scrape:{e}")
        log(traceback.format_exc(), "ERROR")
        if is_browser_dead(e):
            raise

    clear_session(driver)

    # ── Phase B : admin / flotte ──
    if not admin_login(driver):
        record["errors"].append("admin_login_failed")
        return record

    try:
        proof_sink: dict[str, str] = {}
        checked = admin_check_drivers(
            driver,
            partner,
            record.get("drivers") or [],
            proof_sink=proof_sink,
        )
        if owner_scrape_ok and len(owner_drivers) == 0:
            record["drivers"] = []
            log(
                "   [SYNC] fleet-drivers owner vide — state aligné sur le site (0 chauffeur, "
                "sans reprise admin ni ancienne liste)",
            )
        else:
            record["drivers"] = checked
        record["last_admin_check"] = datetime.now().isoformat(timespec="seconds")
        if owner_scrape_ok:
            record["drivers_replace_from_scan"] = True
        if record.get("proof_screenshots") or proof_sink:
            proofs = dict(record.get("proof_screenshots") or {})
            proofs.update(proof_sink)
            proofs["captured_at"] = datetime.now().isoformat(timespec="seconds")
            record["proof_screenshots"] = proofs
        elif owner_scrape_ok:
            record.pop("proof_screenshots", None)
    except Exception as e:
        record["errors"].append(f"admin_check:{e}")
        log(traceback.format_exc(), "ERROR")
        if is_browser_dead(e):
            raise

    clear_session(driver)
    return record


def print_summary(state: dict[str, Any]) -> None:
    total_d = 0
    ok = pending = missing = 0
    for p in state.get("partners", {}).values():
        for d in p.get("drivers") or []:
            total_d += 1
            adm = d.get("admin") or {}
            if adm.get("vehicle_assigned") and "APPROUV" in (adm.get("fleet_status") or "").upper():
                ok += 1
            elif adm.get("vehicle_assigned"):
                pending += 1
            else:
                missing += 1
    log(f"\nBILAN | partenaires={len(state.get('partners', {}))} | chauffeurs={total_d}")
    log(f"   Assignés OK: {ok} | assignés autre statut: {pending} | non assignés / absent: {missing}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Orchestrateur web partenaires UpJunoo")
    parser.add_argument("--start", type=int, default=1, help="Index premier partenaire (défaut 1)")
    parser.add_argument("--end", type=int, default=20, help="Index dernier partenaire (défaut 20)")
    parser.add_argument("--only", type=int, help="Traiter un seul index partenaire")
    parser.add_argument("--excel", help="DOSSIER_PARTENAIRES.xlsx (sinon campagne{N}@upjunoo.com)")
    parser.add_argument("--email-template", default=DEFAULT_EMAIL_TEMPLATE)
    parser.add_argument("--password", default=DEFAULT_PARTNER_PASSWORD)
    parser.add_argument("--top", type=int, default=DRIVERS_TOP_N, help="Nb chauffeurs fleet-drivers")
    parser.add_argument("--headed", action="store_true", help="Navigateur visible")
    parser.add_argument("--loop", action="store_true", help="Boucle continue (temps réel)")
    parser.add_argument("--interval", type=int, default=120, help="Pause entre tours complets (s)")
    parser.add_argument("--limit", type=int, default=0, help="Limiter nb partenaires (test)")
    parser.add_argument("--force", action="store_true", help="Re-traiter même si scan récent")
    parser.add_argument(
        "--config",
        default=str(DEFAULT_CONFIG_JSON),
        help="JSON avec profile_uuid par index (ex. partner_automation_config.json)",
    )
    args = parser.parse_args()
    config_path = Path(args.config)

    if args.only is not None:
        args.start = args.end = args.only

    partners = load_partners(args, config_path)
    if not partners:
        log("Aucun partenaire à traiter.", "ERROR")
        sys.exit(2)

    log("SWEEP PARTENAIRES WEB")
    log(f"   Partenaires: {len(partners)} | top {args.top} chauffeurs/partenaire")
    log(f"   Owner: {OWNER_LOGIN_URL}")
    log(f"   Fleet: {FLEET_DRIVERS_URL}")
    log(f"   Etat: {STATE_FILE}")
    if args.loop:
        log(f"   Mode boucle — intervalle {args.interval}s (Ctrl+C pour arrêter)")

    driver = make_driver(headed=args.headed)
    state = load_state()
    if "partners" not in state:
        state["partners"] = {}

    try:
        round_num = 0
        while True:
            round_num += 1
            if args.loop:
                log(f"\n### TOUR GLOBAL {round_num} ###")

            total_slots = len(partners)
            for slot_pos, partner in enumerate(partners, start=1):
                key = str(partner["index"])
                log(
                    f"\n── Campagne {slot_pos}/{total_slots} "
                    f"(n°{partner['index']}) — {partner.get('name', '')} ──",
                )
                try:
                    record = process_partner(
                        driver,
                        partner,
                        top_n=args.top,
                        force=args.force,
                        show_ui=bool(args.headed),
                    )
                    previous = state["partners"].get(key)
                    state["partners"][key] = merge_partner_record(previous, record)
                    n_done = sum(
                        1
                        for d in state["partners"][key].get("drivers") or []
                        if d.get("transfer_2000_done")
                    )
                    if n_done:
                        log(f"   Etat fusionné — {n_done} transfert(s) 2000 déjà marqué(s) conservé(s)")
                    save_state(state)
                except Exception as e:
                    if is_browser_dead(e):
                        log("Session navigateur perdue — redémarrage Chrome…", "WARNING")
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        driver = make_driver(headed=args.headed)
                        ent = state["partners"].setdefault(
                            key,
                            {"index": partner["index"], "errors": []},
                        )
                        ent["errors"] = list(ent.get("errors") or []) + [
                            "browser_session_perdue",
                        ]
                        save_state(state)
                        continue
                    log(f"Erreur partenaire {key}: {e}", "ERROR")
                    state["partners"].setdefault(key, {"index": partner["index"], "errors": []})
                    state["partners"][key]["errors"] = state["partners"][key].get(
                        "errors", [],
                    ) + [str(e)]
                    save_state(state)

            print_summary(state)

            if not args.loop:
                break
            log(f"Pause {args.interval}s avant prochain tour…")
            time.sleep(args.interval)

    except KeyboardInterrupt:
        log("Interruption Ctrl+C", "WARNING")
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    print_summary(state)
    log(f"Log: {LOG_FILE}")


if __name__ == "__main__":
    main()
