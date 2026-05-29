#!/usr/bin/env python3
"""Renomme les XLSX partenaires : NomPartenaire(Nchauf).xlsx depuis _RECAP_partenaires.xlsx."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def sanitize_filename(name: str, max_len: int = 180) -> str:
    name = (name or "").strip() or "sans_partenaire"
    name = re.sub(r'[\\/:*?"<>|]', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:max_len] or "sans_partenaire"


def partner_label(partner: str, count: int) -> str:
    return f"{partner}({int(count)}chauf)"


def autosize_columns(ws, max_width: int = 52) -> None:
    for col_idx, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, max(12, length + 2))


def apply_prefixes(folder: Path, recap_name: str = "_RECAP_partenaires.xlsx") -> None:
    recap_path = folder / recap_name
    if not recap_path.is_file():
        raise FileNotFoundError(f"Récap introuvable: {recap_path}")

    wb = load_workbook(recap_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    renamed = 0
    skipped = 0
    updated_rows: list[tuple[str, int, str, str]] = []

    for partner, nb, old_file in rows:
        partner = (partner or "").strip()
        if not partner:
            continue
        try:
            count = int(nb)
        except (TypeError, ValueError):
            count = 0
        old_name = (old_file or "").strip()
        if not old_name:
            old_name = sanitize_filename(partner) + ".xlsx"

        label = partner_label(partner, count)
        new_name = sanitize_filename(label) + ".xlsx"
        src = folder / old_name
        dst = folder / new_name

        if src.is_file():
            if src.resolve() != dst.resolve():
                if dst.exists():
                    raise FileExistsError(f"Conflit: {dst} existe déjà")
                src.rename(dst)
                renamed += 1
        elif dst.is_file():
            skipped += 1
        else:
            print(f"⚠️ Fichier absent: {old_name}")
            skipped += 1

        updated_rows.append((label, count, new_name, partner))

    # Récap mis à jour
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Récap"
    out_ws.append(["Partenaire", "Nb conducteurs", "Fichier", "Nom original"])
    for label, count, fname, orig in sorted(updated_rows, key=lambda x: (-x[1], x[3].lower())):
        out_ws.append([label, count, fname, orig])
    autosize_columns(out_ws)
    out_wb.save(recap_path)

    print(f"Dossier: {folder}")
    print(f"Renommés: {renamed} | déjà OK / ignorés: {skipped}")
    print(f"Récap: {recap_path}")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument(
        "-d",
        "--dir",
        default="output/fleet-drivers-par-partenaire",
        help="Dossier des XLSX partenaires",
    )
    args = p.parse_args()
    apply_prefixes(Path(args.dir).resolve())


if __name__ == "__main__":
    main()
