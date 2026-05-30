#!/usr/bin/env python3
"""
export_partner_fleet_drivers_report.py
======================================

Rapport admin : flotte + chauffeurs par campagne (1→20). JSON + Excel.
Script autonome (ne dépend pas de partner_fleet_orchestrator.py).

Usage :
  python export_partner_fleet_drivers_report.py --start 1 --end 20 --headed
  python export_partner_fleet_drivers_report.py --only 3 --headed

Partenaires par défaut : campagne1@upjunoo.com … campagne20@upjunoo.com
(option --excel seulement pour surcharger avec DOSSIER_PARTENAIRES.xlsx)
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import (
    InvalidSessionIdException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait

import quick_approve_all_vehicle_vps as qa

# ─── Config ───────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "output" / "partner_automation"
STATE_FILE = OUTPUT_DIR / "state.json"
LOG_FILE = OUTPUT_DIR / "rapport_export.log"
DEFAULT_CONFIG_JSON = SCRIPT_DIR / "partner_automation_config.json"

BASE_URL = os.getenv("UPJUNOO_BASE_URL", "https://upjunoo-server-new.junooapps.com")
MANAGE_OWNERS_URL = f"{BASE_URL}/manage-owners"
VIEW_PROFILE_URL = f"{BASE_URL}/manage-owners/view-profile"
LOGOUT_URL = f"{BASE_URL}/logout"

DEFAULT_EMAIL_TEMPLATE = os.getenv("PARTNER_EMAIL_TEMPLATE", "campagne{index}@upjunoo.com")
DEFAULT_PARTNER_PASSWORD = os.getenv("PARTNER_PASSWORD", "123456789@")
ADMIN_EMAIL = os.getenv("UPJUNOO_EMAIL", "admin@upjunoo.com")

VIS_GREEN = "#2e7d32"
VIS_ORANGE = "#ef6c00"
VIS_RED = "#c62828"
VIS_BLUE = "#1565c0"
VIS_TEAL = "#00695c"
VIS_PURPLE = "#6a1b9a"


def log(msg: str, level: str = "INFO") -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}][{level}] {msg}"
    try:
        print(line, flush=True)
    except UnicodeEncodeError:
        enc = getattr(sys.stdout, "encoding", None) or "utf-8"
        print(line.encode(enc, errors="replace").decode(enc, errors="replace"), flush=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with LOG_FILE.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def is_browser_dead(exc: BaseException) -> bool:
    if isinstance(exc, InvalidSessionIdException):
        return True
    if isinstance(exc, WebDriverException):
        m = str(exc).lower()
        return any(
            k in m
            for k in ("invalid session", "session deleted", "disconnected", "not reachable")
        )
    return False


def wait_for_table(driver: webdriver.Chrome, timeout: int = 30, min_rows: int = 1) -> bool:
    ignore = ("chargement", "loading", "aucune", "no data", "vide")
    end = time.time() + timeout
    while time.time() < end:
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            if len(rows) >= min_rows:
                txt = (rows[0].text or "").lower()
                if not any(w in txt for w in ignore):
                    return True
        except Exception:
            pass
        time.sleep(0.8)
    return False


def inject_banner(driver: webdriver.Chrome, text: str, color: str = VIS_BLUE) -> None:
    try:
        driver.execute_script(
            """
            let el = document.getElementById('pa-export-banner');
            if (!el) {
                el = document.createElement('div');
                el.id = 'pa-export-banner';
                document.body.prepend(el);
            }
            el.style.cssText = [
                'position:fixed','top:0','left:0','right:0','z-index:99999',
                'padding:10px 14px','color:#fff','font:600 13px sans-serif',
                'background:' + arguments[1], 'box-shadow:0 2px 8px rgba(0,0,0,.25)'
            ].join(';');
            el.textContent = arguments[0];
            document.body.style.marginTop = '48px';
            """,
            text,
            color,
        )
    except Exception:
        pass


def inject_progress_panel(
    driver: webdriver.Chrome,
    *,
    slot: int,
    total: int,
    step: str,
    detail: str,
    color: str = VIS_BLUE,
) -> None:
    try:
        driver.execute_script(
            """
            let root = document.getElementById('pa-progress-panel');
            if (!root) {
                root = document.createElement('div');
                root.id = 'pa-progress-panel';
                root.style.cssText = [
                    'position:fixed','top:52px','right:12px','z-index:99998',
                    'min-width:240px','padding:12px 14px','border-radius:8px',
                    'color:#fff','font:600 12px/1.45 sans-serif',
                    'box-shadow:0 4px 14px rgba(0,0,0,.35)'
                ].join(';');
                document.body.appendChild(root);
            }
            root.style.background = arguments[4];
            const pct = arguments[1] > 0
                ? Math.round(100 * arguments[0] / arguments[1]) : 0;
            root.innerHTML =
                '<div style="font-size:11px;opacity:.92">EXPORT RAPPORT</div>'
                + '<div style="font-size:17px;margin:6px 0">Campagne '
                + arguments[0] + ' &rarr; ' + arguments[1] + '</div>'
                + '<div style="background:rgba(255,255,255,.28);height:7px;border-radius:4px;margin:8px 0">'
                + '<div style="background:#fff;height:100%;width:' + pct
                + '%;border-radius:4px"></div></div>'
                + '<div style="font-size:13px">' + arguments[2] + '</div>'
                + '<div style="font-size:11px;margin-top:5px;opacity:.88">'
                + arguments[3] + '</div>';
            """,
            slot,
            total,
            step,
            detail,
            color,
        )
    except Exception:
        pass


def flash_viewport_border(driver: webdriver.Chrome, color: str, *, active: bool = True) -> None:
    try:
        driver.execute_script(
            "document.body.style.outline = arguments[0] ? ('5px solid ' + arguments[1]) : '';",
            active,
            color,
        )
    except Exception:
        pass


def show_phase_ui(
    driver: webdriver.Chrome,
    *,
    partner: dict[str, Any],
    slot_pos: int,
    total_slots: int,
    step: str,
    detail: str,
    color: str,
    show_ui: bool,
) -> None:
    if not show_ui:
        return
    idx = partner.get("index", "?")
    name = partner.get("name") or f"Campagne UPJUNOO {idx}"
    inject_progress_panel(
        driver,
        slot=slot_pos,
        total=total_slots,
        step=step,
        detail=f"{name} | {detail}",
        color=color,
    )
    inject_banner(
        driver,
        f"▶ Campagne {slot_pos}/{total_slots} (n°{idx}) — {step}",
        color,
    )
    flash_viewport_border(driver, color, active=True)


def log_campaign_plan(partners: list[dict[str, Any]], *, start: int, end: int) -> None:
    log(f"   Ordre : Campagne {start} → {end} ({len(partners)} slot(s))")
    for p in partners:
        src = "Excel" if p.get("login_from_excel") else "campagne@upjunoo"
        em = p.get("email", "")
        mask = em[:4] + "…@" + em.split("@", 1)[1] if "@" in em else em
        log(f"      [{p['index']:>2}] {p.get('name', '')} | {mask} ({src})")


def _normalize_ui_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def _is_empty_data_message(text: str) -> bool:
    """Message DataTables « vraiment vide » (pas confondre avec 10 entrées / 500 entrées)."""
    low = _normalize_ui_text(text)
    if not low:
        return False
    if re.search(r"aucune\s+donn[eé]e\s+trouv[eé]e", low):
        return True
    if re.search(r"no\s+(matching\s+)?records?\s+found", low):
        return True
    if re.search(r"affichage\s+0\s+de\s+0", low):
        return True
    if re.search(r"affichage\s+(?:\d+\s+[àa]\s+)?0\s+de\s+0\s+entr", low):
        return True
    return False


def _is_placeholder_row(row: Any) -> bool:
    try:
        txt = (row.text or "").strip()
        if not txt:
            return True
        if _is_empty_data_message(txt):
            return True
        low = txt.lower()
        if any(w in low for w in ("chargement", "loading")):
            return True
        cells = row.find_elements(By.TAG_NAME, "td")
        if len(cells) == 1 and (cells[0].get_attribute("colspan") or "").strip():
            cell_txt = (cells[0].text or "").strip()
            if _is_empty_data_message(cell_txt) or not cell_txt:
                return True
        if len(cells) < 3:
            return True
        if not any((c.text or "").strip() not in ("", "-", "—", "N/A") for c in cells):
            return True
        first = (cells[0].text or "").strip().lower()
        if first in ("nom", "name", "#", "type de véhicule", "type de vehicule"):
            return True
    except Exception:
        return True
    return False


FLEET_RETRY_MAX = 3
FLEET_SUSPECT_PROFILE_RELOADS = 2


def _valid_data_rows_in_pane(pane) -> int:
    """Compte les lignes tbody réelles (hors en-tête / « Aucune donnée trouvée »)."""
    n = 0
    for row in pane.find_elements(By.CSS_SELECTOR, "table tbody tr"):
        if _is_placeholder_row(row):
            continue
        cells = row.find_elements(By.TAG_NAME, "td")
        if len(cells) >= 2 and any((c.text or "").strip() for c in cells):
            n += 1
    return n


def _pane_has_empty_data_row(pane) -> bool:
    """Une ligne unique « Aucune donnée trouvée » dans tbody."""
    rows = pane.find_elements(By.CSS_SELECTOR, "table tbody tr")
    if not rows:
        return False
    for row in rows:
        txt = (row.text or "").strip()
        if _is_empty_data_message(txt):
            return True
        cells = row.find_elements(By.TAG_NAME, "td")
        if len(cells) == 1:
            if _is_empty_data_message((cells[0].text or "").strip()):
                return True
    return False


def _pane_text_confirms_empty(txt: str) -> bool:
    """Texte du panneau confirme un tableau vide réel (strict)."""
    return _is_empty_data_message(txt)


def _get_active_tab_pane(driver: webdriver.Chrome):
    for sel in (
        ".tab-pane.active.show",
        ".tab-pane.active",
        ".tab-content .tab-pane.show.active",
        "[role='tabpanel'].active",
    ):
        for el in driver.find_elements(By.CSS_SELECTOR, sel):
            try:
                if el.is_displayed():
                    return el
            except Exception:
                continue
    return None


def _active_tab_has_no_data(driver: webdriver.Chrome) -> bool:
    """
    True seulement si le vide est confirmé (« Aucune donnée trouvée », affichage 0 de 0).
    Sinon False → on tente quand même le parse (évite les faux vides).
    """
    pane = _get_active_tab_pane(driver)
    if not pane:
        return False
    if _valid_data_rows_in_pane(pane) > 0:
        return False
    if _pane_has_empty_data_row(pane):
        return True
    return _pane_text_confirms_empty(pane.text or "")


def fleet_tab_confirmed_empty(driver: webdriver.Chrome) -> bool:
    """Indique si l'onglet flotte affiche un vide réel UpJunoo."""
    return _active_tab_has_no_data(driver)


def _get_active_table_rows(driver: webdriver.Chrome) -> list[Any]:
    if _active_tab_has_no_data(driver):
        return []
    pane = _get_active_tab_pane(driver)
    if pane:
        return pane.find_elements(By.CSS_SELECTOR, "table tbody tr")
    for sel in (".tab-pane.active.show table tbody tr", ".tab-pane.active table tbody tr"):
        rows = driver.find_elements(By.CSS_SELECTOR, sel)
        if rows:
            return rows
    return []


def wait_for_tab_table(driver: webdriver.Chrome, timeout: int = 30) -> bool:
    """Attend données ou état vide stable dans l'onglet actif du profil partenaire."""
    end = time.time() + timeout
    empty_streak = 0
    while time.time() < end:
        pane = _get_active_tab_pane(driver)
        if pane and _valid_data_rows_in_pane(pane) > 0:
            return True
        if _active_tab_has_no_data(driver):
            empty_streak += 1
            if empty_streak >= 3:
                return True
        else:
            empty_streak = 0
        time.sleep(0.5)
    return _active_tab_has_no_data(driver)


def is_fleet_scrape_suspect(record: dict[str, Any]) -> bool:
    """
    0 véhicule + chauffeurs présents sans message « Aucune donnée trouvée »
    → probable faux négatif (bug timing / parse).
    """
    if record.get("fleet_confirmed_empty"):
        return False
    return int(record.get("vehicles_count") or 0) == 0 and int(record.get("drivers_count") or 0) > 0


def scrape_fleet_vehicles(
    driver: webdriver.Chrome,
    *,
    max_attempts: int = FLEET_RETRY_MAX,
) -> tuple[list[dict[str, str]], list[str], bool]:
    """Ouvre l'onglet flotte et parse les véhicules (plusieurs essais)."""
    errors: list[str] = []
    vehicles: list[dict[str, str]] = []
    confirmed_empty = False
    for attempt in range(1, max_attempts + 1):
        if not open_fleet_tab(driver):
            errors.append("fleet_tab_failed")
            if attempt < max_attempts:
                log(f"   Flotte : onglet inaccessible — essai {attempt + 1}/{max_attempts}", "WARNING")
                time.sleep(2.5)
            continue
        if fleet_tab_confirmed_empty(driver):
            confirmed_empty = True
            log("   Flotte : « Aucune donnée trouvée » — vide réel confirmé", "OK")
            return [], errors, True
        vehicles = parse_fleet_detail_rows(driver)
        if vehicles:
            if attempt > 1:
                log(
                    f"   ✓ Flotte récupérée (essai {attempt}/{max_attempts}) — {len(vehicles)} véhicule(s)",
                    "OK",
                )
            return vehicles, errors, False
        if fleet_tab_confirmed_empty(driver):
            confirmed_empty = True
            log("   Flotte : « Aucune donnée trouvée » — vide réel confirmé", "OK")
            return [], errors, True
        if attempt < max_attempts:
            log(
                f"   Flotte : 0 ligne sans message vide (essai {attempt}/{max_attempts}) — retry…",
                "WARNING",
            )
            time.sleep(2.5)
    return vehicles, errors, confirmed_empty


def retry_fleet_if_suspect(
    driver: webdriver.Chrome,
    partner: dict[str, Any],
    record: dict[str, Any],
) -> None:
    """Re-scrape flotte si 0 véh. alors que des chauffeurs existent."""
    if not is_fleet_scrape_suspect(record):
        return
    idx = partner.get("index")
    log(
        f"   ⚠ Campagne {idx} suspecte : 0 véh. / {record['drivers_count']} chauf. — retry flotte",
        "WARNING",
    )
    for reload in range(1, FLEET_SUSPECT_PROFILE_RELOADS + 1):
        if partner.get("profile_uuid"):
            open_partner_profile_by_uuid(driver, str(partner["profile_uuid"]))
            time.sleep(2)
        vehicles, errs = scrape_fleet_vehicles(driver, max_attempts=FLEET_RETRY_MAX)
        for e in errs:
            if e not in record["errors"]:
                record["errors"].append(e)
        if vehicles:
            record["vehicles"] = vehicles
            record["vehicles_count"] = len(vehicles)
            record["fleet_confirmed_empty"] = False
            if "fleet_suspect_empty" in record["errors"]:
                record["errors"].remove("fleet_suspect_empty")
            log(f"   ✓ Flotte corrigée après retry : {len(vehicles)} véhicule(s)", "OK")
            attach_status_breakdown(record)
            return
        if fleet_tab_confirmed_empty(driver):
            record["fleet_confirmed_empty"] = True
            if "fleet_suspect_empty" in record["errors"]:
                record["errors"].remove("fleet_suspect_empty")
            log("   Flotte vide réelle confirmée au retry — pas d'erreur", "OK")
            attach_status_breakdown(record)
            return
        log(f"   Retry flotte {reload}/{FLEET_SUSPECT_PROFILE_RELOADS} — toujours vide", "WARNING")
    if "fleet_suspect_empty" not in record["errors"]:
        record["errors"].append("fleet_suspect_empty")
    attach_status_breakdown(record)


def list_suspect_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [r for r in records if is_fleet_scrape_suspect(r)]


def validate_and_rescrape_suspects(
    driver: webdriver.Chrome,
    partners: list[dict[str, Any]],
    records: list[dict[str, Any]],
) -> int:
    """Validation finale : re-scrape flotte pour les campagnes suspectes."""
    by_idx = {int(p.get("index") or 0): p for p in partners}
    fixed = 0
    for record in list(records):
        if not is_fleet_scrape_suspect(record):
            continue
        idx = int(record.get("index") or 0)
        partner = by_idx.get(idx)
        if not partner:
            continue
        log(f"\n── Validation : re-scrape flotte campagne {idx} ──")
        before = record["vehicles_count"]
        retry_fleet_if_suspect(driver, partner, record)
        if record["vehicles_count"] > before:
            fixed += 1
    remaining = list_suspect_records(records)
    if remaining:
        ids = ", ".join(f"P{int(r['index']):02d}" for r in remaining)
        log(f"   ⚠ Campagnes encore suspectes après validation : {ids}", "WARNING")
    else:
        log("   ✓ Validation flotte : aucune campagne suspecte restante", "OK")
    return fixed


def set_page_size(
    driver: webdriver.Chrome,
    size: int = 500,
    *,
    active_tab_only: bool = False,
) -> bool:
    try:
        time.sleep(1.0)
        roots: list[Any] = []
        if active_tab_only:
            pane = _get_active_tab_pane(driver)
            if pane:
                roots.append(pane)
        if not roots:
            roots.append(driver)
        for root in roots:
            selects = root.find_elements(
                By.CSS_SELECTOR,
                "select.form-select-sm, select[name*='length'], .dataTables_length select, select.form-select",
            )
            for el in selects:
                try:
                    if not el.is_displayed():
                        continue
                    s = Select(el)
                    opts = [
                        o for o in s.options if (o.get_attribute("value") or "").strip() != ""
                    ]
                    if not opts:
                        continue
                    values = [o.get_attribute("value") for o in opts]
                    texts = [(o.text or "").strip() for o in opts]
                    target = str(size)
                    if target in values:
                        s.select_by_value(target)
                    else:
                        for pref in (str(size), "100", "50", "25", "10", "All", "Tout"):
                            if pref in texts:
                                s.select_by_visible_text(pref)
                                break
                        else:
                            s.select_by_index(len(opts) - 1)
                    log(f"   [PAGE] Pagination → {s.first_selected_option.text.strip()}", "OK")
                    time.sleep(2.5)
                    if active_tab_only:
                        wait_for_tab_table(driver, timeout=30)
                    else:
                        wait_for_table(driver, timeout=30, min_rows=1)
                    return True
                except Exception:
                    continue
        return False
    except Exception as e:
        log(f"   Pagination: {e}", "WARNING")
        return False


def try_search(driver: webdriver.Chrome, query: str) -> bool:
    if not (query or "").strip():
        return False
    try:
        box = WebDriverWait(driver, 8).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "input[type='search'], .dataTables_filter input"),
            ),
        )
        box.clear()
        time.sleep(0.2)
        box.send_keys(query)
        time.sleep(2.5)
        return True
    except Exception:
        return False


def wait_table_row_count_stable(driver: webdriver.Chrome, checks: int = 3, pause_s: float = 2) -> int:
    last_count = -1
    stable_hits = 0
    for _ in range(checks * 3):
        try:
            count = len(driver.find_elements(By.CSS_SELECTOR, "table tbody tr"))
            if count == last_count and count > 0:
                stable_hits += 1
                if stable_hits >= checks:
                    return count
            else:
                stable_hits = 0
            last_count = count
        except Exception:
            stable_hits = 0
        time.sleep(pause_s)
    return last_count


def find_row_by_cell_text_with_pagination(
    driver: webdriver.Chrome,
    target_text: str,
    max_pages: int = 25,
) -> Any | None:
    q = (target_text or "").strip().lower()
    if not q:
        return None
    next_btn_xpath = "//li[contains(@class, 'next') and not(contains(@class, 'disabled'))]/a"
    for _ in range(max_pages):
        if not wait_for_table(driver, timeout=15, min_rows=1):
            break
        for row in driver.find_elements(By.CSS_SELECTOR, "table tbody tr"):
            if not row.is_displayed():
                continue
            try:
                for cell in row.find_elements(By.TAG_NAME, "td"):
                    if q in (cell.text or "").strip().lower():
                        return row
            except Exception:
                continue
        try:
            nxt = driver.find_element(By.XPATH, next_btn_xpath)
            if nxt.is_displayed():
                driver.execute_script("arguments[0].click();", nxt)
                time.sleep(2.5)
                continue
        except Exception:
            pass
        break
    return None


def partner_search_queries(partner: dict[str, Any]) -> list[str]:
    idx = partner.get("index")
    email = (partner.get("email") or "").strip()
    name = (partner.get("name") or "").strip()
    display = (partner.get("display_name") or "").strip()
    queries: list[str] = []
    for q in (email, display, name):
        if q and q not in queries:
            queries.append(q)
    is_upjunoo_campagne = (
        email.lower().endswith("@upjunoo.com") and "campagne" in email.lower()
    )
    if is_upjunoo_campagne and idx is not None:
        for q in (
            f"Campagne UPJUNOO {idx}",
            f"Campagne UPJUNOO {idx}".upper(),
            f"campagne{idx}",
            f"partenaire{idx}",
            f"Partenaire {idx}",
        ):
            if q not in queries:
                queries.append(q)
    if email and "@" in email:
        local = email.split("@", 1)[0]
        if len(local) >= 4 and local not in queries:
            queries.append(local)
    return queries


def is_on_partner_profile(driver: webdriver.Chrome) -> bool:
    return "/manage-owners/view-profile/" in (driver.current_url or "").lower()


def extract_profile_uuid_from_row(row: Any) -> str:
    try:
        for link in row.find_elements(By.CSS_SELECTOR, "a[href*='view-profile']"):
            href = link.get_attribute("href") or ""
            m = re.search(r"/view-profile/([a-f0-9-]{36})", href, re.I)
            if m:
                return m.group(1).lower()
        for link in row.find_elements(By.CSS_SELECTOR, "a[href*='/document/']"):
            href = link.get_attribute("href") or ""
            m = re.search(r"/document/([a-f0-9-]{36})", href, re.I)
            if m:
                return m.group(1).lower()
    except Exception:
        pass
    return ""


def clear_session(driver: webdriver.Chrome) -> None:
    try:
        driver.get(LOGOUT_URL)
        time.sleep(0.8)
    except Exception:
        pass
    try:
        driver.delete_all_cookies()
    except Exception:
        pass


def admin_login(driver: webdriver.Chrome) -> bool:
    log(f"   [ADMIN] Connexion {ADMIN_EMAIL}")
    clear_session(driver)
    ok = qa.admin_login(driver, 0)
    if ok:
        log(f"   [ADMIN] OK → {driver.current_url}", "OK")
    else:
        log("   [ADMIN] Échec — vérifie UPJUNOO_EMAIL / UPJUNOO_PASSWORD", "ERROR")
    return ok


def navigate_manage_owners_table(driver: webdriver.Chrome) -> None:
    log("   [ADMIN] Ouverture manage-owners…")
    driver.get(MANAGE_OWNERS_URL)
    time.sleep(4)
    wait_for_table(driver, timeout=40)
    set_page_size(driver, 500)
    wait_for_table(driver, timeout=45, min_rows=1)
    time.sleep(4)
    n = wait_table_row_count_stable(driver, checks=3, pause_s=2)
    log(f"   [ADMIN] Tableau partenaires stabilisé (~{n} lignes visibles)")


def find_partner_row_admin(driver: webdriver.Chrome, partner: dict[str, Any]) -> Any | None:
    navigate_manage_owners_table(driver)
    for query in partner_search_queries(partner):
        log(f"   [ADMIN] Recherche partenaire : {query!r}")
        try_search(driver, "")
        time.sleep(0.5)
        try_search(driver, query)
        row = find_row_by_cell_text_with_pagination(driver, query, max_pages=25)
        if row:
            log(f"   [ADMIN] Partenaire trouvé via {query!r}", "OK")
            return row
    return None


def open_partner_profile_by_uuid(driver: webdriver.Chrome, profile_uuid: str) -> bool:
    uuid = (profile_uuid or "").strip()
    if not uuid:
        return False
    url = f"{VIEW_PROFILE_URL}/{uuid}"
    log(f"   [ADMIN] Profil direct → {url}")
    try:
        driver.get(url)
        time.sleep(4)
        WebDriverWait(driver, 30).until(lambda d: is_on_partner_profile(d))
        return is_on_partner_profile(driver)
    except Exception as e:
        log(f"   [ADMIN] Profil UUID échoué: {e}", "WARNING")
        return False


def open_partner_profile(driver: webdriver.Chrome, row: Any) -> bool:
    uuid = extract_profile_uuid_from_row(row)
    if uuid and open_partner_profile_by_uuid(driver, uuid):
        log(f"   [ADMIN] Profil ouvert (UUID) → {driver.current_url}", "OK")
        return True
    try:
        profile_btn = None
        try:
            profile_btn = row.find_element(By.XPATH, ".//a[contains(@href, 'view-profile')]")
        except Exception:
            pass
        if not profile_btn:
            try:
                profile_btn = row.find_element(
                    By.XPATH,
                    ".//a[contains(@href, 'profile') or contains(@href, 'view')]",
                )
            except Exception:
                pass
        if not profile_btn:
            return False
        href = (profile_btn.get_attribute("href") or "").strip()
        m = re.search(r"/view-profile/([a-f0-9-]{36})", href, re.I)
        if m:
            return open_partner_profile_by_uuid(driver, m.group(1).lower())
        log("   [ADMIN] Clic « voir profil »…")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", profile_btn)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", profile_btn)
        WebDriverWait(driver, 35).until(lambda d: is_on_partner_profile(d))
        time.sleep(2)
        ok = is_on_partner_profile(driver)
        if ok:
            log(f"   [ADMIN] Profil ouvert → {driver.current_url}", "OK")
        return ok
    except Exception as e:
        log(f"   [ADMIN] Profil: {e}", "WARNING")
        return False


def open_partner_profile_admin(driver: webdriver.Chrome, partner: dict[str, Any]) -> bool:
    uuid = (partner.get("profile_uuid") or "").strip()
    if uuid and open_partner_profile_by_uuid(driver, uuid):
        return True
    row = find_partner_row_admin(driver, partner)
    if not row:
        return False
    found_uuid = extract_profile_uuid_from_row(row)
    if found_uuid:
        partner["profile_uuid"] = found_uuid
        log(f"   [ADMIN] UUID mémorisé …{found_uuid[-8:]}", "OK")
        return open_partner_profile_by_uuid(driver, found_uuid)
    return open_partner_profile(driver, row)


def open_fleet_tab(driver: webdriver.Chrome) -> bool:
    if not is_on_partner_profile(driver):
        log("   [ADMIN] Onglet flotte refusé — pas sur view-profile", "WARNING")
        return False
    for sel in (
        "//a[contains(., 'Détails de la flotte')]",
        "//a[contains(., 'flotte') or contains(., 'Flotte')]",
        "//span[contains(., 'flotte') or contains(., 'Flotte')]",
    ):
        try:
            for el in driver.find_elements(By.XPATH, sel):
                if el.is_displayed():
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    driver.execute_script("arguments[0].click();", el)
                    time.sleep(4)
                    wait_for_tab_table(driver, timeout=30)
                    set_page_size(driver, 500, active_tab_only=True)
                    wait_for_tab_table(driver, timeout=25)
                    return True
        except Exception:
            continue
    return False


def parse_fleet_detail_rows(driver: webdriver.Chrome) -> list[dict[str, str]]:
    if _active_tab_has_no_data(driver):
        return []
    rows_data: list[dict[str, str]] = []
    for row in _get_active_table_rows(driver):
        if _is_placeholder_row(row):
            continue
        try:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) < 4:
                continue
            texts = [(c.text or "").strip() for c in cells]
            if not any(t for t in texts if t and t not in ("-", "—", "N/A")):
                continue
            joined = " ".join(texts).lower()
            if any(w in joined for w in ("aucune donnée", "affichage", "montrer ", " entrées")):
                continue
            plate = ""
            driver_name = ""
            status = ""
            for t in texts:
                tu = t.upper()
                if "APPROUV" in tu or "ATTENTE" in tu:
                    status = t
                elif t and t not in ("-", "—", "N/A") and len(t) > 2:
                    if re.search(r"[A-Z]{2,}[-\s]?\d", t.upper()) or len(t) <= 12:
                        if not plate and any(ch.isdigit() for ch in t):
                            plate = t
            for t in reversed(texts):
                if t in ("-", "—", ""):
                    continue
                if "APPROUV" in t.upper() or "ATTENTE" in t.upper():
                    continue
                if len(t) > 3 and not plate.endswith(t):
                    driver_name = t
                    break
            if not driver_name and len(texts) >= 8:
                driver_name = texts[7]
            if not plate and len(texts) >= 4:
                candidate = texts[3]
                if candidate and any(ch.isalnum() for ch in candidate):
                    plate = candidate
            if not status and len(texts) >= 6:
                status = texts[5]
            if not status:
                for t in texts:
                    if "APPROUV" in t.upper() or "ATTENTE" in t.upper():
                        status = t
                        break
            if not (plate or driver_name or status):
                continue
            rows_data.append(
                {
                    "plate": plate,
                    "driver_name": driver_name if driver_name not in ("-", "—") else "",
                    "fleet_status": normalize_status_label(status, " | ".join(texts), kind="fleet"),
                    "row_text": " | ".join(texts[:9]),
                },
            )
        except Exception:
            continue
    return rows_data


def open_driver_tab(driver: webdriver.Chrome) -> bool:
    if not is_on_partner_profile(driver):
        return False
    for sel in (
        "//a[contains(., 'Détails du conducteur')]",
        "//a[contains(., 'Driver Details')]",
        "//span[contains(., 'Détails du conducteur')]",
    ):
        try:
            for el in driver.find_elements(By.XPATH, sel):
                if el.is_displayed():
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    driver.execute_script("arguments[0].click();", el)
                    time.sleep(4)
                    wait_for_tab_table(driver, timeout=30)
                    set_page_size(driver, 500, active_tab_only=True)
                    wait_for_tab_table(driver, timeout=30)
                    return True
        except Exception:
            continue
    return False


def parse_driver_detail_rows(driver: webdriver.Chrome) -> list[dict[str, str]]:
    if _active_tab_has_no_data(driver):
        return []
    drivers: list[dict[str, str]] = []
    for row in _get_active_table_rows(driver):
        if _is_placeholder_row(row):
            continue
        try:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) < 2:
                continue
            name = (cells[0].text or "").strip()
            if not name or len(name) < 2:
                continue
            row_text = " | ".join((c.text or "").strip() for c in cells[:8])
            approval_raw = (cells[4].text or "").strip() if len(cells) > 4 else ""
            drivers.append(
                {
                    "name": name,
                    "phone": (cells[2].text or "").strip() if len(cells) > 2 else "",
                    "location": (cells[1].text or "").strip() if len(cells) > 1 else "",
                    "transport_type": (cells[3].text or "").strip() if len(cells) > 3 else "",
                    "approval_status": normalize_status_label(
                        approval_raw, row_text, kind="driver",
                    ),
                    "rating": (cells[5].text or "").strip() if len(cells) > 5 else "",
                    "vehicle_type": (cells[6].text or "").strip() if len(cells) > 6 else "",
                    "row_text": row_text,
                },
            )
        except Exception:
            continue
    return drivers


# ─── Répartition par statut ───────────────────────────────────────────────────

def normalize_status_label(raw: str, row_text: str = "", *, kind: str = "fleet") -> str:
    """Libellé homogène pour regrouper les statuts (flotte ou conducteur)."""
    src = (raw or "").strip()
    blob = f"{src} {row_text}".upper()
    if not src and not (row_text or "").strip():
        return "Non renseigné"
    if "DESAPPROUV" in blob or "DÉSAPPROUV" in f"{src} {row_text}":
        return "Désapprouvé"
    if "APPROUV" in blob:
        return "Approuvé"
    if "ATTENTE" in blob or "PENDING" in blob:
        return "En attente"
    if any(w in blob for w in ("REJET", "REFUS", "REJECT")):
        return "Rejeté"
    if kind == "driver" and not src:
        return "Non renseigné"
    return src if src else "Autre"


def status_breakdown(
    items: list[dict[str, Any]],
    field: str,
    *,
    kind: str,
) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in items:
        label = item.get(field) or normalize_status_label(
            "", item.get("row_text", ""), kind=kind,
        )
        if not label:
            label = "Non renseigné"
        counts[label] = counts.get(label, 0) + 1
    return dict(sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])))


def merge_status_breakdowns(parts: list[dict[str, int]]) -> dict[str, int]:
    merged: dict[str, int] = {}
    for part in parts:
        for label, n in part.items():
            merged[label] = merged.get(label, 0) + n
    return dict(sorted(merged.items(), key=lambda kv: (-kv[1], kv[0])))


def attach_status_breakdown(record: dict[str, Any]) -> None:
    record["vehicles_by_status"] = status_breakdown(
        record.get("vehicles") or [], "fleet_status", kind="fleet",
    )
    record["drivers_by_status"] = status_breakdown(
        record.get("drivers") or [], "approval_status", kind="driver",
    )


def format_status_breakdown(counts: dict[str, int]) -> str:
    if not counts:
        return "—"
    return ", ".join(f"{label}: {n}" for label, n in counts.items())


# ─── Partenaires (slots 1→20) ─────────────────────────────────────────────────

def load_state() -> dict[str, Any]:
    if STATE_FILE.exists():
        try:
            with STATE_FILE.open(encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"version": 1, "partners": {}, "meta": {}}


def save_state(state: dict[str, Any]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    state.setdefault("meta", {})["updated_at"] = datetime.now().isoformat(timespec="seconds")
    with STATE_FILE.open("w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)
    log(f"   UUID sauvegardés → {STATE_FILE}")


def load_partners_from_excel(path: Path, *, limit: int = 0) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        num, email, password = (list(row) + [None, None, None])[:3]
        if num is None:
            continue
        try:
            idx = int(num)
        except (TypeError, ValueError):
            continue
        em = str(email or "").strip() or DEFAULT_EMAIL_TEMPLATE.format(index=idx)
        pwd = str(password or "").strip() or DEFAULT_PARTNER_PASSWORD
        out.append({"index": idx, "email": em, "password": pwd})
        if limit and len(out) >= limit:
            break
    wb.close()
    return out


def build_campaign_slot_partners(
    start: int,
    end: int,
    *,
    excel_path: Path | None = None,
    email_template: str = DEFAULT_EMAIL_TEMPLATE,
    password_default: str = DEFAULT_PARTNER_PASSWORD,
    limit: int = 0,
) -> list[dict[str, Any]]:
    excel_by_slot: dict[int, dict[str, Any]] = {}
    if excel_path and excel_path.is_file():
        for row in load_partners_from_excel(excel_path):
            idx = int(row["index"])
            if start <= idx <= end:
                excel_by_slot[idx] = row

    partners: list[dict[str, Any]] = []
    for idx in range(start, end + 1):
        if idx in excel_by_slot:
            p = dict(excel_by_slot[idx])
            p["login_from_excel"] = True
        else:
            p = {
                "index": idx,
                "email": email_template.format(index=idx),
                "password": password_default,
                "login_from_excel": False,
            }
        p["name"] = f"Campagne UPJUNOO {idx}"
        partners.append(p)
        if limit and len(partners) >= limit:
            break
    return partners


def load_partner_config_overrides(config_path: Path) -> dict[int, dict[str, Any]]:
    if not config_path.is_file():
        return {}
    try:
        with config_path.open(encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return {}
    out: dict[int, dict[str, Any]] = {}
    for item in data.get("partners") or []:
        try:
            out[int(item["index"])] = item
        except (TypeError, ValueError, KeyError):
            continue
    return out


def apply_partner_overrides(partners: list[dict[str, Any]], config_path: Path) -> None:
    """profile_uuid uniquement — les emails restent campagne{N}@upjunoo.com."""
    overrides = load_partner_config_overrides(config_path)
    for p in partners:
        o = overrides.get(p.get("index"))
        if not o:
            continue
        if o.get("profile_uuid"):
            p["profile_uuid"] = o["profile_uuid"]


def enrich_partners_from_state(partners: list[dict[str, Any]], state: dict[str, Any]) -> None:
    pmap = state.get("partners") or {}
    for p in partners:
        prev = pmap.get(str(p.get("index")))
        if not prev:
            continue
        if prev.get("profile_uuid") and not p.get("profile_uuid"):
            p["profile_uuid"] = prev["profile_uuid"]
        if prev.get("display_name") and not p.get("display_name"):
            p["display_name"] = prev["display_name"]


def resolve_partners(args: argparse.Namespace) -> list[dict[str, Any]]:
    state = load_state()
    excel_path: Path | None = None
    if (getattr(args, "excel", None) or "").strip():
        excel_path = Path(args.excel)
    partners = build_campaign_slot_partners(
        args.start,
        args.end,
        excel_path=excel_path,
        email_template=args.email_template,
        password_default=args.password,
        limit=args.limit or 0,
    )
    apply_partner_overrides(partners, Path(args.config))
    enrich_partners_from_state(partners, state)
    if args.only is not None:
        partners = [p for p in partners if p["index"] == args.only]
    return partners


def capture_profile_meta(driver: webdriver.Chrome, partner: dict[str, Any]) -> None:
    url = driver.current_url or ""
    partner["profile_url"] = url
    m = re.search(r"view-profile/([a-f0-9\-]+)", url, re.I)
    if m:
        partner["profile_uuid"] = m.group(1)


def scrape_partner_report(
    driver: webdriver.Chrome,
    partner: dict[str, Any],
    *,
    show_ui: bool = False,
    slot_pos: int = 1,
    total_slots: int = 1,
) -> dict[str, Any]:
    idx = partner.get("index")
    name = partner.get("name", "") or f"Campagne UPJUNOO {idx}"
    email = partner.get("email", "")
    scraped_at = datetime.now().isoformat(timespec="seconds")

    record: dict[str, Any] = {
        "index": idx,
        "name": name,
        "email": email,
        "profile_uuid": partner.get("profile_uuid", ""),
        "profile_url": partner.get("profile_url", ""),
        "vehicles_count": 0,
        "drivers_count": 0,
        "vehicles_by_status": {},
        "drivers_by_status": {},
        "vehicles": [],
        "drivers": [],
        "errors": [],
        "fleet_confirmed_empty": False,
        "scraped_at": scraped_at,
    }

    log(f"\n{'='*50}")
    log(f"[Campagne {slot_pos}/{total_slots}] n°{idx} | {name} | {email}")

    show_phase_ui(
        driver, partner=partner, slot_pos=slot_pos, total_slots=total_slots,
        step="1/4 RECHERCHE admin", detail=email, color=VIS_BLUE, show_ui=show_ui,
    )
    if not open_partner_profile_admin(driver, partner):
        record["errors"].append("partner_not_found")
        log(f"   Partenaire introuvable ({partner_search_queries(partner)})", "WARNING")
        return record

    if not is_on_partner_profile(driver):
        record["errors"].append("profile_page_not_open")
        log(f"   URL invalide → {driver.current_url}", "WARNING")
        if show_ui:
            show_phase_ui(
                driver, partner=partner, slot_pos=slot_pos, total_slots=total_slots,
                step="ERREUR profil", detail="manage-owners", color=VIS_RED, show_ui=True,
            )
        return record

    show_phase_ui(
        driver, partner=partner, slot_pos=slot_pos, total_slots=total_slots,
        step="2/4 PROFIL OK", detail="view-profile", color=VIS_GREEN, show_ui=show_ui,
    )
    capture_profile_meta(driver, partner)
    record["profile_uuid"] = partner.get("profile_uuid", "")
    record["profile_url"] = partner.get("profile_url", "")
    log(f"   Profil → {record['profile_url']}")

    show_phase_ui(
        driver, partner=partner, slot_pos=slot_pos, total_slots=total_slots,
        step="3/4 FLOTTE", detail="véhicules", color=VIS_ORANGE, show_ui=show_ui,
    )
    vehicles, fleet_errs, fleet_empty = scrape_fleet_vehicles(driver)
    record["vehicles"] = vehicles
    record["vehicles_count"] = len(vehicles)
    record["fleet_confirmed_empty"] = fleet_empty
    for e in fleet_errs:
        if e not in record["errors"]:
            record["errors"].append(e)

    show_phase_ui(
        driver, partner=partner, slot_pos=slot_pos, total_slots=total_slots,
        step="4/4 CONDUCTEURS", detail="chauffeurs", color=VIS_TEAL, show_ui=show_ui,
    )
    if not open_driver_tab(driver):
        record["errors"].append("driver_tab_failed")
        log("   Onglet conducteur inaccessible", "WARNING")
    else:
        record["drivers"] = parse_driver_detail_rows(driver)
        record["drivers_count"] = len(record["drivers"])

    retry_fleet_if_suspect(driver, partner, record)

    attach_status_breakdown(record)
    log(f"   {record['vehicles_count']} véhicule(s) — {format_status_breakdown(record['vehicles_by_status'])}")
    log(f"   {record['drivers_count']} chauffeur(s) — {format_status_breakdown(record['drivers_by_status'])}")

    if record["errors"] and show_ui:
        show_phase_ui(
            driver, partner=partner, slot_pos=slot_pos, total_slots=total_slots,
            step="FIN — erreurs", detail="; ".join(record["errors"]),
            color=VIS_RED, show_ui=True,
        )
    elif show_ui:
        show_phase_ui(
            driver, partner=partner, slot_pos=slot_pos, total_slots=total_slots,
            step="FIN — OK",
            detail=f"{record['vehicles_count']} véh. / {record['drivers_count']} chauf.",
            color=VIS_GREEN, show_ui=True,
        )
    return record


# ─── Export ───────────────────────────────────────────────────────────────────

def build_report(partner_records: list[dict[str, Any]]) -> dict[str, Any]:
    veh_by_status = merge_status_breakdowns(
        [r.get("vehicles_by_status") or {} for r in partner_records],
    )
    drv_by_status = merge_status_breakdowns(
        [r.get("drivers_by_status") or {} for r in partner_records],
    )
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": "admin",
        "partners_count": len(partner_records),
        "totals": {
            "vehicles": sum(r.get("vehicles_count", 0) for r in partner_records),
            "drivers": sum(r.get("drivers_count", 0) for r in partner_records),
            "vehicles_by_status": veh_by_status,
            "drivers_by_status": drv_by_status,
        },
        "partners": partner_records,
    }


def autosize_columns(ws, max_width: int = 48) -> None:
    for col_idx, column_cells in enumerate(ws.columns, 1):
        length = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, max(10, length + 2))


def export_excel(report: dict[str, Any], path: Path) -> None:
    wb = Workbook()
    ws_veh = wb.active
    ws_veh.title = "Vehicules"
    ws_drv = wb.create_sheet("Chauffeurs")
    ws_stat = wb.create_sheet("Repartition_statuts")
    veh_h = [
        "partenaire_index", "partenaire_nom", "partenaire_email", "profile_uuid",
        "profile_url", "plaque", "conducteur_assigne", "statut_flotte",
        "row_text", "erreur", "scraped_at",
    ]
    drv_h = [
        "partenaire_index", "partenaire_nom", "partenaire_email", "profile_uuid",
        "profile_url", "nom", "telephone", "emplacement", "type_transport",
        "statut_approuve", "notation", "type_vehicule", "row_text", "erreur", "scraped_at",
    ]
    stat_h = [
        "partenaire_index", "partenaire_nom", "partenaire_email",
        "categorie", "statut", "nombre", "scraped_at",
    ]
    ws_veh.append(veh_h)
    ws_drv.append(drv_h)
    ws_stat.append(stat_h)
    for p in report.get("partners") or []:
        err = "; ".join(p.get("errors") or [])
        base = [p.get("index"), p.get("name", ""), p.get("email", ""),
                p.get("profile_uuid", ""), p.get("profile_url", "")]
        scraped = p.get("scraped_at", "")
        stat_base = [p.get("index"), p.get("name", ""), p.get("email", "")]
        for v in p.get("vehicles") or []:
            ws_veh.append(base + [
                v.get("plate", ""), v.get("driver_name", ""), v.get("fleet_status", ""),
                v.get("row_text", ""), err, scraped,
            ])
        for d in p.get("drivers") or []:
            ws_drv.append(base + [
                d.get("name", ""), d.get("phone", ""), d.get("location", ""),
                d.get("transport_type", ""), d.get("approval_status", ""),
                d.get("rating", ""), d.get("vehicle_type", ""),
                d.get("row_text", ""), err, scraped,
            ])
        for statut, nombre in (p.get("vehicles_by_status") or {}).items():
            ws_stat.append(stat_base + ["Véhicule", statut, nombre, scraped])
        for statut, nombre in (p.get("drivers_by_status") or {}).items():
            ws_stat.append(stat_base + ["Chauffeur", statut, nombre, scraped])

    totals = report.get("totals") or {}
    ws_stat.append([])
    ws_stat.append(["", "", "TOTAL GLOBAL", "Véhicule", "", "", ""])
    for statut, nombre in (totals.get("vehicles_by_status") or {}).items():
        ws_stat.append(["", "", "TOTAL GLOBAL", "Véhicule", statut, nombre, ""])
    ws_stat.append([])
    ws_stat.append(["", "", "TOTAL GLOBAL", "Chauffeur", "", "", ""])
    for statut, nombre in (totals.get("drivers_by_status") or {}).items():
        ws_stat.append(["", "", "TOTAL GLOBAL", "Chauffeur", statut, nombre, ""])

    autosize_columns(ws_veh)
    autosize_columns(ws_drv)
    autosize_columns(ws_stat)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def export_json(report: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export rapport admin flotte + chauffeurs")
    parser.add_argument("--start", type=int, default=1)
    parser.add_argument("--end", type=int, default=20)
    parser.add_argument("--only", type=int)
    parser.add_argument(
        "--excel",
        default="",
        help="Optionnel : DOSSIER_PARTENAIRES.xlsx pour surcharger les logins. "
        "Sans --excel → campagne1@upjunoo.com … campagne20@upjunoo.com",
    )
    parser.add_argument("--email-template", default=DEFAULT_EMAIL_TEMPLATE)
    parser.add_argument("--password", default=DEFAULT_PARTNER_PASSWORD)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--headed", action="store_true")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG_JSON))
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR))
    args = parser.parse_args()

    if args.only is not None:
        args.start = args.end = args.only

    partners = resolve_partners(args)
    if not partners:
        log("Aucun partenaire à traiter.", "ERROR")
        sys.exit(2)

    out_dir = Path(args.output_dir)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = out_dir / f"rapport_partenaires_{ts}.json"
    xlsx_path = out_dir / f"rapport_partenaires_{ts}.xlsx"

    log("EXPORT RAPPORT PARTENAIRES (admin) — script autonome")
    log(f"   Slots: {len(partners)} — Campagne {args.start} → {args.end}")
    log(f"   Emails : {args.email_template.replace('{index}', 'N')} (défaut UpJunoo)")
    if args.excel:
        log(f"   ⚠ Surcharge Excel active : {args.excel}", "WARNING")
    log_campaign_plan(partners, start=args.start, end=args.end)
    log(f"   Sortie: {json_path}")
    if args.headed:
        log("   UI : bannière + panneau progression (droite)")

    log("   Démarrage Chrome…")
    driver = qa.setup_driver(headed=args.headed)
    log("   Chrome prêt.")
    records: list[dict[str, Any]] = []

    try:
        if not admin_login(driver):
            sys.exit(1)

        total_slots = len(partners)
        for slot_pos, partner in enumerate(partners, start=1):
            log(f"\n── Campagne {slot_pos}/{total_slots} (n°{partner['index']}) ──")
            try:
                records.append(
                    scrape_partner_report(
                        driver, partner,
                        show_ui=bool(args.headed),
                        slot_pos=slot_pos,
                        total_slots=total_slots,
                    ),
                )
            except Exception as e:
                log(f"Erreur partenaire {partner.get('index')}: {e}", "ERROR")
                if is_browser_dead(e):
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    driver = qa.setup_driver(headed=args.headed)
                    if not admin_login(driver):
                        sys.exit(1)
                records.append({
                    "index": partner.get("index"),
                    "name": partner.get("name", ""),
                    "email": partner.get("email", ""),
                    "profile_uuid": partner.get("profile_uuid", ""),
                    "profile_url": "",
                    "vehicles_count": 0,
                    "drivers_count": 0,
                    "vehicles_by_status": {},
                    "drivers_by_status": {},
                    "vehicles": [],
                    "drivers": [],
                    "errors": [f"exception:{e}"],
                    "scraped_at": datetime.now().isoformat(timespec="seconds"),
                })

        suspects_before = list_suspect_records(records)
        if suspects_before:
            log(
                f"\nValidation post-export : {len(suspects_before)} campagne(s) suspecte(s)",
                "WARNING",
            )
            validate_and_rescrape_suspects(driver, partners, records)

        report = build_report(records)
        export_json(report, json_path)
        export_excel(report, xlsx_path)

        state = load_state()
        state.setdefault("partners", {})
        for rec in records:
            key = str(rec.get("index"))
            if key and key != "None":
                ent = state["partners"].setdefault(key, {"index": rec.get("index")})
                if rec.get("profile_uuid"):
                    ent["profile_uuid"] = rec["profile_uuid"]
                ent["email"] = rec.get("email", "")
        save_state(state)

        log("\n" + "=" * 50)
        log("TERMINÉ")
        log(
            f"   {report['partners_count']} partenaires | "
            f"{report['totals']['vehicles']} véhicules | "
            f"{report['totals']['drivers']} chauffeurs",
        )
        log(f"   Véhicules (global) : {format_status_breakdown(report['totals'].get('vehicles_by_status') or {})}")
        log(f"   Chauffeurs (global) : {format_status_breakdown(report['totals'].get('drivers_by_status') or {})}")
        final_suspects = list_suspect_records(records)
        if final_suspects:
            log(
                f"   ⚠ ATTENTION : {len(final_suspects)} campagne(s) avec 0 véh. et des chauffeurs : "
                + ", ".join(f"P{int(r['index']):02d}" for r in final_suspects),
                "WARNING",
            )
        log(f"   JSON: {json_path}")
        log(f"   Excel: {xlsx_path}")

    except KeyboardInterrupt:
        log("Interruption Ctrl+C", "WARNING")
        if records:
            report = build_report(records)
            export_json(report, json_path)
            export_excel(report, xlsx_path)
            log(f"Export partiel → {json_path}", "WARNING")
    finally:
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
