#!/usr/bin/env python3
"""Partenaires du JSON absents du RECAP fleet-drivers."""

from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

JSON_PARTNERS = [
    {"nom_complet": "ADONI FIRMIN", "partenaire": 79},
    {"nom_complet": "AFFI Nathanael", "partenaire": 73},
    {"nom_complet": "AKAFFOU Abondo Charles", "partenaire": 45},
    {"nom_complet": "AKE YAPO CASIMIR", "partenaire": 33},
    {"nom_complet": "C'KOOL SERVICES", "partenaire": 16},
    {"nom_complet": "BAÏ Deassa Dorgeles", "partenaire": 85},
    {"nom_complet": "SICOPP-CI", "partenaire": 88},
    {"nom_complet": "BLE Gongo Fulgence", "partenaire": 110},
    {"nom_complet": "BOHOUO Ange Patrick", "partenaire": 47},
    {"nom_complet": "ZICARL SARL", "partenaire": 63},
    {"nom_complet": "COULIBALY Yasmine", "partenaire": 78},
    {"nom_complet": "DANIOKO Hamed Ibrahim", "partenaire": 105},
    {"nom_complet": "DEGBOU Yao Marie Paule", "partenaire": 46},
    {"nom_complet": "DEKA HOLDING", "partenaire": 12},
    {"nom_complet": "DIABAGATE Digueta", "partenaire": 104},
    {"nom_complet": "VGD", "partenaire": 19},
    {"nom_complet": "DIOMANDE Adama", "partenaire": 64},
    {"nom_complet": "ETOILE DE SONGON", "partenaire": 18},
    {"nom_complet": "DRO Kadidjatou Flora", "partenaire": 97},
    {"nom_complet": "UNIVERS SERVICES IMMOBILIER", "partenaire": 48},
    {"nom_complet": "GOPOU Sonia", "partenaire": 71},
    {"nom_complet": "GUE GBATO Jean Marc", "partenaire": 83},
    {"nom_complet": "KASSI Venance Desire", "partenaire": 82},
    {"nom_complet": "KOFFI Axel", "partenaire": 26},
    {"nom_complet": "PROSPERITY MARKETING", "partenaire": 99},
    {"nom_complet": "KONAN Yao Afon Anaice", "partenaire": 8},
    {"nom_complet": "IVOIRE PLUS GROUP (FLOTTE 2)", "partenaire": 30},
    {"nom_complet": "KOUADIO N'DAYA BIENVENUE VICTOIRE", "partenaire": 9},
    {"nom_complet": "KOUAKOU Yao Jean Marc", "partenaire": 6},
    {"nom_complet": "KOUASSI AMOIN ARIELLE EPSE AMAND", "partenaire": 53},
    {"nom_complet": "KOUDOU Marc", "partenaire": 49},
    {"nom_complet": "KRAMO Fortuné", "partenaire": 65},
    {"nom_complet": "LIDET Delmas Houphouet", "partenaire": 59},
    {"nom_complet": "MHG DELIVERY (FLOTTE 2)", "partenaire": 75},
    {"nom_complet": "N'DRI Yao Anicet", "partenaire": 66},
    {"nom_complet": "NAL ANGEL'S GROUPE INTERNATIONAL", "partenaire": 113},
    {"nom_complet": "N'GORAN Jacques", "partenaire": 17},
    {"nom_complet": "OPEA Jean Yves", "partenaire": 58},
    {"nom_complet": "ORIANE BUSINESS CENTER", "partenaire": 52},
    {"nom_complet": "TRAMEX IVOIRE TRANSPORT", "partenaire": 119},
    {"nom_complet": "SERI Legre Hugues", "partenaire": 60},
    {"nom_complet": "SIAMA Abdramane", "partenaire": 93},
    {"nom_complet": "KOUADIO KOFFI HERMAN-HUGUES", "partenaire": 51},
    {"nom_complet": "TANOH Koffi Augustin", "partenaire": 56},
    {"nom_complet": "TIEHI Henri Joel", "partenaire": 57},
    {"nom_complet": "LT BENEDICTION SERVICES", "partenaire": 69},
    {"nom_complet": "2HU SERVICES", "partenaire": 43},
    {"nom_complet": "KING TELECOM ET SERVICES", "partenaire": 31},
    {"nom_complet": "UHUD MULTISERVICES (FLOTTE 2)", "partenaire": 44},
    {"nom_complet": "ETS TOURE (FLOTTE 2)", "partenaire": 108},
    {"nom_complet": "TRE LOU LINDA ANNE", "partenaire": 118},
    {"nom_complet": "VOUE Elisée Parfait", "partenaire": 55},
    {"nom_complet": "AS COM", "partenaire": 28},
    {"nom_complet": "YOUAN BI Dieudonné", "partenaire": 10},
    {"nom_complet": "YOUZAN FRANCK STEPHANE", "partenaire": 90},
    {"nom_complet": "ETS BAYANO SERVICES", "partenaire": 20},
    {"nom_complet": "ZOU Grace Bénédicte", "partenaire": 29},
]

BASE_DIR = Path("output/fleet-drivers-par-partenaire")
RECAP_PATH = BASE_DIR / "_RECAP_partenaires.xlsx"
NON_TRAITE_DIR = BASE_DIR / "Partenaire non Traite"


def norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^a-z0-9 '\-]+", "", s)
    return s


def strip_chauf_suffix(s: str) -> str:
    return re.sub(r"\(\d+chauf\)\s*$", "", s, flags=re.I).strip()


def load_recap_names() -> tuple[set[str], list[str]]:
    wb = load_workbook(RECAP_PATH, read_only=True, data_only=True)
    ws = wb.active
    normalized: set[str] = set()
    display: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        partenaire, _nb, fichier, nom_orig = (list(row) + [None] * 4)[:4]
        for raw in (partenaire, fichier, nom_orig):
            if not raw:
                continue
            text = strip_chauf_suffix(str(raw))
            display.append(text)
            normalized.add(norm(text))
    wb.close()
    return normalized, display


def match_recap(json_name: str, recap_norm: set[str], recap_display: list[str]) -> str | None:
    n = norm(json_name)
    if n in recap_norm:
        return "exact"
    # Variantes connues (troncature fichier / libellés différents)
    aliases = {
        norm("UNIVERS SERVICES IMMOBILIER"): [
            norm("UNIVERS SERVICES IMMOBILI"),
        ],
        norm("DIABAGATE Digueta"): [norm("DIABAGATE DIGUETTA")],
        norm("KOUADIO N'DAYA BIENVENUE VICTOIRE"): [norm("KOUADIO N'DAYA BIENVENUE")],
        norm("KOUASSI AMOIN ARIELLE EPSE AMAND"): [norm("KOUASSI AMOIN ARIELLE EPS")],
        norm("NAL ANGEL'S GROUPE INTERNATIONAL"): [norm("NAL ANGEL'S GROUPE INTERN")],
        norm("LT BENEDICTION SERVICES"): [norm("LT BENEDICTION SERVICES")],
        norm("ADONI FIRMIN"): [],  # vraiment absent sauf si autre nom
    }
    for alt in aliases.get(n, []):
        if alt in recap_norm:
            return "alias"

    for r in recap_display:
        rn = norm(r)
        if n == rn or (len(n) >= 10 and (n in rn or rn in n)):
            return f"proche:{r}"
    return None


def main() -> None:
    recap_norm, recap_display = load_recap_names()
    missing: list[dict] = []
    found: list[tuple[dict, str]] = []

    for item in JSON_PARTNERS:
        m = match_recap(item["nom_complet"], recap_norm, recap_display)
        if m:
            found.append((item, m))
        else:
            missing.append(item)

    print(f"JSON: {len(JSON_PARTNERS)} partenaires")
    print(f"Présents dans RECAP: {len(found)}")
    print(f"ABSENTS du RECAP: {len(missing)}\n")

    if missing:
        print("--- Dans le JSON mais PAS dans le RECAP (export fleet-drivers) ---")
        for m in sorted(missing, key=lambda x: x["partenaire"]):
            print(f"  #{m['partenaire']:3d} | {m['nom_complet']}")
    else:
        print("Tous les partenaires JSON sont dans le RECAP.")

    # Vrais absents stricts (sans match proche)
    strict = [m for m, _ in [(x, match_recap(x["nom_complet"], recap_norm, recap_display)) for x in JSON_PARTNERS] if _ is None]
    # fix - rerun strict only for missing from first pass


def load_recap_rows() -> list[dict]:
    wb = load_workbook(RECAP_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        partenaire, nb, fichier, nom_orig = (list(row) + [None] * 4)[:4]
        name = strip_chauf_suffix(str(nom_orig or partenaire or ""))
        rows.append(
            {
                "nom": name,
                "nb": int(nb or 0),
                "fichier": str(fichier or "").strip(),
                "partenaire_label": str(partenaire or "").strip(),
                "nom_original": str(nom_orig or "").strip(),
            },
        )
    wb.close()
    return rows


def is_in_json(recap_name: str) -> bool:
    for item in JSON_PARTNERS:
        if match_recap(item["nom_complet"], {norm(recap_name)}, [recap_name]):
            return True
    return False


def split_recap_rows() -> tuple[list[dict], list[dict]]:
    traites: list[dict] = []
    non_traites: list[dict] = []
    for r in load_recap_rows():
        if is_in_json(r["nom"]):
            traites.append(r)
        else:
            non_traites.append(r)
    return traites, non_traites


def write_recap(path: Path, rows: list[dict], title: str = "Récap") -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(["Partenaire", "Nb conducteurs", "Fichier", "Nom original"])
    for r in sorted(rows, key=lambda x: (-x["nb"], x["nom"].lower())):
        label = r.get("partenaire_label") or f"{r['nom']}({r['nb']}chauf)"
        ws.append([label, r["nb"], r["fichier"], r["nom"]])
    for col_idx, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(52, max(12, length + 2))
    wb.save(path)


def move_non_traites() -> None:
    traites, non_traites = split_recap_rows()
    NON_TRAITE_DIR.mkdir(parents=True, exist_ok=True)

    moved = 0
    missing: list[str] = []
    for r in non_traites:
        fname = r["fichier"]
        if not fname:
            missing.append(r["nom"])
            continue
        src = BASE_DIR / fname
        dst = NON_TRAITE_DIR / fname
        if not src.is_file():
            if dst.is_file():
                moved += 1
                continue
            missing.append(fname)
            continue
        if dst.exists():
            raise FileExistsError(f"Déjà présent: {dst}")
        src.rename(dst)
        moved += 1

    write_recap(RECAP_PATH, traites)
    write_recap(NON_TRAITE_DIR / "_RECAP_partenaires_non_traites.xlsx", non_traites)

    print(f"Dossier traités: {BASE_DIR}")
    print(f"Dossier non traités: {NON_TRAITE_DIR}")
    print(f"Déplacés: {moved} / {len(non_traites)}")
    if missing:
        print(f"Fichiers introuvables: {len(missing)}")
        for m in missing[:10]:
            print(f"  - {m}")
    print(f"RECAP principal: {len(traites)} partenaires")
    print(f"RECAP non traités: {NON_TRAITE_DIR / '_RECAP_partenaires_non_traites.xlsx'}")


def recap_not_in_json() -> list[dict]:
    return [r for r in load_recap_rows() if not is_in_json(r["nom"])]


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == "--recap-only":
        rows = recap_not_in_json()
        print(f"RECAP hors JSON: {len(rows)}\n")
        for r in sorted(rows, key=lambda x: (-x["nb"], x["nom"].lower())):
            print(f"  {r['nb']:4d} chauf | {r['nom']}")
    elif len(sys.argv) > 1 and sys.argv[1] == "--move-non-traites":
        move_non_traites()
    else:
        main()
