#!/usr/bin/env python3
"""Découpe un export fleet-drivers CSV en un fichier XLSX par partenaire."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def sanitize_filename(name: str, max_len: int = 180) -> str:
    name = (name or "").strip() or "sans_partenaire"
    name = re.sub(r'[\\/:*?"<>|]', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:max_len] or "sans_partenaire"


def partner_filename(partner: str, count: int) -> str:
    """Ex: Campagne UPJUNOO 1(8chauf).xlsx"""
    label = f"{partner}({int(count)}chauf)"
    return sanitize_filename(label) + ".xlsx"


def autosize_columns(ws, max_width: int = 48) -> None:
    for col_idx, column_cells in enumerate(ws.columns, 1):
        length = 0
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            length = max(length, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, max(10, length + 2))


def split_csv_to_xlsx(csv_path: Path, out_dir: Path) -> dict[str, int]:
    out_dir.mkdir(parents=True, exist_ok=True)

    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    fieldnames: list[str] = []

    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        for row in reader:
            partner = (row.get("Partenaire") or "").strip() or "(sans partenaire)"
            groups[partner].append(row)

    used_names: dict[str, int] = {}
    counts: dict[str, int] = {}
    file_by_partner: dict[str, str] = {}

    for partner, rows in sorted(groups.items(), key=lambda x: x[0].lower()):
        n = len(rows)
        fname = partner_filename(partner, n)
        base_key = fname.lower()
        used_names[base_key] = used_names.get(base_key, 0) + 1
        if used_names[base_key] > 1:
            stem = Path(fname).stem
            fname = sanitize_filename(f"{stem}_{used_names[base_key]}") + ".xlsx"
        xlsx_path = out_dir / fname
        file_by_partner[partner] = fname

        wb = Workbook()
        ws = wb.active
        ws.title = "Conducteurs"
        ws.append(fieldnames)
        for row in rows:
            ws.append([row.get(col, "") for col in fieldnames])
        autosize_columns(ws)
        wb.save(xlsx_path)

        counts[partner] = n

    # Récap
    recap_path = out_dir / "_RECAP_partenaires.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Récap"
    ws.append(["Partenaire", "Nb conducteurs", "Fichier", "Nom original"])
    for partner, n in sorted(counts.items(), key=lambda x: (-x[1], x[0].lower())):
        label = f"{partner}({n}chauf)"
        ws.append([label, n, file_by_partner[partner], partner])
    autosize_columns(ws)
    wb.save(recap_path)

    return counts


def main() -> None:
    parser = argparse.ArgumentParser(description="Découpe fleet-drivers CSV par partenaire")
    parser.add_argument(
        "csv",
        nargs="?",
        default="fleet-drivers-20260521-091257.csv",
        help="Chemin du CSV source",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="output/fleet-drivers-par-partenaire",
        help="Dossier de sortie",
    )
    args = parser.parse_args()

    csv_path = Path(args.csv).resolve()
    out_dir = Path(args.output).resolve()
    if not csv_path.is_file():
        raise SystemExit(f"CSV introuvable: {csv_path}")

    counts = split_csv_to_xlsx(csv_path, out_dir)
    total = sum(counts.values())
    print(f"Source: {csv_path}")
    print(f"Sortie: {out_dir}")
    print(f"Partenaires: {len(counts)} | Conducteurs: {total}")
    print(f"Récap: {out_dir / '_RECAP_partenaires.xlsx'}")


if __name__ == "__main__":
    main()
