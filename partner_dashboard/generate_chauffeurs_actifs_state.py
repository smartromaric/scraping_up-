#!/usr/bin/env python3
"""Génère chauffeurs_actifs_state.xlsx depuis state.json + export Godseye."""

from __future__ import annotations

import argparse
import html
import json
import re
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "Campagne",
    "Chauffeur",
    "Telephone",
    "Owner Status",
    "Fleet Status",
    "Reason",
    "Transfer 2000 Done",
    "Transfer 2000 At",
    "En ligne",
]

FILL_ONLINE = PatternFill("solid", fgColor="C8E6C9")
FILL_OFFLINE = PatternFill("solid", fgColor="FFCDD2")
FILL_HEADER = PatternFill("solid", fgColor="1565C0")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="CCCCCC")


def default_output_dir() -> Path:
    """Dossier partner_automation (state.json, Godseye, Excel générés)."""
    here = Path(__file__).resolve().parent
    data_dir = here.parent / "output" / "partner_automation"
    if data_dir.is_dir() or (data_dir / "state.json").is_file():
        return data_dir
    return here


def norm_text(s: str) -> str:
    s = (s or "").strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_phone_digits(phone: str) -> str:
    digits = re.sub(r"\D", "", phone or "")
    if len(digits) >= 10:
        return digits[-10:]
    return digits


def phone_e164(phone: str) -> str:
    d = norm_phone_digits(phone)
    return f"+225{d}" if d else ""


def camp_num(partner_name: str) -> int | None:
    m = re.search(r"(\d+)\s*$", partner_name or "")
    return int(m.group(1)) if m else None


def parse_godseye(path: Path) -> tuple[set[str], set[str]]:
    """Retourne (phones 10 chiffres, noms normalisés)."""
    text = path.read_text(encoding="utf-8", errors="ignore")
    phones: set[str] = set()
    names: set[str] = set()
    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", text, flags=re.I | re.S)
    for row in rows[1:]:
        cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row, flags=re.I | re.S)
        clean = []
        for c in cells:
            c = re.sub(r"<[^>]+>", " ", c)
            c = html.unescape(c)
            c = re.sub(r"\s+", " ", c).strip()
            clean.append(c)
        if len(clean) < 3:
            continue
        name = clean[1] if len(clean) > 1 else ""
        mobile = clean[2] if len(clean) > 2 else ""
        d = norm_phone_digits(mobile)
        if d:
            phones.add(d)
        n = norm_text(name)
        if n:
            names.add(n)
    return phones, names


def is_online(name: str, phone: str, g_phones: set[str], g_names: set[str]) -> bool:
    d = norm_phone_digits(phone)
    if d and d in g_phones:
        return True
    nn = norm_text(name)
    if not nn:
        return False
    if nn in g_names:
        return True
    for gn in g_names:
        if nn in gn or gn in nn:
            return True
        ta, tb = set(nn.split()), set(gn.split())
        if len(ta & tb) >= 2:
            return True
    return False


def load_active_rows(state_path: Path, g_phones: set[str], g_names: set[str]) -> list[dict]:
    data = json.loads(state_path.read_text(encoding="utf-8"))
    rows: list[dict] = []
    partners = data.get("partners", {})
    iterable = partners.values() if isinstance(partners, dict) else partners
    for p in iterable:
        pname = p.get("name", "")
        cn = camp_num(pname)
        for drv in p.get("drivers", []):
            admin = drv.get("admin") or {}
            if admin.get("reason") != "assigned_approved":
                continue
            phone = phone_e164(drv.get("phone", ""))
            done = drv.get("transfer_2000_done")
            rows.append(
                {
                    "camp_num": cn or 0,
                    "campagne": pname,
                    "chauffeur": drv.get("name", ""),
                    "telephone": phone,
                    "owner": drv.get("owner_status", ""),
                    "fleet": admin.get("fleet_status", ""),
                    "reason": admin.get("reason", ""),
                    "transfer_done": True if done else False if done is False else None,
                    "transfer_at": drv.get("transfer_2000_at") or "",
                    "en_ligne": is_online(drv.get("name", ""), phone, g_phones, g_names),
                }
            )
    rows.sort(key=lambda r: (r["camp_num"], norm_text(r["chauffeur"])))
    return rows


def style_sheet(ws, data_rows: int) -> None:
    ws.freeze_panes = "A2"
    for col, w in enumerate([22, 36, 18, 14, 14, 22, 18, 22, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for r in range(1, data_rows + 2):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if r > 1:
                cell.alignment = Alignment(vertical="center", wrap_text=(c == 2))
                fill = FILL_ONLINE if ws.cell(r, 9).value == "Oui" else FILL_OFFLINE
                cell.fill = fill
    if data_rows:
        ws.auto_filter.ref = f"A1:I{data_rows + 1}"


def output_filename_with_timestamp(when: datetime | None = None) -> str:
    """Nom horodaté : chauffeurs_actifs_state_YYYYMMDD_HHMMSS.xlsx"""
    ts = (when or datetime.now()).strftime("%Y%m%d_%H%M%S")
    return f"chauffeurs_actifs_state_{ts}.xlsx"


def format_generated_at(when: datetime | None = None) -> str:
    return (when or datetime.now()).strftime("%d/%m/%Y %H:%M:%S")


def write_footer(
    ws,
    start_row: int,
    online: int,
    offline: int,
    *,
    generated_at: str,
    godseye_file: str,
    godseye_at: str = "",
) -> None:
    r = start_row + 1
    label_font = Font(bold=True, size=10, color="424242")
    ws.cell(r, 1, "Généré le").font = label_font
    ws.cell(r, 2, generated_at)
    r += 1
    ws.cell(r, 1, "Export Godseye").font = label_font
    ws.cell(r, 2, godseye_file)
    if godseye_at:
        ws.cell(r, 3, godseye_at)
    r += 1
    ws.cell(r, 1, "STATUT GLOBAL").font = Font(bold=True, size=12)
    r += 1
    ws.cell(r, 1, "En ligne (vert léger)")
    ws.cell(r, 2, online)
    r += 1
    ws.cell(r, 1, "Pas en ligne (rouge)")
    ws.cell(r, 2, offline)


def build_workbook(
    rows: list[dict],
    *,
    generated_at: str,
    godseye_file: str,
    godseye_at: str = "",
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    sheets = {"1-10": [], "11-20": []}
    for row in rows:
        n = row["camp_num"]
        if 1 <= n <= 10:
            sheets["1-10"].append(row)
        elif 11 <= n <= 20:
            sheets["11-20"].append(row)
    for title, subset in sheets.items():
        ws = wb.create_sheet(title)
        ws.append(HEADERS)
        online = offline = 0
        for row in subset:
            en = "Oui" if row["en_ligne"] else "Non"
            if row["en_ligne"]:
                online += 1
            else:
                offline += 1
            ws.append(
                [
                    row["campagne"],
                    row["chauffeur"],
                    row["telephone"],
                    row["owner"],
                    row["fleet"],
                    row["reason"],
                    row["transfer_done"],
                    row["transfer_at"] or None,
                    en,
                ]
            )
        style_sheet(ws, len(subset))
        write_footer(
            ws,
            len(subset) + 2,
            online,
            offline,
            generated_at=generated_at,
            godseye_file=godseye_file,
            godseye_at=godseye_at,
        )
    return wb


def find_latest_godseye(directory: Path | None = None) -> Path | None:
    """Dernier export drivers-godseye-*.xls (HTML) dans le dossier partner_automation."""
    base = directory or default_output_dir()
    files = [p for p in base.glob("drivers-godseye-*.xls") if p.is_file()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def find_latest_chauffeurs_xlsx(directory: Path | None = None) -> Path | None:
    """Dernier fichier Excel horodaté chauffeurs_actifs_state_YYYYMMDD_HHMMSS.xlsx.

    La génération écrit aussi une copie "chauffeurs_actifs_state.xlsx" sans suffixe.
    Pour le suivi (et le téléchargement côté dashboard), on préfère retourner le fichier horodaté.
    """
    base = directory or default_output_dir()
    files = [
        p
        for p in base.glob("chauffeurs_actifs_state*.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    ]
    if not files:
        return None
    ts_re = re.compile(r"^chauffeurs_actifs_state_\d{8}_\d{6}\.xlsx$")
    timestamped = [p for p in files if ts_re.match(p.name)]
    if timestamped:
        return max(timestamped, key=lambda p: p.stat().st_mtime)
    # Fallback : uniquement la copie "latest" existe.
    return max(files, key=lambda p: p.stat().st_mtime)


def generate_report(
    *,
    state_path: Path,
    godseye_path: Path | None = None,
    output_path: Path | None = None,
    timestamped: bool = True,
) -> dict:
    """Génère l'Excel et retourne des statistiques pour l'API dashboard."""
    base = default_output_dir()
    state_path = Path(state_path)
    now = datetime.now()
    if output_path is None:
        if timestamped:
            output_path = base / output_filename_with_timestamp(now)
        else:
            output_path = base / "chauffeurs_actifs_state.xlsx"
    else:
        output_path = Path(output_path)
    godseye = Path(godseye_path) if godseye_path else find_latest_godseye(base)
    if not state_path.is_file():
        raise FileNotFoundError(f"state.json introuvable: {state_path}")
    if not godseye or not godseye.is_file():
        raise FileNotFoundError(
            "Aucun export Godseye (drivers-godseye-*.xls). "
            "Lancez d'abord le téléchargement depuis le dashboard."
        )

    g_st = godseye.stat()
    godseye_at = datetime.fromtimestamp(g_st.st_mtime).strftime("%d/%m/%Y %H:%M:%S")
    generated_label = format_generated_at(now)

    g_phones, g_names = parse_godseye(godseye)
    rows = load_active_rows(state_path, g_phones, g_names)
    wb = build_workbook(
        rows,
        generated_at=generated_label,
        godseye_file=godseye.name,
        godseye_at=godseye_at,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    latest_copy = base / "chauffeurs_actifs_state.xlsx"
    if output_path.resolve() != latest_copy.resolve():
        shutil.copy2(output_path, latest_copy)

    online = sum(1 for r in rows if r["en_ligne"])
    recharged = sum(1 for r in rows if r["transfer_done"] is True)
    to_recharge = sum(1 for r in rows if r["transfer_done"] is not True)
    return {
        "output": str(output_path.resolve()),
        "output_name": output_path.name,
        "latest_copy": str(latest_copy.resolve()),
        "generated_at": generated_label,
        "generated_at_iso": now.isoformat(timespec="seconds"),
        "godseye": str(godseye.resolve()),
        "godseye_name": godseye.name,
        "godseye_at": godseye_at,
        "state": str(state_path.resolve()),
        "actifs": len(rows),
        "en_ligne": online,
        "hors_ligne": len(rows) - online,
        "recharges": recharged,
        "a_recharger": to_recharge,
        "godseye_phones": len(g_phones),
        "godseye_names": len(g_names),
    }


def main() -> None:
    base = default_output_dir()
    latest = find_latest_godseye(base)
    parser = argparse.ArgumentParser(description="Génère chauffeurs_actifs_state.xlsx")
    parser.add_argument("--state", type=Path, default=base / "state.json")
    parser.add_argument("--godseye", type=Path, default=latest)
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Chemin de sortie (défaut : nom horodaté chauffeurs_actifs_state_YYYYMMDD_HHMMSS.xlsx)",
    )
    parser.add_argument(
        "--no-timestamp",
        action="store_true",
        help="Écrire chauffeurs_actifs_state.xlsx sans suffixe horodaté",
    )
    args = parser.parse_args()

    stats = generate_report(
        state_path=args.state,
        godseye_path=args.godseye,
        output_path=args.output,
        timestamped=not args.no_timestamp and args.output is None,
    )
    print(f"Fichier: {stats['output_name']} ({stats['generated_at']})")
    print(
        f"Actifs: {stats['actifs']} | En ligne: {stats['en_ligne']} | "
        f"Rechargés: {stats['recharges']}"
    )
    print(f"Godseye: {stats['godseye_phones']} téléphones, {stats['godseye_names']} noms")
    try:
        from partner_dashboard.chauffeurs_actifs_service import create_archive_for_xlsx

        zip_path = create_archive_for_xlsx(Path(stats["output"]), stats)
        print(f"Archive: {zip_path.name}")
    except Exception as e:
        print(f"Archive (dashboard) ignorée: {e}")


if __name__ == "__main__":
    main()
