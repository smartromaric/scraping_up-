#!/usr/bin/env python3
"""
Mode assisté UpJunoo Driver - transferts en série.

Ce script:
1) Se connecte au compte partenaire.
2) Ouvre Comptes -> Portefeuille -> modal Transférer.
3) Remplit numéro + montant pour chaque ligne d'un fichier (CSV/XLSX).
4) Te laisse choisir le destinataire + cliquer Transférer + valider popup.
5) Passe à la ligne suivante quand tu appuies Entrée dans le terminal.
"""

import csv
import glob
import os
import shutil
import subprocess
import sys
import time
from typing import Iterable, List, Dict, Tuple

import openpyxl
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


APPIUM_URL = os.getenv("APPIUM_URL", "http://127.0.0.1:4723")
APP_PACKAGE = os.getenv("UPJUNOO_PACKAGE", "com.upjunoo.driver")
APP_ACTIVITY = os.getenv("UPJUNOO_ACTIVITY", ".MainActivity")

PARTNER_EMAIL = os.getenv("PARTNER_EMAIL", "blessingcania@gmail.com")
PARTNER_PASSWORD = os.getenv("PARTNER_PASSWORD", "123456789@")

# Fichier source des transferts (2 colonnes minimum: numero, montant)
TRANSFER_FILE = os.getenv("TRANSFER_FILE", "transfer_list.csv")

Locator = Tuple[str, str]


def _find_adb_executable() -> str | None:
    p = shutil.which("adb")
    if p:
        return p
    for base in (os.environ.get("ANDROID_HOME"), os.environ.get("ANDROID_SDK_ROOT")):
        if base:
            cand = os.path.join(base, "platform-tools", "adb.exe")
            if os.path.isfile(cand):
                return cand
    winget_pkgs = os.path.join(os.environ.get("LOCALAPPDATA", ""), "Microsoft", "WinGet", "Packages")
    if os.path.isdir(winget_pkgs):
        matches = glob.glob(os.path.join(winget_pkgs, "Google.PlatformTools*", "platform-tools", "adb.exe"))
        if matches:
            return matches[0]
    return None


def list_adb_device_serials() -> List[str]:
    adb = _find_adb_executable()
    if not adb:
        return []
    creationflags = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
    try:
        r = subprocess.run(
            [adb, "devices"],
            capture_output=True,
            text=True,
            timeout=20,
            creationflags=creationflags,
        )
    except (OSError, subprocess.TimeoutExpired):
        return []
    serials: List[str] = []
    for line in (r.stdout or "").splitlines():
        line = line.strip()
        if not line or line.startswith("List of devices"):
            continue
        parts = line.split()
        if len(parts) >= 2 and parts[-1] == "device":
            serials.append(parts[0])
    return serials


def resolve_android_udid(requested: str) -> str:
    """
    Si ANDROID_UDID pointe vers un ancien téléphone (variable système Windows),
    on bascule sur le seul appareil réellement branché.
    """
    serials = list_adb_device_serials()
    if not serials:
        if _find_adb_executable() is None:
            print("⚠️ adb introuvable (PATH). Impossible de vérifier le téléphone.")
        if requested:
            print(f"→ Utilisation de ANDROID_UDID={requested!r} sans vérification adb.")
            return requested
        raise SystemExit(
            "Aucun appareil dans `adb devices`. Branche le téléphone, active le débogage USB, autorise le PC."
        )

    if requested and requested in serials:
        return requested

    if requested and requested not in serials:
        if len(serials) == 1:
            print(
                f"⚠️ ANDROID_UDID={requested!r} ne correspond pas au téléphone branché "
                f"(souvent une variable d'environnement Windows obsolète)."
            )
            print(f"→ Utilisation de l'appareil connecté : {serials[0]}")
            return serials[0]
        raise SystemExit(
            f"ANDROID_UDID={requested!r} absent. Appareils vus par adb : {serials}. "
            "Mets à jour ou supprime ANDROID_UDID dans les variables d'environnement Windows."
        )

    if len(serials) == 1:
        print(f"📱 Appareil détecté : {serials[0]}")
        return serials[0]

    raise SystemExit(
        f"Plusieurs appareils USB : {serials}. Définis ANDROID_UDID sur celui à utiliser "
        "(voir la colonne de gauche dans `adb devices`)."
    )


def wait_click(driver, wait: WebDriverWait, locators: Iterable[Locator], label: str):
    last_err = None
    for by, value in locators:
        try:
            el = wait.until(EC.element_to_be_clickable((by, value)))
            el.click()
            print(f"✅ Click: {label} [{by}={value}]")
            return el
        except Exception as e:
            last_err = e
    raise TimeoutException(f"Element non cliquable: {label} | err={last_err}")


def wait_visible(driver, wait: WebDriverWait, locators: Iterable[Locator], label: str):
    last_err = None
    for by, value in locators:
        try:
            el = wait.until(EC.visibility_of_element_located((by, value)))
            print(f"✅ Visible: {label} [{by}={value}]")
            return el
        except Exception as e:
            last_err = e
    raise TimeoutException(f"Element non visible: {label} | err={last_err}")


def load_rows(path: str) -> List[Dict[str, str]]:
    p = os.path.abspath(path)
    if not os.path.exists(p):
        raise FileNotFoundError(f"Fichier introuvable: {p}")

    rows: List[Dict[str, str]] = []
    lower = p.lower()

    phone_keys = {"numero", "numéro", "telephone", "téléphone", "phone", "mobile"}
    amount_keys = {"montant", "amount", "somme", "prix", "total"}

    def _norm_col(s: str) -> str:
        return (s or "").strip().lower()

    if lower.endswith(".csv"):
        with open(p, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError("CSV sans en-têtes de colonnes.")
            fields = [_norm_col(c) for c in reader.fieldnames]
            phone_col = next((reader.fieldnames[i] for i, c in enumerate(fields) if c in phone_keys), None)
            amount_col = next((reader.fieldnames[i] for i, c in enumerate(fields) if c in amount_keys), None)
            status_col = next((reader.fieldnames[i] for i, c in enumerate(fields) if c in {"statut", "status"}), None)
            if not phone_col or not amount_col:
                raise ValueError("Colonnes requises introuvables. Attendu: numero + montant")
            for idx, r in enumerate(reader, start=2):
                phone = str(r.get(phone_col, "")).strip()
                amount = str(r.get(amount_col, "")).strip()
                status = str(r.get(status_col, "")).strip() if status_col else ""
                if phone and amount:
                    rows.append({"row_index": str(idx), "numero": phone, "montant": amount, "statut": status})
    elif lower.endswith(".xlsx"):
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        fields = [_norm_col(c) for c in header]
        phone_idx = next((i for i, c in enumerate(fields) if c in phone_keys), None)
        amount_idx = next((i for i, c in enumerate(fields) if c in amount_keys), None)
        if phone_idx is None or amount_idx is None:
            raise ValueError("Colonnes requises introuvables dans XLSX. Attendu: numero + montant")
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            phone = str(vals[phone_idx]).strip() if phone_idx < len(vals) and vals[phone_idx] is not None else ""
            amount = str(vals[amount_idx]).strip() if amount_idx < len(vals) and vals[amount_idx] is not None else ""
            if phone and amount:
                rows.append({"row_index": str(len(rows) + 2), "numero": phone, "montant": amount, "statut": ""})
    else:
        raise ValueError("Format non supporté. Utilise .csv ou .xlsx")

    if not rows:
        raise ValueError("Aucune ligne valide à traiter.")
    return rows


def ensure_csv_status_column(path: str):
    if not path.lower().endswith(".csv"):
        return
    p = os.path.abspath(path)
    with open(p, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        data = list(reader)
    if "statut" in fieldnames:
        return
    fieldnames = list(fieldnames) + ["statut"]
    for r in data:
        r["statut"] = r.get("statut", "")
    with open(p, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(data)


def mark_status_csv(path: str, row_index_1_based: int, status: str):
    p = os.path.abspath(path)
    with open(p, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        rows = list(reader)
    if "statut" not in fieldnames:
        fieldnames = list(fieldnames) + ["statut"]
        for r in rows:
            r["statut"] = r.get("statut", "")
    target = row_index_1_based - 2
    if 0 <= target < len(rows):
        rows[target]["statut"] = status
    with open(p, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def open_app(driver):
    try:
        driver.terminate_app(APP_PACKAGE)
    except Exception:
        pass
    driver.activate_app(APP_PACKAGE)
    time.sleep(0.8)
    try:
        driver.start_activity(APP_PACKAGE, APP_ACTIVITY)
    except Exception:
        pass
    time.sleep(1.0)


def login_partner(driver, wait: WebDriverWait, email: str, password: str):
    try:
        wait_click(
            driver,
            wait,
            [
                (AppiumBy.ACCESSIBILITY_ID, "Connexion partenaire"),
                (By.XPATH, "//*[@text='Connexion partenaire']"),
            ],
            "Connexion partenaire",
        )
        time.sleep(0.5)
    except Exception:
        pass

    # Onglet E-mail: certaines builds exposent différemment le libellé.
    email_tab_clicked = False
    for by, value in [
        (AppiumBy.ACCESSIBILITY_ID, "E-mail"),
        (By.XPATH, "//*[@text='E-mail']"),
        (By.XPATH, "//*[contains(@text,'E-mail')]"),
        (By.XPATH, "//*[contains(@text,'mail')]"),
        (By.XPATH, "//*[contains(@content-desc,'mail')]"),
        (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().textContains("mail")'),
    ]:
        try:
            el = WebDriverWait(driver, 3, poll_frequency=0.2).until(
                EC.element_to_be_clickable((by, value))
            )
            el.click()
            print(f"✅ Click: Onglet E-mail [{by}={value}]")
            email_tab_clicked = True
            break
        except Exception:
            continue

    if not email_tab_clicked:
        print("⚠️ Onglet E-mail non trouvé, tentative de saisie directe des champs visibles.")

    edits = [e for e in driver.find_elements(By.CLASS_NAME, "android.widget.EditText") if e.is_displayed()]
    if len(edits) < 2:
        raise RuntimeError("Champs login introuvables.")
    edits[0].click()
    edits[0].clear()
    edits[0].send_keys(email)
    edits[1].click()
    edits[1].clear()
    edits[1].send_keys(password)

    wait_click(
        driver,
        wait,
        [
            (AppiumBy.ACCESSIBILITY_ID, "Se connecter"),
            (By.XPATH, "//*[@text='Se connecter']"),
        ],
        "Se connecter",
    )
    time.sleep(2)


def go_to_transfer_modal(driver, wait: WebDriverWait):
    wait_click(
        driver,
        wait,
        [
            (AppiumBy.ACCESSIBILITY_ID, "Comptes"),
            (By.XPATH, "//*[@text='Comptes']"),
        ],
        "Onglet Comptes",
    )
    wait_click(
        driver,
        wait,
        [
            (AppiumBy.ACCESSIBILITY_ID, "Portefeuille"),
            (By.XPATH, "//*[@text='Portefeuille']"),
        ],
        "Portefeuille",
    )
    wait_click(
        driver,
        wait,
        [
            (AppiumBy.ACCESSIBILITY_ID, "Transférer"),
            (By.XPATH, "//*[@text='Transférer']"),
        ],
        "Bouton Transférer (ouvrir modal)",
    )
    time.sleep(0.7)

    # signal modal (tolérant)
    modal_locators = [
        (By.XPATH, "//*[@text='Annuler']"),
        (By.XPATH, "//*[contains(@text,'Sélectionnez le type de destinataire')]"),
        (By.XPATH, "//*[@text='Transférer']"),
        (By.XPATH, "//*[@text='Utilisateur']"),
        (By.XPATH, "//*[@text='Conducteur']"),
    ]
    for by, value in modal_locators:
        try:
            WebDriverWait(driver, 3, poll_frequency=0.2).until(
                EC.visibility_of_element_located((by, value))
            )
            print(f"✅ Modal détectée [{by}={value}]")
            return
        except Exception:
            continue

    # fallback: 2 champs EditText visibles dans le bas-sheet
    edits = [e for e in driver.find_elements(By.CLASS_NAME, "android.widget.EditText") if e.is_displayed()]
    if len(edits) >= 2:
        print("✅ Modal détectée [fallback EditText visibles]")
        return

    # fallback texte page source
    src = driver.page_source or ""
    if "Sélectionnez le type de destinataire" in src or "Annuler" in src:
        print("✅ Modal détectée [fallback page_source]")
        return

    raise TimeoutException("Modal Transférer non détectée après clic.")


def _visible_edittexts(driver):
    return [e for e in driver.find_elements(By.CLASS_NAME, "android.widget.EditText") if e.is_displayed()]


def fill_transfer_fields(driver, phone: str, amount: str):
    """
    Remplit montant (1er EditText) puis numéro (2e). Re-interroge le DOM à chaque étape :
    après transfert / fermeture modale, les références UiAutomator deviennent souvent « stale ».
    """

    def fill_slot(index: int, value: str, label: str):
        last_err: Exception | None = None
        for attempt in range(8):
            try:
                edits = _visible_edittexts(driver)
                if len(edits) <= index:
                    last_err = RuntimeError(f"EditText insuffisants ({len(edits)} visibles, besoin index {index})")
                    time.sleep(0.35)
                    continue
                el = edits[index]
                el.click()
                time.sleep(0.12)
                try:
                    el.clear()
                except StaleElementReferenceException:
                    continue
                el.send_keys(str(value))
                return
            except StaleElementReferenceException:
                last_err = StaleElementReferenceException("stale")
                time.sleep(0.28)
            except Exception as e:
                last_err = e
                time.sleep(0.28)
        raise RuntimeError(f"Impossible de remplir {label} (UI instable): {last_err}")

    time.sleep(0.25)
    fill_slot(0, amount, "montant")
    time.sleep(0.2)
    fill_slot(1, phone, "numero")
    print(f"✅ Prérempli -> montant={amount} | numero={phone}")


def wait_user_then_prepare_next(driver, wait: WebDriverWait, row: Dict[str, str]) -> bool:
    input("➡️ Fais le choix destinataire + Transférer + validation popup sur le téléphone, puis Entrée ici...")
    ok = input("✅ Le transfert est-il bien passé ? (o/n): ").strip().lower() in {"o", "ok", "oui", "y", "yes"}
    if ok and TRANSFER_FILE.lower().endswith(".csv"):
        mark_status_csv(TRANSFER_FILE, int(row["row_index"]), "OK")
        print(f"📝 Statut mis à jour: {row['numero']} -> OK")
    # Ré-ouvrir modal pour la ligne suivante si fermée
    if not driver.find_elements(By.XPATH, "//*[@text='Annuler']"):
        try:
            wait_click(
                driver,
                WebDriverWait(driver, 5, poll_frequency=0.2),
                [
                    (AppiumBy.ACCESSIBILITY_ID, "Transférer"),
                    (By.XPATH, "//*[@text='Transférer']"),
                ],
                "Réouvrir modal Transférer",
            )
            wait_visible(
                driver,
                WebDriverWait(driver, 5, poll_frequency=0.2),
                [
                    (By.XPATH, "//*[@text='Annuler']"),
                    (By.XPATH, "//*[contains(@text,'Sélectionnez le type de destinataire')]"),
                ],
                "Modal Transférer (suivant)",
            )
        except Exception:
            pass
    return ok


def main():
    ensure_csv_status_column(TRANSFER_FILE)
    rows = load_rows(TRANSFER_FILE)
    pending_rows = [r for r in rows if str(r.get("statut", "")).strip().upper() != "OK"]
    print(f"📄 {len(rows)} ligne(s) chargée(s) depuis {TRANSFER_FILE} | à traiter: {len(pending_rows)}")

    udid = resolve_android_udid(os.getenv("ANDROID_UDID", "").strip())

    opts = UiAutomator2Options()
    opts.platform_name = "Android"
    opts.automation_name = "UiAutomator2"
    if udid:
        opts.udid = udid
    opts.app_package = APP_PACKAGE
    opts.app_activity = APP_ACTIVITY
    opts.no_reset = True
    opts.auto_launch = False
    opts.new_command_timeout = 240
    # Certains téléphones refusent `settings delete global hidden_api_*` (WRITE_SECURE_SETTINGS) : sans ça la session échoue au démarrage.
    opts.ignore_hidden_api_policy_error = True
    # Évite `adb install -g` sur l'APK io.appium.settings (certains OEM refusent INSTALL_GRANT_RUNTIME_PERMISSIONS pour le shell).
    opts.skip_device_initialization = True

    driver = webdriver.Remote(APPIUM_URL, options=opts)
    wait = WebDriverWait(driver, 10, poll_frequency=0.2)

    try:
        print("🚀 Démarrage app + login...")
        open_app(driver)
        login_partner(driver, wait, PARTNER_EMAIL, PARTNER_PASSWORD)
        go_to_transfer_modal(driver, wait)

        for i, row in enumerate(pending_rows, 1):
            phone = row["numero"]
            amount = row["montant"]
            print(f"\n[{i}/{len(pending_rows)}] Préparation transfert -> {phone} / {amount}")
            fill_transfer_fields(driver, phone, amount)
            wait_user_then_prepare_next(driver, wait, row)

        print("\n✅ Traitement terminé: toutes les lignes ont été préremplies.")
    finally:
        driver.quit()
        print("🧹 Session Appium fermée.")


if __name__ == "__main__":
    main()

