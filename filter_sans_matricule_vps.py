#!/usr/bin/env python3
"""
filter_sans_matricule.py
------------------------
Lit all_partners_enriched.json (sur le VPS ou en local),
filtre les partenaires Partenaire1 → Partenaire-120,
garde uniquement les conducteurs SANS matricule réel (N/A, vide ou absent),
et exporte le résultat en Excel.
"""

import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl manquant. Installation : pip install openpyxl")
    sys.exit(1)

# ── Chemins ────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
INPUT_JSON = BASE_DIR / "output" / "organized_by_partner" / "all_partners_enriched.json"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_XLS = OUTPUT_DIR / "conducteurs_sans_matricule.xlsx"

# ── Helpers ────────────────────────────────────────────────────────────────────
NO_PLATE_VALUES = {"n/a", "", "null", "none", "-", "—"}

def is_missing_matricule(vehicle) -> bool:
    """Retourne True si le conducteur n'a pas de vrai matricule."""
    if not vehicle:
        return True
    m = str(vehicle.get("matricule", "")).strip().lower()
    return m in NO_PLATE_VALUES

def partner_in_scope(nom: str) -> bool:
    """
    Garde seulement les partenaires dans la plage Partenaire1 … Partenaire-120.
    Noms attendus (insensible à la casse) :
      - Partenaire1, Partenaire2, …, Partenaire9
      - partenaire10, partenaire11, …  (sans tiret)
      - Partenaire-102, Partenaire-103, … Partenaire-120
      - Partenaires-51, Partenaires-52, … Partenaires-100
      - partenaire-43, partenaire-44, …  (minuscule avec tiret)
    """
    n = nom.strip().lower()
    # Enlève les préfixes variantes pour extraire le numéro
    for prefix in ("partenaires-", "partenaire-", "partenaire"):
        if n.startswith(prefix):
            num_str = n[len(prefix):]
            try:
                num = int(num_str)
                return 1 <= num <= 120
            except ValueError:
                return False
    return False


# ── Lecture JSON ───────────────────────────────────────────────────────────────
print(f"Lecture : {INPUT_JSON}")
if not INPUT_JSON.exists():
    print(f"ERREUR : fichier introuvable → {INPUT_JSON}")
    sys.exit(1)

with open(INPUT_JSON, encoding="utf-8") as f:
    partners = json.load(f)

print(f"  {len(partners)} partenaires trouvés dans le JSON")

# ── Filtrage ───────────────────────────────────────────────────────────────────
rows = []
partners_kept = set()

for partner in partners:
    nom_partenaire = partner.get("nom", "")
    if not partner_in_scope(nom_partenaire):
        continue

    for driver in partner.get("drivers", []):
        vehicle = driver.get("vehicle")
        if is_missing_matricule(vehicle):
            partners_kept.add(nom_partenaire)
            rows.append({
                "Partenaire":      nom_partenaire,
                "Email Partenaire": partner.get("email", ""),
                "Tel Partenaire":  partner.get("telephone", ""),
                "Conducteur":      driver.get("nom", ""),
                "Tel Conducteur":  driver.get("telephone", ""),
                "Emplacement":     driver.get("emplacement", ""),
                "Type Transport":  driver.get("type_transport", ""),
                "Statut Véhicule": driver.get("type_vehicule", ""),
                "Véhicule Type":   vehicle.get("type", "N/A") if vehicle else "N/A",
                "Marque":          vehicle.get("marque", "N/A") if vehicle else "N/A",
                "Modèle":          vehicle.get("modele", "N/A") if vehicle else "N/A",
                "Matricule":       vehicle.get("matricule", "N/A") if vehicle else "ABSENT",
            })

print(f"  {len(rows)} conducteurs sans matricule dans {len(partners_kept)} partenaires")

# ── Export Excel ───────────────────────────────────────────────────────────────
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sans Matricule"

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")
CENTER       = Alignment(horizontal="center", vertical="center")

headers = list(rows[0].keys()) if rows else []
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font       = HEADER_FONT
    cell.fill       = HEADER_FILL
    cell.alignment  = CENTER

for row_idx, row in enumerate(rows, 2):
    fill = ALT_FILL if row_idx % 2 == 0 else None
    for col_idx, key in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
        cell.alignment = Alignment(vertical="center")
        if fill:
            cell.fill = fill

# Largeurs auto
col_widths = {h: len(h) for h in headers}
for row in rows:
    for k, v in row.items():
        col_widths[k] = max(col_widths[k], len(str(v)))

for col_idx, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = min(col_widths[h] + 2, 45)

ws.freeze_panes = "A2"

# Onglet résumé par partenaire
ws2 = wb.create_sheet("Résumé par Partenaire")
ws2.append(["Partenaire", "Nb conducteurs sans matricule"])
ws2["A1"].font = HEADER_FONT
ws2["A1"].fill = HEADER_FILL
ws2["B1"].font = HEADER_FONT
ws2["B1"].fill = HEADER_FILL

summary: dict[str, int] = {}
for row in rows:
    summary[row["Partenaire"]] = summary.get(row["Partenaire"], 0) + 1

for r_idx, (pname, count) in enumerate(sorted(summary.items()), 2):
    ws2.cell(row=r_idx, column=1, value=pname)
    ws2.cell(row=r_idx, column=2, value=count)

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 30

wb.save(OUTPUT_XLS)
print(f"\n✅ Fichier Excel généré : {OUTPUT_XLS}")
print(f"   Lignes : {len(rows)} | Partenaires : {len(partners_kept)}")
